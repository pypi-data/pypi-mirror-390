from typing import Union, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


class MappingParser:
    @staticmethod
    def extract_dictionary_from_worksheet(
        worksheet: Worksheet,
        value_column_names: Union[Dict[str, str], str],
        key_column: str = "A",
        min_row: int = 1,
        max_row: int = None,
    ):
        dictionary = {}

        for row in range(min_row, max_row or worksheet.max_row + 1):
            key = worksheet[f"{key_column}{row}"].value
            dictionary[key] = {}

            if isinstance(value_column_names, str):
                value = worksheet[f"{value_column_names}{row}"].value
                dictionary[key] = value
            else:
                for k, v in value_column_names.items():
                    value = worksheet[f"{k}{row}"].value
                    dictionary[key][v] = value
        return dictionary

    @staticmethod
    def extract_list_from_worksheet(
        worksheet: Worksheet,
        value_column_names: Dict[str, str],
        min_row: int = 1,
        max_row: int = None,
    ):
        output_list = []

        for row in range(min_row, max_row or worksheet.max_row + 1):
            obj = {}

            for k, v in value_column_names.items():
                value = worksheet[f"{k}{row}"].value
                obj[v] = value
            output_list.append(obj)
        return output_list

