import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .exceptions import WorkbookError

logger = logging.getLogger(__name__)

def create_workbook(filepath: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Create a new Excel workbook with optional custom sheet name.

    Behavior change: if the default sheet created by openpyxl is empty, remove it and
    create the requested sheet so we don't end up with an empty default sheet plus a new one.
    """
    try:
        wb = Workbook()
        # openpyxl creates a default sheet (usually named "Sheet").
        default_ws = wb.active
        default_name = default_ws.title

        if default_name == sheet_name:
            # requested name equals default -> nothing to do
            pass
        else:
            # determine if the default sheet is empty by scanning all cells
            # Consider a cell empty if its value is None or an empty/whitespace-only string
            is_empty = True
            for row in default_ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and (not (isinstance(cell, str) and str(cell).strip() == "")):
                        is_empty = False
                        break
                if not is_empty:
                    break

            if is_empty:
                # remove the empty default sheet and create the requested one
                wb.remove(default_ws)
                wb.create_sheet(sheet_name)
            else:
                # default has content; keep it and create an additional sheet with requested name
                wb.create_sheet(sheet_name)

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return {
            "message": f"Created workbook: {filepath}",
            "active_sheet": sheet_name,
            "workbook": wb
        }
    except Exception as e:
        logger.error(f"Failed to create workbook: {e}")
        raise WorkbookError(f"Failed to create workbook: {e!s}")

def get_or_create_workbook(filepath: str) -> Workbook:
    """Get existing workbook or create new one if it doesn't exist"""
    try:
        return load_workbook(filepath)
    except FileNotFoundError:
        return create_workbook(filepath)["workbook"]

def create_sheet(filepath: str, sheet_name: str) -> dict:
    """Create a new worksheet in the workbook if it doesn't exist.

    When creating a new sheet, if the default sheet (e.g., "Sheet1" or "Sheet") exists and is empty,
    remove it first to avoid leaving an unused default sheet.
    """
    try:
        wb = load_workbook(filepath)

        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            raise WorkbookError(f"Sheet {sheet_name} already exists")

        # If a default sheet exists and is empty, remove it before creating the requested sheet
        for default_name in ("Sheet1", "Sheet"):
            if default_name in wb.sheetnames and default_name != sheet_name:
                default_ws = wb[default_name]
                is_empty = True
                for row in default_ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None and (not (isinstance(cell, str) and str(cell).strip() == "")):
                            is_empty = False
                            break
                    if not is_empty:
                        break
                if is_empty:
                    wb.remove(default_ws)
                    break

        # Create new sheet
        wb.create_sheet(sheet_name)
        wb.save(filepath)
        wb.close()
        return {"message": f"Sheet {sheet_name} created successfully"}
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to create sheet: {e}")
        raise WorkbookError(str(e))

def get_workbook_info(filepath: str, include_ranges: bool = False) -> dict[str, Any]:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        path = Path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")
            
        wb = load_workbook(filepath, read_only=False)
        
        info = {
            "filename": path.name,
            "sheets": wb.sheetnames,
            "size": path.stat().st_size,
            "modified": path.stat().st_mtime
        }
        
        if include_ranges:
            # Add used ranges for each sheet
            ranges = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row > 0 and ws.max_column > 0:
                    ranges[sheet_name] = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            info["used_ranges"] = ranges
            
        wb.close()
        return info
        
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get workbook info: {e}")
        raise WorkbookError(str(e))
