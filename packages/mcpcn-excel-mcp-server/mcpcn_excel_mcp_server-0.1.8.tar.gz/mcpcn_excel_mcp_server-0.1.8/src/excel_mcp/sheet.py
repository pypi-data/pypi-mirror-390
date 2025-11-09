import logging
from typing import Any, Dict, Optional
from copy import copy

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Border, PatternFill, Side

from .cell_utils import parse_cell_range
from .exceptions import SheetError, ValidationError

logger = logging.getLogger(__name__)

def copy_sheet(filepath: str, source_sheet: str, target_sheet: str) -> Dict[str, Any]:
    """Copy a worksheet within the same workbook."""
    try:
        wb = load_workbook(filepath)
        if source_sheet not in wb.sheetnames:
            raise SheetError(f"Source sheet '{source_sheet}' not found")
            
        if target_sheet in wb.sheetnames:
            raise SheetError(f"Target sheet '{target_sheet}' already exists")
            
        source = wb[source_sheet]
        target = wb.copy_worksheet(source)
        target.title = target_sheet
        
        wb.save(filepath)
        return {"message": f"Sheet '{source_sheet}' copied to '{target_sheet}'"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to copy sheet: {e}")
        raise SheetError(str(e))

def delete_sheet(filepath: str, sheet_name: str) -> Dict[str, Any]:
    """Delete a worksheet from the workbook."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        if len(wb.sheetnames) == 1:
            raise SheetError("Cannot delete the only sheet in workbook")
            
        del wb[sheet_name]
        wb.save(filepath)
        return {"message": f"Sheet '{sheet_name}' deleted"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete sheet: {e}")
        raise SheetError(str(e))

def rename_sheet(filepath: str, old_name: str, new_name: str) -> Dict[str, Any]:
    """Rename a worksheet."""
    try:
        wb = load_workbook(filepath)
        if old_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{old_name}' not found")
            
        if new_name in wb.sheetnames:
            raise SheetError(f"Sheet '{new_name}' already exists")
            
        sheet = wb[old_name]
        sheet.title = new_name
        wb.save(filepath)
        return {"message": f"Sheet renamed from '{old_name}' to '{new_name}'"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to rename sheet: {e}")
        raise SheetError(str(e))

def format_range_string(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    """Format range string from row and column indices."""
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

def copy_range(
    source_ws: Worksheet,
    target_ws: Worksheet,
    source_range: str,
    target_start: Optional[str] = None,
) -> None:
    """Copy range from source worksheet to target worksheet."""
    # Parse source range
    if ':' in source_range:
        source_start, source_end = source_range.split(':')
    else:
        source_start = source_range
        source_end = None
        
    src_start_row, src_start_col, src_end_row, src_end_col = parse_cell_range(
        source_start, source_end
    )

    if src_end_row is None:
        src_end_row = src_start_row
        src_end_col = src_start_col

    if target_start is None:
        target_start = source_start

    tgt_start_row, tgt_start_col, _, _ = parse_cell_range(target_start)

    for i, row in enumerate(range(src_start_row, src_end_row + 1)):
        for j, col in enumerate(range(src_start_col, src_end_col + 1)):
            source_cell = source_ws.cell(row=row, column=col)
            target_cell = target_ws.cell(row=tgt_start_row + i, column=tgt_start_col + j)

            target_cell.value = source_cell.value

            try:
                # Copy font
                font_kwargs = {}
                if hasattr(source_cell.font, 'name'):
                    font_kwargs['name'] = source_cell.font.name
                if hasattr(source_cell.font, 'size'):
                    font_kwargs['size'] = source_cell.font.size
                if hasattr(source_cell.font, 'bold'):
                    font_kwargs['bold'] = source_cell.font.bold
                if hasattr(source_cell.font, 'italic'):
                    font_kwargs['italic'] = source_cell.font.italic
                if hasattr(source_cell.font, 'color'):
                    font_color = None
                    if source_cell.font.color:
                        font_color = source_cell.font.color.rgb
                    font_kwargs['color'] = font_color
                target_cell.font = Font(**font_kwargs)

                # Copy border
                new_border = Border()
                for side in ['left', 'right', 'top', 'bottom']:
                    source_side = getattr(source_cell.border, side)
                    if source_side and source_side.style:
                        side_color = source_side.color.rgb if source_side.color else None
                        setattr(new_border, side, Side(
                            style=source_side.style,
                            color=side_color
                        ))
                target_cell.border = new_border

                # Copy fill
                if hasattr(source_cell, 'fill'):
                    fill_kwargs = {'patternType': source_cell.fill.patternType}
                    if hasattr(source_cell.fill, 'fgColor') and source_cell.fill.fgColor:
                        fg_color = None
                        if hasattr(source_cell.fill.fgColor, 'rgb'):
                            fg_color = source_cell.fill.fgColor.rgb
                        fill_kwargs['fgColor'] = fg_color
                    if hasattr(source_cell.fill, 'bgColor') and source_cell.fill.bgColor:
                        bg_color = None
                        if hasattr(source_cell.fill.bgColor, 'rgb'):
                            bg_color = source_cell.fill.bgColor.rgb
                        fill_kwargs['bgColor'] = bg_color
                    target_cell.fill = PatternFill(**fill_kwargs)

                # Copy number format and alignment
                if source_cell.number_format:
                    target_cell.number_format = source_cell.number_format
                if source_cell.alignment:
                    target_cell.alignment = source_cell.alignment

            except Exception:
                continue

def delete_range(worksheet: Worksheet, start_cell: str, end_cell: Optional[str] = None) -> None:
    """Delete contents and formatting of a range."""
    start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)

    if end_row is None:
        end_row = start_row
        end_col = start_col

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = None
            cell.font = Font()
            cell.border = Border()
            cell.fill = PatternFill()
            cell.number_format = "General"
            cell.alignment = None

def merge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> Dict[str, Any]:
    """Merge a range of cells."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)

        if end_row is None or end_col is None:
            raise SheetError("Both start and end cells must be specified for merging")

        range_string = format_range_string(start_row, start_col, end_row, end_col)
        worksheet = wb[sheet_name]
        worksheet.merge_cells(range_string)
        wb.save(filepath)
        return {"message": f"Range '{range_string}' merged in sheet '{sheet_name}'"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to merge range: {e}")
        raise SheetError(str(e))

def unmerge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> Dict[str, Any]:
    """Unmerge a range of cells."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        
        if end_row is None or end_col is None:
            raise SheetError("Both start and end cells must be specified for unmerging")

        range_string = format_range_string(start_row, start_col, end_row, end_col)
        
        # Check if range is actually merged
        merged_ranges = worksheet.merged_cells.ranges
        target_range = range_string.upper()
        
        if not any(str(merged_range).upper() == target_range for merged_range in merged_ranges):
            raise SheetError(f"Range '{range_string}' is not merged")
            
        worksheet.unmerge_cells(range_string)
        wb.save(filepath)
        return {"message": f"Range '{range_string}' unmerged successfully"}
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to unmerge range: {e}")
        raise SheetError(str(e))

def get_merged_ranges(filepath: str, sheet_name: str) -> list[str]:
    """Get merged cells in a worksheet."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
        worksheet = wb[sheet_name]
        return [str(merged_range) for merged_range in worksheet.merged_cells.ranges]
    except SheetError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get merged cells: {e}")
        raise SheetError(str(e))

def copy_range_operation(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None
) -> Dict:
    """Copy a range of cells to another location."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found")
            raise ValidationError(f"Sheet '{sheet_name}' not found")

        source_ws = wb[sheet_name]
        target_ws = wb[target_sheet] if target_sheet else source_ws

        # Parse source range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(source_start, source_end)
        except ValueError as e:
            logger.error(f"Invalid source range: {e}")
            raise ValidationError(f"Invalid source range: {str(e)}")

        # Parse target starting point
        try:
            target_row = int(''.join(filter(str.isdigit, target_start)))
            target_col = column_index_from_string(''.join(filter(str.isalpha, target_start)))
        except ValueError as e:
            logger.error(f"Invalid target cell: {e}")
            raise ValidationError(f"Invalid target cell: {str(e)}")

        # Copy the range
        row_offset = target_row - start_row
        col_offset = target_col - start_col

        for i in range(start_row, end_row + 1):
            for j in range(start_col, end_col + 1):
                source_cell = source_ws.cell(row=i, column=j)
                target_cell = target_ws.cell(row=i + row_offset, column=j + col_offset)
                target_cell.value = source_cell.value
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)

        wb.save(filepath)
        return {"message": f"Range copied successfully"}

    except (ValidationError, SheetError):
        raise
    except Exception as e:
        logger.error(f"Failed to copy range: {e}")
        raise SheetError(f"Failed to copy range: {str(e)}")

def delete_range_operation(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    shift_direction: str = "up"
) -> Dict[str, Any]:
    """Delete a range of cells and shift remaining cells."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        # Validate range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
            if end_row and end_row > worksheet.max_row:
                raise SheetError(f"End row {end_row} out of bounds (1-{worksheet.max_row})")
            if end_col and end_col > worksheet.max_column:
                raise SheetError(f"End column {end_col} out of bounds (1-{worksheet.max_column})")
        except ValueError as e:
            raise SheetError(f"Invalid range: {str(e)}")
            
        # Validate shift direction
        if shift_direction not in ["up", "left"]:
            raise ValidationError(f"Invalid shift direction: {shift_direction}. Must be 'up' or 'left'")
            
        range_string = format_range_string(
            start_row, start_col,
            end_row or start_row,
            end_col or start_col
        )
        
        # Delete range contents
        delete_range(worksheet, start_cell, end_cell)
        
        # Shift cells if needed
        if shift_direction == "up":
            worksheet.delete_rows(start_row, (end_row or start_row) - start_row + 1)
        elif shift_direction == "left":
            worksheet.delete_cols(start_col, (end_col or start_col) - start_col + 1)
            
        wb.save(filepath)
        
        return {"message": f"Range {range_string} deleted successfully"}
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete range: {e}")
        raise SheetError(str(e))

def insert_row(filepath: str, sheet_name: str, start_row: int, count: int = 1) -> Dict[str, Any]:
    """Insert one or more rows starting at the specified row."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        # Validate parameters
        if start_row < 1:
            raise ValidationError("Start row must be 1 or greater")
        if count < 1:
            raise ValidationError("Count must be 1 or greater")
            
        worksheet.insert_rows(start_row, count)
        wb.save(filepath)
        
        return {"message": f"Inserted {count} row(s) starting at row {start_row} in sheet '{sheet_name}'"}
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to insert rows: {e}")
        raise SheetError(str(e))

def insert_cols(filepath: str, sheet_name: str, start_col: int, count: int = 1) -> Dict[str, Any]:
    """Insert one or more columns starting at the specified column."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        # Validate parameters
        if start_col < 1:
            raise ValidationError("Start column must be 1 or greater")
        if count < 1:
            raise ValidationError("Count must be 1 or greater")
            
        worksheet.insert_cols(start_col, count)
        wb.save(filepath)
        
        return {"message": f"Inserted {count} column(s) starting at column {start_col} in sheet '{sheet_name}'"}
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to insert columns: {e}")
        raise SheetError(str(e))

def delete_rows(filepath: str, sheet_name: str, start_row: int, count: int = 1) -> Dict[str, Any]:
    """Delete one or more rows starting at the specified row."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        # Validate parameters
        if start_row < 1:
            raise ValidationError("Start row must be 1 or greater")
        if count < 1:
            raise ValidationError("Count must be 1 or greater")
        if start_row > worksheet.max_row:
            raise ValidationError(f"Start row {start_row} exceeds worksheet bounds (max row: {worksheet.max_row})")
            
        worksheet.delete_rows(start_row, count)
        wb.save(filepath)
        
        return {"message": f"Deleted {count} row(s) starting at row {start_row} in sheet '{sheet_name}'"}
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete rows: {e}")
        raise SheetError(str(e))

def delete_cols(filepath: str, sheet_name: str, start_col: int, count: int = 1) -> Dict[str, Any]:
    """Delete one or more columns starting at the specified column."""
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{sheet_name}' not found")
            
        worksheet = wb[sheet_name]
        
        # Validate parameters
        if start_col < 1:
            raise ValidationError("Start column must be 1 or greater")
        if count < 1:
            raise ValidationError("Count must be 1 or greater")
        if start_col > worksheet.max_column:
            raise ValidationError(f"Start column {start_col} exceeds worksheet bounds (max column: {worksheet.max_column})")
            
        worksheet.delete_cols(start_col, count)
        wb.save(filepath)
        
        return {"message": f"Deleted {count} column(s) starting at column {start_col} in sheet '{sheet_name}'"}
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to delete columns: {e}")
        raise SheetError(str(e))

def sort_rows_by_column(
    filepath: str,
    column_index: int,
    sheet_name: Optional[str] = None,
    ascending: bool = True,
    has_header: bool = True,
    start_row: Optional[int] = None,
    end_row: Optional[int] = None,
) -> Dict[str, Any]:
    """Sort rows in a worksheet by a specific column.

    If ``sheet_name`` is not provided, the first sheet in the workbook will be used.
    Column index is required. Sorting order defaults to ascending and the
    first row in the sort range can be treated as header (not included in sort)
    when `has_header` is True.

    Args:
        filepath: Path to the Excel workbook.
        column_index: 1-based target column index to sort by (required).
        sheet_name: Name of the sheet to sort. Defaults to the first sheet.
        ascending: Sort order; True for ascending, False for descending.
        has_header: Whether the first row in the sort range is a header to keep on top.
        start_row: Optional start row for sort; defaults to sheet's ``min_row``.
        end_row: Optional end row for sort; defaults to sheet's ``max_row``.

    Returns:
        Dict with a human-readable message of the operation performed.
    """
    try:
        wb = load_workbook(filepath)

        # Resolve sheet
        target_sheet_name = sheet_name or (wb.sheetnames[0] if wb.sheetnames else None)
        if not target_sheet_name:
            raise SheetError("Workbook has no sheets to sort")
        if target_sheet_name not in wb.sheetnames:
            raise SheetError(f"Sheet '{target_sheet_name}' not found")
        ws = wb[target_sheet_name]

        # Determine boundaries (cover from first column to max used column)
        min_col = 1
        max_col = ws.max_column
        if max_col < column_index:
            raise ValidationError(
                f"Target column {column_index} exceeds sheet bounds (max column: {max_col})"
            )

        sr = start_row or ws.min_row
        er = end_row or ws.max_row

        if sr < 1 or er < sr:
            raise ValidationError("Invalid row range for sorting")
        if er > ws.max_row:
            er = ws.max_row

        # Extract rows data (values + styles) within [sr, er] and [min_col, max_col]
        def extract_row(r: int):
            row_cells = []
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                row_cells.append({
                    "value": cell.value,
                    "style": copy(cell._style) if cell.has_style else None,
                })
            return row_cells

        rows = [extract_row(r) for r in range(sr, er + 1)]
        if not rows:
            wb.close()
            return {"message": f"No rows to sort in sheet '{target_sheet_name}'"}

        # Determine target index within row array
        target_idx = column_index - min_col
        if target_idx < 0 or target_idx >= (max_col - min_col + 1):
            raise ValidationError("Computed target column index out of bounds")

        # Optionally separate header
        header_row = None
        data_rows = rows
        if has_header and len(rows) > 0:
            header_row = rows[0]
            data_rows = rows[1:]

        # Normalize values for sorting
        def norm_value(val: Any):
            # Ensure keys are always comparable across rows
            # Return tuple: (none_flag, type_order, normalized_value)
            # none_flag: 1 for None (sorted last), 0 otherwise
            # type_order: 0 for numeric-like, 1 for string-like
            # normalized_value: float for numeric, lowercase str for string
            if val is None:
                return (1, 1, "")
            # Numeric-like (includes ints, floats, and numeric strings)
            try:
                num = float(val)
                return (0, 0, num)
            except (TypeError, ValueError):
                pass
            # Fallback to strings
            return (0, 1, str(val).lower())

        # Sort data rows
        data_rows.sort(
            key=lambda row: norm_value(row[target_idx]["value"]),
            reverse=not ascending,
        )

        # Recombine with header if present
        sorted_rows = ([header_row] + data_rows) if header_row is not None else data_rows

        # Write back to sheet
        for i, row in enumerate(sorted_rows):
            dest_r = sr + i
            for j, cell_info in enumerate(row):
                dest_c = min_col + j
                tgt = ws.cell(row=dest_r, column=dest_c)
                tgt.value = cell_info["value"]
                if cell_info["style"] is not None:
                    try:
                        tgt._style = copy(cell_info["style"])  # preserve formatting when possible
                    except Exception:
                        # If style copy fails, continue without raising
                        pass

        wb.save(filepath)
        wb.close()

        col_letter = get_column_letter(column_index)
        range_str = f"{get_column_letter(min_col)}{sr}:{get_column_letter(max_col)}{er}"
        order_str = "ascending" if ascending else "descending"
        hdr_str = " (header preserved)" if has_header else ""
        return {"message": f"Sorted rows {range_str} by column {col_letter} in sheet '{target_sheet_name}' ({order_str}){hdr_str}"}
    except (ValidationError, SheetError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to sort rows: {e}")
        raise SheetError(str(e))
