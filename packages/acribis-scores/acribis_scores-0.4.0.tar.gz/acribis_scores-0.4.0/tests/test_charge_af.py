import unittest
import platform
import os

from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By

from parameter_generator import generate_charge_af_parameters
from acribis_scores.charge_af import calc_charge_af_score, Parameters

if platform.system() == 'Windows':
    import openpyxl
    import win32com.client as win32


class TestCHARGEAFCalculator(unittest.TestCase):
    def setUp(self):
        self.driver = webdriver.Firefox()
        self.driver.maximize_window()

    def tearDown(self):
        self.driver.quit()

    def test_charge_af(self):
        for i in range(10):
            parameters = generate_charge_af_parameters()

            r_score = self.__get_r_score(parameters)
            python_score = calc_charge_af_score(parameters)
            if platform.system() == 'Windows':
                excel_score, equal = self.__get_excel_score(parameters)
                if equal:
                    self.assertAlmostEqual(python_score, excel_score, msg='CHARGE-AF', delta=0.2)
                else:
                    self.assertGreaterEqual(round(python_score, 2), round(excel_score, 2))
            self.assertEqual(round(python_score, 2), r_score, 'CHARGE-AF')

    def __get_r_score(self, parameters: Parameters) -> float:
        self.driver.get("http://localhost/")
        self.driver.find_element(By.CSS_SELECTOR, "a[data-value='CHARGE-AF Score']").click()
        mapping: dict[str, str] = {
            'Race (white)': 'race_white',
            'Smoking (current)': 'smoking_current',
            'Antihypertensive Medication Use (Yes)': 'antihypertensive_use',
            'Diabetes (Yes)': 'diabetes_yes',
            'Heart failure (Yes)': 'heart_failure_yes',
            'Myocardial infarction (Yes)': 'mi_yes'
        }
        element = self.driver.find_element(By.ID, "charge_af_age")
        element.click()
        element.send_keys(Keys.CONTROL, "a")
        element.send_keys(str(parameters['Age']))
        element = self.driver.find_element(By.ID, "height")
        element.click()
        element.send_keys(Keys.CONTROL, "a")
        element.send_keys(str(parameters['Height']))
        element = self.driver.find_element(By.ID, "weight")
        element.click()
        element.send_keys(Keys.CONTROL, "a")
        element.send_keys(str(parameters['Weight']))
        element = self.driver.find_element(By.ID, "systolic_bp")
        element.click()
        element.send_keys(Keys.CONTROL, "a")
        element.send_keys(str(parameters['Systolic Blood Pressure']))
        element = self.driver.find_element(By.ID, "diastolic_bp")
        element.click()
        element.send_keys(Keys.CONTROL, "a")
        element.send_keys(str(parameters['Diastolic Blood Pressure']))
        for key, value in parameters.items():
            if key not in mapping:
                continue
            if value != (self.driver.find_element(By.ID, mapping[key]).get_attribute('checked') is not None):
                self.driver.find_element(By.ID, mapping[key]).click()
        self.driver.find_element(By.ID, "calculate_charge_af").click()
        text = self.driver.find_element(By.ID, "score_output_charge_af").text
        return float(text.split(': ')[1])

    @staticmethod
    def __get_excel_score(parameters: Parameters) -> tuple[float, bool]:
        directory = os.path.dirname(os.path.abspath(__file__))
        excel_original = os.path.join(directory, 'resources/CHARGE_AF_Calculator.xlsx')
        excel_tmp = os.path.join(directory, 'resources/CHARGE_AF_Calculator_TMP.xlsx')
        wb = openpyxl.load_workbook(excel_original)
        ws = wb['Sheet1']

        ws['C6'] = parameters['Age']
        ws['C7'] = 'yes' if parameters['Race (white)'] else 'no'
        ws['C8'] = parameters['Height']
        ws['C9'] = parameters['Weight']
        ws['C10'] = 'yes' if parameters['Smoking (current)'] else 'no'
        ws['C11'] = parameters['Systolic Blood Pressure']
        ws['C12'] = parameters['Diastolic Blood Pressure']
        ws['C13'] = 'yes' if parameters['Antihypertensive Medication Use (Yes)'] else 'no'
        ws['C14'] = 'yes' if parameters['Diabetes (Yes)'] else 'no'
        ws['C15'] = 'yes' if parameters['Heart failure (Yes)'] else 'no'
        ws['C16'] = 'yes' if parameters['Myocardial infarction (Yes)'] else 'no'

        wb.save(excel_tmp)
        wb.close()

        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workbook = excel.Workbooks.Open(excel_tmp)
        workbook.Save()
        workbook.Close()
        excel.Quit()

        score = openpyxl.load_workbook(excel_tmp, data_only=True)['Sheet1']['B20'].value
        if type(score) is str:
            score = score.strip('>%')
            return float(score), False
        return score * 100, True


if __name__ == '__main__':
    unittest.main()
