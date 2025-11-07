from pathlib import Path
from typing import List

from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from .base_writer import BaseWriter
import logging
import pandas as pd



class XLSXWriter(BaseWriter):
    @property
    def fmt(self) -> str:
        return "xlsx"

    @property
    def format(self) -> str:
        return 'excel'

    def __init__(self, project_id: str, pipeline_id_list: List[str], output_dir: str, recursive_extract: bool, run_for_all: bool = False):
        super().__init__(project_id, pipeline_id_list, output_dir, recursive_extract, run_for_all)
        self._configure_styles()
        self._column_widths = {
            'A': 20, 'B': 30, 'C': 30,
            'D': 40, 'E': 30, 'F': 30, 'G': 80, 'H':40, 'I': 30,
            'J': 80, 'K': 80
        }

    def _configure_styles(self):
        self.header_fill = PatternFill(start_color="FFDD99", fill_type="solid")
        self.header_font = Font(bold=True)

    def write_to_format(self) -> Path:
        logging.info(f"DEBUG_WRITER: Starting: Writing to Excel Format; OUTPUT_DIR = {self.get_output_path()}")
        wb = Workbook()

        # Process each pipeline
        pipeline_nm_full_str = ""
        for pipeline_id, datasets in self.pipeline_dataset_map.items():
            pipeline_nm_full_str = pipeline_nm_full_str + "_" + self._get_name(pipeline_id)
            logging.info(f"DEBUG_WRITER: Starting: Writing SHEET for pipeline = {self._get_name(pipeline_id)}")
            df = self._process_pipeline(pipeline_id, datasets)
            if df is None:
                logging.error(f"NO CSV AVAILABLE FOR pipeline_id={pipeline_id} datasets={datasets}")
                continue
            self._create_sheet(wb, df, self._get_name(pipeline_id))
            logging.info(f"DEBUG_WRITER: Ending: Writing SHEET for pipeline = {self._get_name(pipeline_id)}")


        # Create overall project sheet if run_for_all is True
        if self.run_for_all:
            logging.info("DEBUG_WRITER: Starting: Writing OVERALL SHEET")
            df = self.read_csvs()
            if df is None:
                logging.error("NO CSV AVAILABLE IN TEMP FOLDER. FAILED AT RUN_FOR_ALL")
            self._create_sheet(wb, self._process_dataframe(df), "Overall Project", overall = True)
            output_path = self.get_output_path()
            logging.info("DEBUG_WRITER: Ending: Writing OVERALL SHEET")
        else:
            logging.info("DEBUG_WRITER: Skipping: Writing OVERALL SHEET")
            output_path = self.get_output_path(pipeline_nm_full_str.strip("_"))

        wb.remove(wb['Sheet'])
        wb.save(output_path)
        logging.info(f"Excel report generated: {output_path}")
        return output_path

    def _create_sheet(self, wb: Workbook, df: pd.DataFrame, sheet_name: str, overall = False):
        if overall:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb.create_sheet(sheet_name, 0)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            if r_idx == 1:
                for cell in ws[r_idx]:
                    cell.fill = self.header_fill
                    cell.font = self.header_font

        for col, width in self._column_widths.items():
            ws.column_dimensions[col].width = width
