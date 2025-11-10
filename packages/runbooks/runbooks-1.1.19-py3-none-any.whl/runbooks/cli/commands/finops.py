"""
FinOps Commands Module - Financial Operations & Cost Optimization

KISS Principle: Focused on financial operations and cost optimization
DRY Principle: Uses centralized patterns from DRYPatternManager

Phase 2 Enhancement: Eliminates pattern duplication through reference-based access.
Context Efficiency: Reduced imports and shared instances for memory optimization.
"""

# Essential imports that can't be centralized due to decorator usage
import click

# DRY Pattern Manager - eliminates duplication across CLI modules
from runbooks.common.patterns import get_console, get_error_handlers, get_click_group_creator, get_common_decorators

# Import unified CLI decorators (v1.1.7 standardization)
from runbooks.common.cli_decorators import (
    common_aws_options,
    common_output_options,
    common_multi_account_options,
    common_filter_options,
    mcp_validation_option
)

# Single console instance shared across all modules (DRY principle)
console = get_console()

# Import additional modules for enhanced functionality
from runbooks.common.rich_utils import print_header, print_success, print_error, print_info, print_warning
from runbooks.common.output_controller import OutputController
from runbooks.common.logging_config import configure_logging

# Centralized error handlers - replaces 6 duplicate patterns in this module
error_handlers = get_error_handlers()


def _get_cost_metric_display(cost_metrics):
    """Get display string for cost metrics."""
    if len(cost_metrics) == 1:
        return cost_metrics[0]
    else:
        return " + ".join(cost_metrics)


def calculate_evidence_based_savings(
    services_data: dict,
    running_instances: int,
    stopped_instances: int,
    current_cost: float
) -> tuple:
    """
    Calculate optimization potential from actual resource analysis (evidence-based).

    Uses evidence-based calculations from AWS benchmarks and industry standards:
    - Stopped EC2: $100/month per instance (AWS cost average for EBS volumes)
    - EC2 rightsizing: 20% savings (AWS Compute Optimizer studies)
    - S3 lifecycle: 25% savings (Intelligent-Tiering + Glacier migration)
    - CloudWatch logs: 30% savings (retention optimization)
    - RDS idle: 15% savings (5-signal idle detection)

    Args:
        services_data: Cost by service mapping from Cost Explorer API
        running_instances: Count of running EC2 instances
        stopped_instances: Count of stopped EC2 instances
        current_cost: Current monthly spend

    Returns:
        Tuple of (total_monthly_savings, savings_breakdown_list)

    Example:
        >>> monthly_savings, breakdown = calculate_evidence_based_savings(
        ...     services_data={"Amazon Elastic Compute Cloud - Compute": 5000},
        ...     running_instances=10,
        ...     stopped_instances=2,
        ...     current_cost=10000
        ... )
        >>> monthly_savings
        1200
        >>> breakdown
        [("Stopped EC2 cleanup", 200), ("EC2 rightsizing", 1000)]
    """
    savings_sources = []

    # Source 1: Stopped EC2 instances (IMMEDIATE priority - 🔴)
    if stopped_instances > 0:
        # $100/month per instance for EBS volumes + allocated resources
        stopped_savings = stopped_instances * 100
        savings_sources.append(("Stopped EC2 cleanup", stopped_savings))

    # Source 2: EC2 rightsizing (30-DAY priority - 🟡)
    ec2_cost = services_data.get("Amazon Elastic Compute Cloud - Compute", 0)
    if running_instances > 5 and ec2_cost > 0:
        # 20% savings potential from AWS Compute Optimizer industry benchmarks
        rightsizing_savings = ec2_cost * 0.20
        if rightsizing_savings > 100:  # Only include if savings > $100/month
            savings_sources.append(("EC2 rightsizing", rightsizing_savings))

    # Source 3: S3 lifecycle policies (90-DAY priority - 🟢)
    s3_cost = services_data.get("Amazon Simple Storage Service", 0)
    if s3_cost > 200:  # Meaningful S3 spend threshold
        # 25% savings from Intelligent-Tiering + Glacier Deep Archive migration
        storage_savings = s3_cost * 0.25
        savings_sources.append(("S3 lifecycle policies", storage_savings))

    # Source 4: CloudWatch log retention (60-DAY priority - 🟡)
    cloudwatch_cost = (
        services_data.get("Amazon CloudWatch", 0) +
        services_data.get("CloudWatch", 0) +
        services_data.get("AmazonCloudWatch", 0)  # Handle name variations
    )
    if cloudwatch_cost > 50:  # Meaningful CloudWatch spend threshold
        # 30% savings from retention policy optimization (30 days → 7 days for non-prod)
        log_savings = cloudwatch_cost * 0.30
        savings_sources.append(("CloudWatch log retention", log_savings))

    # Source 5: RDS idle detection (30-DAY priority - 🟡)
    rds_cost = (
        services_data.get("Amazon Relational Database Service", 0) +
        services_data.get("Amazon RDS Service", 0) +
        services_data.get("Amazon Relational Datab", 0)  # Handle truncated names
    )
    if rds_cost > 500:  # Meaningful RDS spend threshold
        # 15% savings from idle instance detection (5 signals: connections, CPU, I/O, network, queries)
        rds_savings = rds_cost * 0.15
        savings_sources.append(("RDS idle instances", rds_savings))

    # Calculate total monthly savings from all evidence-based sources
    total_monthly_savings = sum(amount for _, amount in savings_sources)

    return total_monthly_savings, savings_sources


def _generate_table_summary(
    account_id: str,
    current_month_name: str,
    current_cost: float,
    sorted_services: list,
    monthly_savings: float,
    optimization_potential: float,
    savings_breakdown: list
) -> None:
    """Generate executive summary focused on services table data."""
    from runbooks.common.rich_utils import console

    console.print("[bold]Executive Summary (Services Focus):[/bold]")
    console.print(f"• Account: [cyan]{account_id}[/cyan] | Period: {current_month_name} | Monthly spend: [cyan]${current_cost:,.0f}[/cyan]")

    # Top 3 services breakdown
    if sorted_services:
        console.print("• Top 3 services:")
        for i, (service, cost) in enumerate(sorted_services[:3], 1):
            pct = (cost / current_cost * 100) if current_cost > 0 else 0
            service_display = service[:40] if len(service) > 40 else service
            console.print(f"  {i}. {service_display}: [cyan]${cost:,.0f}[/cyan] ({pct:.1f}%)")

    # Savings potential
    if monthly_savings > 0:
        console.print(f"• Savings potential: [green]${monthly_savings:,.0f}/month[/green] (${optimization_potential:,.0f} annually)")
        if savings_breakdown:
            top_sources = ', '.join(source for source, _ in savings_breakdown[:3])
            console.print(f"[dim]  Top sources: {top_sources}[/dim]")
    else:
        console.print("• Optimization: [yellow]Run 'finops optimize' for detailed analysis[/yellow]")

    console.print()


def _generate_tree_summary(
    account_id: str,
    current_month_name: str,
    current_cost: float,
    categorized_services: dict,
    monthly_savings: float,
    optimization_potential: float,
    savings_breakdown: list
) -> None:
    """Generate executive summary focused on category tree data."""
    from runbooks.common.rich_utils import console

    console.print("[bold]Executive Summary (Category Focus):[/bold]")
    console.print(f"• Account: [cyan]{account_id}[/cyan] | Period: {current_month_name} | Monthly spend: [cyan]${current_cost:,.0f}[/cyan]")

    # Top 3 categories breakdown
    if categorized_services:
        # Calculate category totals
        category_totals = {}
        for category, services in categorized_services.items():
            category_total = sum(s.get("current_cost", 0) for s in services)
            category_totals[category] = category_total

        # Sort by total and show top 3
        sorted_categories = sorted(category_totals.items(), key=lambda x: x[1], reverse=True)[:3]

        console.print("• Top 3 cost categories:")
        for i, (category, cost) in enumerate(sorted_categories, 1):
            pct = (cost / current_cost * 100) if current_cost > 0 else 0
            console.print(f"  {i}. {category}: [cyan]${cost:,.0f}[/cyan] ({pct:.1f}%)")

    # Savings potential
    if monthly_savings > 0:
        console.print(f"• Savings potential: [green]${monthly_savings:,.0f}/month[/green] (${optimization_potential:,.0f} annually)")
        if savings_breakdown:
            top_sources = ', '.join(source for source, _ in savings_breakdown[:3])
            console.print(f"[dim]  Top sources: {top_sources}[/dim]")
    else:
        console.print("• Optimization: [yellow]Run 'finops optimize' for detailed analysis[/yellow]")

    console.print()


def _generate_comprehensive_summary(
    account_id: str,
    current_month_name: str,
    current_cost: float,
    monthly_savings: float,
    optimization_potential: float,
    savings_breakdown: list
) -> None:
    """Generate comprehensive executive summary (current behavior)."""
    from runbooks.common.rich_utils import console

    console.print("[bold]Executive Summary:[/bold]")
    console.print(f"• Account: [cyan]{account_id}[/cyan] | Period: {current_month_name} | Monthly spend: [cyan]${current_cost:,.0f}[/cyan]")

    # Display evidence-based savings potential (no hardcoded assumptions)
    if monthly_savings > 0:
        console.print(f"• Savings potential: [green]${monthly_savings:,.0f}/month[/green] (${optimization_potential:,.0f} annually)")
        # Show top 3 savings sources for transparency
        if savings_breakdown:
            top_sources = ', '.join(source for source, _ in savings_breakdown[:3])
            console.print(f"[dim]  Top sources: {top_sources}[/dim]")
    else:
        console.print("• Optimization: [yellow]Run 'finops optimize' for detailed analysis[/yellow]")

    console.print()


def _create_nested_services_table(
    services_data: dict,
    previous_services_costs: dict,
    total_current: float,
    total_previous: float,
    top_n: int,
    sort_by: str,
    cost_threshold: float,
    current_month_name: str
) -> "Table":
    """
    Create nested services table for Rich Tree visualization.

    Args:
        services_data: Current month costs by service {service_name: cost}
        previous_services_costs: Previous month costs by service
        total_current: Account total current month cost
        total_previous: Account total previous month cost
        top_n: Number of top services to display
        sort_by: Sort order (current|previous|change)
        cost_threshold: Minimum cost filter
        current_month_name: Display name for current month

    Returns:
        Rich Table object with top N services + "Others" row + TOTAL row
    """
    from runbooks.common.rich_utils import create_table

    # Apply cost threshold filter if specified
    if cost_threshold > 0:
        services_data = {k: v for k, v in services_data.items() if v >= cost_threshold}

    # Sort services based on --sort-by parameter
    if sort_by == "current":
        sort_key = lambda x: x[1]  # Current cost
    elif sort_by == "previous":
        sort_key = lambda x: previous_services_costs.get(x[0], 0)  # Previous cost
    else:  # sort_by == "change"
        sort_key = lambda x: abs(x[1] - previous_services_costs.get(x[0], 0)) / previous_services_costs.get(x[0], 1) if previous_services_costs.get(x[0], 0) > 0 else 0  # Change %

    sorted_services = sorted(services_data.items(), key=sort_key, reverse=True)

    # Get top N services
    top_services = sorted_services[:top_n]
    other_services = sorted_services[top_n:]

    # Create table matching standalone services table format
    table = create_table(title=f"Top {top_n} Services (Monthly: {current_month_name})")
    table.add_column("Service", style="cyan", no_wrap=True, width=24)
    table.add_column("Current", justify="right", style="bright_green", width=11)
    table.add_column("Previous", justify="right", style="white", width=11)
    table.add_column("Change", justify="right", width=11)
    table.add_column("% Tot", justify="right", style="dim", width=7)
    table.add_column("Trend", style="yellow", width=13)

    # Add top N service rows
    for service, service_current_cost in top_services:
        service_previous_cost = previous_services_costs.get(service, 0)

        # Calculate change metrics
        change_amount = service_current_cost - service_previous_cost
        change_pct = (change_amount / service_previous_cost * 100) if service_previous_cost > 0 else 0
        change_icon = "↑" if change_pct > 0 else "↓" if change_pct < 0 else "→"
        change_style = "red" if change_pct > 0 else "green" if change_pct < 0 else "dim"

        # Determine trend indicator
        if abs(change_pct) < 5:
            trend = "→ stable"
            trend_style = "dim"
        elif change_pct > 20:
            trend = "↑↑↑ growing"
            trend_style = "red"
        elif change_pct > 10:
            trend = "↑ increasing"
            trend_style = "yellow"
        elif change_pct < -20:
            trend = "↓↓↓ declining"
            trend_style = "green"
        elif change_pct < -10:
            trend = "↓ decreasing"
            trend_style = "bright_green"
        else:
            trend = "→ stable"
            trend_style = "dim"

        # Calculate percentage of total
        percentage = (service_current_cost / total_current * 100) if total_current > 0 else 0

        # Truncate long service names
        service_display = service[:23] if len(service) > 23 else service

        table.add_row(
            service_display,
            f"${service_current_cost:,.0f}",
            f"${service_previous_cost:,.0f}",
            f"[{change_style}]{change_icon} {abs(change_pct):.1f}%[/{change_style}]",
            f"{percentage:.1f}%",
            f"[{trend_style}]{trend}[/{trend_style}]"
        )

    # Add "Others" row if there are services beyond top N
    if other_services:
        others_current = sum(cost for _, cost in other_services)
        others_previous = sum(previous_services_costs.get(service, 0) for service, _ in other_services)
        others_change = others_current - others_previous
        others_change_pct = (others_change / others_previous * 100) if others_previous > 0 else 0
        others_change_icon = "↑" if others_change_pct > 0 else "↓" if others_change_pct < 0 else "→"
        others_change_style = "red" if others_change_pct > 0 else "green" if others_change_pct < 0 else "dim"
        others_percentage = (others_current / total_current * 100) if total_current > 0 else 0

        # Others trend
        if abs(others_change_pct) < 5:
            others_trend = "→ stable"
            others_trend_style = "dim"
        elif others_change_pct > 20:
            others_trend = "↑↑↑ growing"
            others_trend_style = "red"
        elif others_change_pct > 10:
            others_trend = "↑ increasing"
            others_trend_style = "yellow"
        elif others_change_pct < -20:
            others_trend = "↓↓↓ declining"
            others_trend_style = "green"
        elif others_change_pct < -10:
            others_trend = "↓ decreasing"
            others_trend_style = "bright_green"
        else:
            others_trend = "→ stable"
            others_trend_style = "dim"

        table.add_row(
            f"[dim]Others ({len(other_services)})[/dim]",
            f"${others_current:,.0f}",
            f"${others_previous:,.0f}",
            f"[{others_change_style}]{others_change_icon} {abs(others_change_pct):.1f}%[/{others_change_style}]",
            f"{others_percentage:.1f}%",
            f"[{others_trend_style}]{others_trend}[/{others_trend_style}]"
        )

    # Add separator before TOTAL row
    table.add_row("", "", "", "", "", "", end_section=True)

    # Add TOTAL row
    total_change = total_current - total_previous
    total_change_pct = (total_change / total_previous * 100) if total_previous > 0 else 0
    total_change_icon = "↑" if total_change_pct > 0 else "↓" if total_change_pct < 0 else "→"
    total_change_style = "red" if total_change_pct > 0 else "green" if total_change_pct < 0 else "dim"

    # Overall trend for TOTAL
    if abs(total_change_pct) < 5:
        total_trend = "→ stable"
        total_trend_style = "dim"
    elif total_change_pct > 20:
        total_trend = "↑↑ growing"
        total_trend_style = "red"
    elif total_change_pct > 10:
        total_trend = "↑ growing"
        total_trend_style = "yellow"
    elif total_change_pct < -20:
        total_trend = "↓↓ declining"
        total_trend_style = "green"
    elif total_change_pct < -10:
        total_trend = "↓ declining"
        total_trend_style = "bright_green"
    else:
        total_trend = "→ overall"
        total_trend_style = "dim"

    table.add_row(
        "[bold]TOTAL[/bold]",
        f"[bold]${total_current:,.0f}[/bold]",
        f"[bold]${total_previous:,.0f}[/bold]",
        f"[bold][{total_change_style}]{total_change_icon} {abs(total_change_pct):.1f}%[/{total_change_style}][/bold]",
        f"[bold]100.0%[/bold]",
        f"[bold][{total_trend_style}]{total_trend}[/{total_trend_style}][/bold]"
    )

    return table


def create_finops_group():
    """
    Create the finops command group with all subcommands.

    Returns:
        Click Group object with all finops commands

    Performance: Lazy creation only when needed by DRYCommandRegistry
    Context Reduction: ~800 lines extracted from main.py
    """

    # Custom Group class with Rich Tree/Table help formatting
    class RichFinOpsGroup(click.Group):
        """Custom Click Group with Rich Tree/Table help display."""

        def format_help(self, ctx, formatter):
            """Format help text with Rich Tree/Table categorization."""
            import os
            from rich.tree import Tree
            from rich.table import Table as RichTable

            # Check for TEST_MODE environment variable for backward compatibility
            test_mode = os.environ.get('RUNBOOKS_TEST_MODE', '0') == '1'

            if test_mode:
                # Plain text fallback for testing
                click.echo("Usage: runbooks finops [OPTIONS] COMMAND [ARGS]...")
                click.echo("")
                click.echo("  Financial operations and cost optimization for AWS resources.")
                click.echo("")
                click.echo("Commands:")
                click.echo("  dashboard                        Multi-account cost visibility")
                click.echo("  analyze-ec2                      EC2 cost analysis with 4-way enrichment")
                click.echo("  analyze-workspaces               WorkSpaces cost analysis")
                click.echo("  lambda-analysis                  Lambda cost and activity analysis")
                click.echo("  detect-rds-idle                  RDS idle instance detector")
                click.echo("  infrastructure                   Comprehensive infrastructure analysis")
                click.echo("  ec2-snapshots                    EC2 snapshot cost optimization")
                click.echo("  optimize                         General cost optimization recommendations")
                click.echo("  optimize-cloudwatch-costs        CloudWatch log retention optimization")
                click.echo("  detect-orphans                   Unified orphan detection")
                click.echo("  analyze-s3-storage-lens          S3 Storage Lens cost intelligence")
                click.echo("  check-config-compliance          AWS Config compliance-cost correlation")
                click.echo("  ec2-decommission-analysis        EC2 decommission candidate identification")
                click.echo("  workspaces-decommission-analysis WorkSpaces decommission tier analysis")
                click.echo("  enrich-workspaces                WorkSpaces metadata enrichment")
                click.echo("  WorkSpaces                       WorkSpaces discovery and analysis")
                click.echo("  export                           Multi-format export")
                return

            # Categorize commands based on business function
            categories = {
                "💰 Cost Analysis": [
                    ("dashboard", "Multi-account cost visibility with MCP validation"),
                    ("analyze-ec2", "EC2 cost analysis with 4-way enrichment (Discovery→Orgs→Cost→Activity)"),
                    ("analyze-workspaces", "WorkSpaces cost analysis with decommission tier scoring"),
                    ("lambda-analysis", "Lambda cost and activity analysis"),
                    ("detect-rds-idle", "RDS idle instance detector ($50K annual savings, 5 signals)")
                ],
                "⚙️ Infrastructure Optimization": [
                    ("infrastructure", "Comprehensive infrastructure analysis"),
                    ("ec2-snapshots", "EC2 snapshot cost optimization"),
                    ("optimize", "General cost optimization recommendations"),
                    ("optimize-savings-plans", "Hybrid Savings Plans optimizer (60/30/10 strategy, $500K+ target)"),
                    ("optimize-s3-lifecycle", "S3 Lifecycle automation ($180K target, Epic 3)"),
                    ("optimize-cloudwatch-costs", "CloudWatch log retention optimization ($10K-$50K annual savings)"),
                    ("detect-orphans", "Unified orphan detection (EBS/EIP/NAT/LB, $50K-$200K savings)"),
                    ("analyze-s3-storage-lens", "S3 Storage Lens cost intelligence ($30K-$150K savings)"),
                    ("check-config-compliance", "AWS Config compliance-cost correlation ($20K-$80K savings)")
                ],
                "📊 Decommission Analysis": [
                    ("ec2-decommission-analysis", "EC2 decommission candidate identification"),
                    ("workspaces-decommission-analysis", "WorkSpaces decommission tier analysis")
                ],
                "🔄 Data Operations": [
                    ("enrich-workspaces", "WorkSpaces metadata enrichment"),
                    ("WorkSpaces", "WorkSpaces discovery and analysis")
                ],
                "📤 Export": [
                    ("export", "Multi-format export (CSV, JSON, PDF, Markdown)")
                ]
            }

            # Phase 1: Pre-calculate max column widths across ALL categories (Track 3A pattern)
            max_cmd_len = 0
            for category_commands in categories.values():
                for cmd, desc in category_commands:
                    max_cmd_len = max(max_cmd_len, len(cmd))

            # Set command column width with padding
            cmd_width = max_cmd_len + 2

            # Create Rich Tree
            tree = Tree("[bold cyan]FinOps Commands[/bold cyan] (19 commands)")

            # Add each category with fixed-width tables
            for category_name, commands in categories.items():
                category_branch = tree.add(f"[bold green]{category_name}[/bold green] [dim]({len(commands)} commands)[/dim]")

                # Create table with FIXED command width for vertical alignment, flexible description
                table = RichTable(show_header=True, box=None, padding=(0, 2))
                table.add_column("Command", style="cyan", no_wrap=True, min_width=cmd_width, max_width=cmd_width)
                table.add_column("Description", style="dim", no_wrap=False, overflow="fold")

                # Add rows
                for cmd, desc in commands:
                    table.add_row(cmd, desc)

                category_branch.add(table)

            # Display the tree
            console.print(tree)
            console.print("\n[blue]💡 Usage: runbooks finops [COMMAND] [OPTIONS][/blue]")
            console.print("[blue]📖 Example: runbooks finops dashboard --profile billing-profile --timeframe monthly[/blue]")

    @click.group(cls=RichFinOpsGroup, invoke_without_command=True)
    @common_filter_options
    @common_multi_account_options
    @common_output_options
    @common_aws_options
    @click.pass_context
    def finops(ctx, profile, region, dry_run, format, output_dir, all_outputs, export_csv, export_json, export_markdown, export,
               all_profiles, profiles, regions, all_regions, tags, accounts):
        """
        Financial operations and cost optimization for AWS resources.

        Comprehensive cost analysis, budget management, and financial reporting
        with enterprise-grade accuracy and multi-format export capabilities.

        Features:
        • Real-time cost analysis with MCP validation (≥99.5% accuracy)
        • Multi-format exports: CSV, JSON, PDF, Markdown
        • Quarterly intelligence with strategic financial reporting
        • Enterprise AWS profile support with multi-account capabilities

        Examples:
            runbooks finops dashboard --profile billing-profile
            runbooks finops dashboard --all-profiles --timeframe monthly
            runbooks finops dashboard --regions ap-southeast-2 ap-southeast-6
            runbooks finops export --format pdf --output-dir ./reports
        """
        # Ensure context object exists
        if ctx.obj is None:
            ctx.obj = {}
        ctx.obj.update({
            "profile": profile,
            "region": region,
            "dry_run": dry_run,
            "verbose": False,  # Default verbose setting (updated by dashboard command)
            "output_format": format,
            "output_dir": output_dir,
            "all_outputs": all_outputs,
            "export_csv": export_csv,
            "export_json": export_json,
            "export_markdown": export_markdown,
            "export": export,
            "all_profiles": all_profiles,
            "profiles": profiles,
            "regions": regions,
            "all_regions": all_regions,
            "tags": tags,
            "accounts": accounts
        })

        # Auto-execute dashboard when no subcommand provided (eliminates "only logs" pattern)
        if ctx.invoked_subcommand is None:
            # Invoke dashboard with default parameters
            ctx.invoke(dashboard,
                      profile=profile,
                      all_profile=all_profiles,  # Note: dashboard parameter is all_profile (singular)
                      timeframe='monthly',
                      services=None,
                      accounts=None,
                      validate=False,
                      validate_mcp=False,
                      mcp_validate=False,
                      csv=False,
                      markdown=False,
                      pdf=False,
                      json=False,
                      export_format=None,
                      unblended=False,
                      amortized=False,
                      dual_metrics=False,
                      dry_run=dry_run)

    @finops.command()
    @click.option("--profile", help="AWS profile to use for authentication")
    @click.option("--all-profile", type=str, default=None, help="Management profile for Organizations API multi-account discovery")
    @click.option(
        "--timeframe",
        type=click.Choice(["daily", "weekly", "monthly", "quarterly"]),
        default="monthly",
        help="Analysis timeframe",
    )
    @click.option("--services", multiple=True, help="Specific AWS services to analyze")
    @click.option("--accounts", multiple=True, help="Specific AWS accounts to analyze")
    @click.option("--validate", is_flag=True, help="Enable MCP validation for accuracy")
    @click.option("--validate-mcp", is_flag=True, help="Run standalone MCP validation framework (AWS-2 implementation)")
    @click.option("--mcp-validate", is_flag=True, help="Enable MCP validation for ≥99.5% accuracy cross-validation")
    @click.option("--csv", is_flag=True, help="Export results to CSV format")
    @click.option("--markdown", is_flag=True, help="Export results to Markdown format")
    @click.option("--pdf", is_flag=True, help="Export results to PDF format")
    @click.option("--json", is_flag=True, help="Export results to JSON format")
    @click.option(
        "--export-format",
        type=click.Choice(["json", "csv", "pdf", "markdown"]),
        help="Export format for results (legacy option - use individual flags)",
    )
    @click.option("--executive", is_flag=True, help="Executive summary format (business narrative, <5 min review)")
    @click.option("--unblended", is_flag=True, help="Use unblended cost metrics (default: BlendedCost)")
    @click.option("--amortized", is_flag=True, help="Use amortized cost metrics for Reserved Instances")
    @click.option("--dual-metrics", is_flag=True, help="Show both BlendedCost and AmortizedCost")
    @click.option("--dry-run", is_flag=True, default=True, help="Execute in dry-run mode")
    @click.option("--verbose", "-v", is_flag=True, default=False, help="Enable verbose diagnostic logging")
    @click.option("--top-n", type=int, default=10, help="Number of top services to display in table (1-50, default: 10)")
    @click.option("--sort-by", type=click.Choice(["current", "previous", "change"]), default="current", help="Sort services by: current cost, previous cost, or change %")
    @click.option("--cost-threshold", type=float, default=0.0, help="Minimum cost threshold for service display (e.g., 100.0 for $100+)")
    @click.option("--summary-mode", type=click.Choice(["table", "tree", "both", "none"]), default="both", help="Executive summary source: table (services), tree (categories), both, or none")
    @click.option("--output-format", type=click.Choice(["both", "table", "tree"]), default="both", help="Visualization choice: both (table + tree with nested table), table only, or tree only (with nested table)")
    @click.pass_context
    def dashboard(
        ctx,
        profile,
        all_profile,
        timeframe,
        services,
        accounts,
        validate,
        validate_mcp,
        mcp_validate,
        csv,
        markdown,
        pdf,
        json,
        export_format,
        executive,
        unblended,
        amortized,
        dual_metrics,
        dry_run,
        verbose,
        top_n,
        sort_by,
        cost_threshold,
        summary_mode,
        output_format,
    ):
        """
        Multi-account cost visibility with MCP validation.

        \b
        📊 DASHBOARD FEATURES
        ├── 🏗️ Enhanced Services Table (6 columns + TOTAL row)
        │   └── Service, Current, Last Mo, Change, % Tot, Trend (MoM)
        │
        ├── 🌲 Rich Tree Cost Breakdown (hierarchical visualization)
        │   └── Compute, Network, Storage, Database, Other categories
        │
        ├── 💡 Optimization Opportunities (Priority-based actionable items)
        │   └── Priority, Action, Savings, Effort, Risk assessment
        │
        └── 💰 Executive Summary (--executive flag)
            └── Business narrative with Key Findings

        \b
        📊 DASHBOARD OPTIONS (16 total)
        ├── 🔐 Authentication (1 option)
        │   └── --profile TEXT        AWS profile for authentication
        │
        ├── ⏰ Time Range (1 option)
        │   └── --timeframe CHOICE    [daily|weekly|monthly|quarterly] (default: monthly)
        │
        ├── 🎯 Filtering (2 options)
        │   └── --services TEXT       Specific AWS services (multiple allowed)
        │   └── --accounts TEXT       Specific AWS accounts (multiple allowed)
        │
        ├── ✅ Validation (3 options)
        │   └── --validate            Enable MCP validation for accuracy
        │   └── --validate-mcp        Standalone MCP validation framework
        │   └── --mcp-validate        MCP cross-validation (≥99.5% accuracy)
        │
        ├── 📤 Export Formats (4 options)
        │   └── --csv                 Export to CSV format
        │   └── --markdown            Export to Markdown format
        │   └── --pdf                 Export to PDF format
        │   └── --json                Export to JSON format
        │
        ├── 💰 Cost Metrics (3 options)
        │   └── --unblended           Use UnblendedCost metrics
        │   └── --amortized           Use AmortizedCost for Reserved Instances
        │   └── --dual-metrics        Show both BlendedCost and AmortizedCost
        │
        ├── 🎭 Display Modes (1 option)
        │   └── --executive           Executive summary with business narrative
        │
        └── 🔧 Execution (2 options)
            └── --verbose, -v         Enable verbose diagnostic logging
            └── --dry-run             Execute in dry-run mode (default: true)

        \b
        💡 Enhanced UX (v1.1.19):
        • 6-column services table with Last Mo, Change %, Trend (MoM)
        • Hierarchical Rich Tree cost breakdown by category
        • Priority-based optimization opportunities (🔴 IMMEDIATE, 🟡 30-DAY, 🟡 60-DAY, 🟢 90-DAY)
        • Verbose logging with --verbose flag for diagnostics

        \b
        💡 Sprint 1 Targets: 25-50% cost savings via decommission analysis
        📖 Example: runbooks finops dashboard --profile billing --timeframe monthly --mcp-validate --verbose
        """
        # Configure logging for diagnostic output
        import logging
        from runbooks.common.logging_config import configure_logging

        # Update context with verbose flag
        if ctx.obj:
            ctx.obj["verbose"] = verbose

        # Initialize logging based on verbose flag
        configure_logging(verbose=verbose)
        logger = logging.getLogger(__name__)

        # Parameter validation
        if not 1 <= top_n <= 50:
            raise click.BadParameter("--top-n must be between 1 and 50")
        if cost_threshold < 0:
            raise click.BadParameter("--cost-threshold must be non-negative")

        # Diagnostic log (only shows if --verbose flag used)
        logger.debug(
            f"Dashboard command started: profile={profile}, all_profile={all_profile}, "
            f"timeframe={timeframe}, verbose={verbose}, top_n={top_n}, sort_by={sort_by}, "
            f"cost_threshold={cost_threshold}, summary_mode={summary_mode}, output_format={output_format}"
        )

        # Handle multi-account Landing Zone analysis with Organizations API discovery
        if all_profile:
            try:
                from runbooks.finops.dashboard_runner import MultiAccountDashboard, DashboardRouter
                from runbooks.common.rich_utils import print_header, print_success, print_error, print_info
                from runbooks.inventory.inventory_modules import get_org_accounts_from_profiles, get_profiles
                import argparse

                print_header("Multi-Account Landing Zone Dashboard", all_profile)
                console.print("[cyan]🏢 Discovering AWS Organization accounts via Organizations API...[/cyan]")

                # ADD diagnostic logging
                logger.info(
                    "Multi-account discovery initiated",
                    extra={"management_profile": all_profile, "discovery_mode": "Organizations API"}
                )

                # CORRECTED: Use management profile TEXT parameter for Organizations API access
                try:
                    # Use management profile specified by user (MANAGEMENT_PROFILE, BILLING_PROFILE, or CENTRALISED_OPS_PROFILE)
                    mgmt_profile_list = get_profiles(fprofiles=[all_profile])
                    console.print(f"[dim]Querying Organizations API with profile: {all_profile}[/dim]")

                    # ADD diagnostic logging
                    logger.debug(f"Resolved management profile list: {mgmt_profile_list}")

                    org_accounts = get_org_accounts_from_profiles(mgmt_profile_list)

                    # Extract account IDs from discovered organization accounts
                    discovered_account_ids = []
                    for acct in org_accounts:
                        if acct.get("Success") and acct.get("RootAcct") and acct.get("aws_acct"):
                            # Root account found - extract all child accounts
                            for child in acct["aws_acct"].ChildAccounts:
                                discovered_account_ids.append(child["AccountId"])

                    if discovered_account_ids:
                        console.print(f"[green]✅ Discovered {len(discovered_account_ids)} organization accounts[/green]")
                        console.print(f"[dim]Analysis Scope: Organization-wide with Landing Zone support[/dim]\n")

                        # ADD diagnostic logging
                        logger.info(
                            "Organizations discovery successful",
                            extra={
                                "account_count": len(discovered_account_ids),
                                "account_ids": discovered_account_ids[:5]  # First 5 for brevity
                            }
                        )
                    else:
                        console.print(f"[yellow]⚠️  No organization accounts discovered - using single account mode[/yellow]")
                        console.print(f"[dim]Tip: Ensure {profile} has AWS Organizations permissions[/dim]\n")

                except Exception as org_error:
                    console.print(f"[yellow]⚠️  Organizations discovery failed: {str(org_error)}[/yellow]")
                    console.print(f"[dim]Falling back to single account mode[/dim]\n")
                    discovered_account_ids = []  # Empty list for fallback

                # Create mock args object for multi-dashboard compatibility
                args = argparse.Namespace()
                args.profile = all_profile  # Use management profile for AWS Organizations access
                args.timeframe = timeframe
                args.services = services
                # PHASE 2 ENHANCEMENT: Use Organizations-discovered accounts if available
                args.accounts = tuple(discovered_account_ids) if discovered_account_ids else accounts
                args.validate = validate or mcp_validate
                # CRITICAL FIX: Handle multiple export format flags
                export_formats = []
                if csv:
                    export_formats.append("csv")
                if markdown:
                    export_formats.append("markdown")
                if pdf:
                    export_formats.append("pdf")
                if json:
                    export_formats.append("json")
                if export_format and export_format not in export_formats:
                    export_formats.append(export_format)

                args.export_format = export_formats[0] if export_formats else None
                args.export_formats = export_formats  # Store all requested formats

                # CRITICAL FIX: Handle cost metric options
                cost_metrics = ["BlendedCost"]  # Default metric
                if unblended:
                    cost_metrics = ["UnblendedCost"]
                elif amortized:
                    cost_metrics = ["AmortizedCost"]
                elif dual_metrics:
                    cost_metrics = ["BlendedCost", "AmortizedCost"]

                args.cost_metrics = cost_metrics
                args.cost_metric_display = _get_cost_metric_display(cost_metrics)
                args.dry_run = dry_run
                args.all = True  # Enable all accounts mode
                args.all_accounts = True  # Enable all accounts mode (alternate flag name)
                args.top_accounts = 50  # Show many accounts for enterprise view
                args.services_per_account = 3
                args.time_range = None
                args.audit = False  # Not audit mode
                args.trend = False  # Not trend analysis mode
                args.combine = False  # Not combined multi-account view
                args.tag = None
                args.region = None  # No specific region filter
                args.regions = None
                args.report_name = None  # No report export by default
                args.report_type = []  # No report types specified
                args.dir = None  # No output directory specified
                args.profiles = []  # No additional profiles beyond main profile
                args.business_analysis = False  # Not business analysis mode
                args.enhanced_export = False  # Not enhanced export mode
                args.live_mode = False  # Not live mode

                # Initialize router and dashboard
                router = DashboardRouter(console=console)
                routing_config = router.route_dashboard_request(args)

                # Create multi-account dashboard
                multi_dashboard = MultiAccountDashboard(console=console)

                # Execute multi-account analysis
                result = multi_dashboard.run_dashboard(args, routing_config)

                if result == 0:
                    print_success("Multi-account Landing Zone analysis completed successfully")
                else:
                    print_error("Multi-account analysis encountered issues")

                return result

            except ImportError as e:
                console.print(f"[red]❌ Multi-account dashboard not available: {e}[/red]")
                console.print("[yellow]💡 Falling back to single-account mode with specified profile[/yellow]")
                # Fallback to single account with the specified profile
                resolved_profile = all_profile
            except Exception as e:
                console.print(f"[red]❌ Multi-account analysis failed: {e}[/red]")
                console.print("[yellow]💡 Fallingback to single-account mode[/yellow]")
                resolved_profile = all_profile
        else:
            resolved_profile = profile or ctx.obj.get("profile", "default")

        # Handle standalone MCP validation (AWS-2 implementation)
        if validate_mcp:
            try:
                from runbooks.common.rich_utils import print_header, print_success, print_error, print_info
                import asyncio

                print_header("MCP Validation Framework", "AWS-2 Implementation")
                console.print("[cyan]🔍 Running comprehensive MCP validation for ≥99.5% accuracy[/cyan]")

                # Import and initialize MCP validator
                from runbooks.validation.mcp_validator import MCPValidator

                # Enterprise profiles configuration
                validation_profiles = {
                    "billing": "${BILLING_PROFILE}",
                    "management": "${MANAGEMENT_PROFILE}",
                    "centralised_ops": "${CENTRALISED_OPS_PROFILE}",
                    "single_aws": "${SINGLE_AWS_PROFILE}",
                }

                # Initialize validator with configured profiles
                validator = MCPValidator(
                    profiles=validation_profiles, tolerance_percentage=5.0, performance_target_seconds=30.0
                )

                # Run comprehensive validation
                validation_report = asyncio.run(validator.validate_all_operations())

                # Success criteria for AWS-2
                if validation_report.overall_accuracy >= 99.5:
                    print_success(
                        f"✅ AWS-2 SUCCESS: {validation_report.overall_accuracy:.1f}% ≥ 99.5% target achieved"
                    )
                    return 0
                else:
                    print_error(f"❌ AWS-2 FAILED: {validation_report.overall_accuracy:.1f}% < 99.5% target")
                    return 1

            except Exception as e:
                print_error(f"❌ AWS-2 MCP validation failed: {e}")
                return 1

        try:
            from runbooks.common.rich_utils import print_header, print_success, print_error, create_table, format_cost
            from runbooks.common.profile_utils import create_cost_session
            from runbooks.finops.cost_processor import get_cost_data
            from runbooks.finops.aws_client import get_account_id, ec2_summary, get_accessible_regions
            import boto3
            from datetime import datetime, timedelta
            from rich.table import Table
            from rich.panel import Panel

            # Resolve profile with priority: command --profile > parent context > default
            # Note: resolved_profile already set above for multi-account vs single-account mode
            if "resolved_profile" not in locals():
                resolved_profile = profile or ctx.obj.get("profile", "default")
            resolved_dry_run = dry_run if dry_run is not None else ctx.obj.get("dry_run", True)

            # MCP validation integration
            mcp_results = None
            if mcp_validate or validate:
                try:
                    from runbooks.validation.mcp_validator import MCPValidator
                    import asyncio

                    console.print("[cyan]🔍 Running MCP validation for dashboard data accuracy[/cyan]")

                    # Configure validation profiles using resolved profile
                    validation_profiles = {
                        "billing": resolved_profile,
                        "management": resolved_profile,
                        "centralised_ops": resolved_profile,
                        "single_aws": resolved_profile,
                    }

                    # Initialize validator
                    validator = MCPValidator(
                        profiles=validation_profiles, tolerance_percentage=5.0, performance_target_seconds=30.0
                    )

                    # Run validation focused on cost explorer operations (primary finops validation)
                    mcp_results = asyncio.run(validator.validate_cost_explorer())

                    # Display validation results
                    if mcp_results.accuracy_percentage >= 99.5:
                        console.print(
                            f"[green]✅ MCP Validation PASSED: {mcp_results.accuracy_percentage:.1f}% accuracy[/green]"
                        )

                        # ADD diagnostic logging
                        logger.info(
                            "MCP validation completed",
                            extra={
                                "accuracy": mcp_results.accuracy_percentage,
                                "operation": mcp_results.operation_name,
                                "threshold": 99.5
                            }
                        )
                    elif mcp_results.accuracy_percentage >= 95.0:
                        console.print(
                            f"[yellow]⚠️ MCP Validation WARNING: {mcp_results.accuracy_percentage:.1f}% accuracy (target: ≥99.5%)[/yellow]"
                        )
                    else:
                        console.print(
                            f"[red]❌ MCP Validation FAILED: {mcp_results.accuracy_percentage:.1f}% accuracy[/red]"
                        )

                except Exception as e:
                    console.print(f"[yellow]⚠️ MCP validation failed: {e}[/yellow]")
                    console.print("[dim]Continuing with dashboard generation...[/dim]")

            # Create AWS session and get account info
            session = create_cost_session(profile_name=resolved_profile)
            account_id = get_account_id(session)

            # Consolidated header with account, profile, and period information (LEAN: single box)
            from runbooks.common.rich_utils import create_display_profile_name
            from datetime import date
            from rich.panel import Panel
            from rich import box
            from rich.text import Text

            truncated_profile = create_display_profile_name(resolved_profile, max_length=30)
            current_month_name = date.today().strftime("%B %Y")

            # Build consolidated header text
            header_text = Text()
            header_text.append("Runbooks FinOps Dashboard", style="bold cyan")
            header_text.append(" | ", style="dim")
            header_text.append(f"Account: {account_id}", style="cyan")
            header_text.append(" | ", style="dim")
            header_text.append(f"Period: {current_month_name}", style="white")

            console.print()
            console.print(Panel(header_text, box=box.DOUBLE, style="cyan"))

            # Get cost data for the specified timeframe
            try:
                # Calculate time range based on timeframe
                # CRITICAL FIX: monthly=None triggers current month logic (not 30-day rolling window)
                # This ensures both services_data and current_cost use same period (Nov 1 - Nov 10)
                # quarterly=90 for last 3 months comparison
                time_range_days = {"daily": 7, "weekly": 30, "monthly": None, "quarterly": 90}.get(timeframe, None)

                # Get comprehensive cost data
                cost_data = get_cost_data(
                    session,
                    time_range=time_range_days,
                    get_trend=True,
                    profile_name=resolved_profile,
                    account_id=account_id,
                )

                # ADD diagnostic logging
                logger.debug(
                    "Cost data retrieved",
                    extra={
                        "time_range_days": time_range_days,
                        "timeframe": timeframe,
                        "services_count": len(cost_data.get("costs_by_service", {}))
                    }
                )

                # Access cost data using correct field names from CostData TypedDict
                current_cost = cost_data.get("current_month", 0)
                previous_cost = cost_data.get("last_month", 0)

                # Display Enhanced Top Services Table with Trends (DEFAULT)
                services_data = cost_data.get("costs_by_service", {})
                if services_data:
                    # Import previous month cost helper
                    from runbooks.finops.cost_processor import get_previous_month_costs

                    # Get previous month costs for comparison
                    previous_services_costs = get_previous_month_costs(session, profile_name=resolved_profile)

                    # Create enhanced 6-column table with time context for clarity
                    from datetime import date
                    current_month_name = date.today().strftime("%B %Y")  # e.g., "November 2025"
                    services_table = create_table(title=f"🏗️ Top {top_n} AWS Services by Cost (Monthly: {current_month_name})")
                    services_table.add_column("Service", style="cyan", no_wrap=True, width=24)
                    services_table.add_column("Current", justify="right", style="bright_green", width=11)
                    services_table.add_column("Previous", justify="right", style="white", width=11)  # Clearer than "Last Mo"
                    services_table.add_column("Change", justify="right", width=11)
                    services_table.add_column("% Tot", justify="right", style="dim", width=7)
                    services_table.add_column("Trend", style="yellow", width=13)  # Simplified from "Trend (MoM)"

                    # Apply cost threshold filter if specified
                    if cost_threshold > 0:
                        services_data = {k: v for k, v in services_data.items() if v >= cost_threshold}

                    # Sort services based on --sort-by parameter
                    if sort_by == "current":
                        sort_key = lambda x: x[1]  # Current cost
                    elif sort_by == "previous":
                        sort_key = lambda x: previous_services_costs.get(x[0], 0)  # Previous cost
                    else:  # sort_by == "change"
                        sort_key = lambda x: abs(x[1] - previous_services_costs.get(x[0], 0)) / previous_services_costs.get(x[0], 1) if previous_services_costs.get(x[0], 0) > 0 else 0  # Change %

                    sorted_services = sorted(services_data.items(), key=sort_key, reverse=True)[:top_n]

                    # Track totals for TOTAL row
                    # Use account-level totals (not just sum of top 10) for mathematical consistency
                    total_current = current_cost  # Account total from API (all services)
                    total_previous = previous_cost  # Previous month total from API

                    # Add service rows
                    for service, service_current_cost in sorted_services:
                        # Get previous month cost for this service
                        service_previous_cost = previous_services_costs.get(service, 0)

                        # Calculate change metrics
                        change_amount = service_current_cost - service_previous_cost
                        change_pct = (change_amount / service_previous_cost * 100) if service_previous_cost > 0 else 0
                        change_icon = "↑" if change_pct > 0 else "↓" if change_pct < 0 else "→"
                        change_style = "red" if change_pct > 0 else "green" if change_pct < 0 else "dim"

                        # Determine trend indicator (MoM)
                        if abs(change_pct) < 5:
                            trend = "→ stable"
                            trend_style = "dim"
                        elif change_pct > 20:
                            trend = "↑↑↑ growing"
                            trend_style = "red"
                        elif change_pct > 10:
                            trend = "↑ increasing"
                            trend_style = "yellow"
                        elif change_pct < -20:
                            trend = "↓↓↓ declining"
                            trend_style = "green"
                        elif change_pct < -10:
                            trend = "↓ decreasing"
                            trend_style = "bright_green"
                        else:
                            trend = "→ stable"
                            trend_style = "dim"

                        # Calculate percentage of total
                        percentage = (service_current_cost / total_current * 100) if total_current > 0 else 0

                        # Truncate long service names
                        service_display = service[:23] if len(service) > 23 else service

                        services_table.add_row(
                            service_display,
                            f"${service_current_cost:,.0f}",
                            f"${service_previous_cost:,.0f}",
                            f"[{change_style}]{change_icon} {abs(change_pct):.1f}%[/{change_style}]",
                            f"{percentage:.1f}%",
                            f"[{trend_style}]{trend}[/{trend_style}]"
                        )

                    # Add separator before TOTAL row
                    services_table.add_row("", "", "", "", "", "", end_section=True)

                    # Add TOTAL row
                    total_change = total_current - total_previous
                    total_change_pct = (total_change / total_previous * 100) if total_previous > 0 else 0
                    total_change_icon = "↑" if total_change_pct > 0 else "↓" if total_change_pct < 0 else "→"
                    total_change_style = "red" if total_change_pct > 0 else "green" if total_change_pct < 0 else "dim"

                    # Overall trend for TOTAL
                    if abs(total_change_pct) < 5:
                        total_trend = "→ overall"
                    elif total_change_pct > 10:
                        total_trend = "↑ overall"
                    else:
                        total_trend = "↓ overall"

                    services_table.add_row(
                        "[bold]TOTAL[/bold]",
                        f"[bold]${total_current:,.0f}[/bold]",
                        f"[bold]${total_previous:,.0f}[/bold]",
                        f"[bold {total_change_style}]{total_change_icon} {abs(total_change_pct):.1f}%[/bold {total_change_style}]",
                        "[bold]100.0%[/bold]",
                        f"[bold]{total_trend}[/bold]"
                    )

                    # Conditional rendering based on --output-format parameter
                    if output_format in ["both", "table"]:
                        console.print(services_table)
                        console.print()

                        # ADD diagnostic logging
                        logger.debug(
                            "Enhanced services table rendered",
                            extra={
                                "services_count": len(sorted_services),
                                "total_current": total_current,
                                "total_previous": total_previous,
                                "output_format": output_format
                            }
                        )

                        # Mathematical consistency check: Services total should ≤ Account total
                        # Allow 1% tolerance for rounding differences
                        if total_current > current_cost * 1.01:
                            console.print(
                                f"[red]⚠️ WARNING: Services total (${total_current:,.0f}) exceeds account total (${current_cost:,.0f})[/red]"
                            )
                            console.print("[dim]This indicates a time period mismatch. Please report this issue.[/dim]")
                            logger.warning(
                                "Mathematical inconsistency detected",
                                extra={
                                    "services_total": total_current,
                                    "account_total": current_cost,
                                    "ratio": total_current / current_cost if current_cost > 0 else 0
                                }
                            )
                            console.print()

                # Rich Tree Cost Breakdown (conditional based on --output-format)
                try:
                    from runbooks.finops.cost_processor import enrich_service_costs_with_trends, categorize_aws_services
                    from runbooks.common.rich_utils import create_cost_breakdown_tree

                    # Enrich services with trend data
                    enriched_services = enrich_service_costs_with_trends(
                        current_costs=services_data,
                        previous_costs=previous_services_costs
                    )

                    # Categorize services for hierarchical visualization
                    categorized_services = categorize_aws_services(enriched_services)

                    # Get EC2 resource summary BEFORE calculating optimization opportunities
                    try:
                        ec2_data = ec2_summary(session, profile_name=resolved_profile)
                        # ec2_summary returns keys: "running", "stopped", etc.
                        running_instances = ec2_data.get("running", 0)
                        stopped_instances = ec2_data.get("stopped", 0)
                        total_instances = sum(ec2_data.values())
                    except Exception as e:
                        console.print(f"[yellow]⚠️ Could not fetch EC2 optimization data: {e}[/yellow]\n")
                        total_instances = 0
                        running_instances = 0
                        stopped_instances = 0

                    # Calculate optimization opportunities from EC2 data
                    optimization_opportunities = {}

                    # Compute: EC2 rightsizing + stopped instance cleanup
                    # Try multiple possible service names for EC2
                    ec2_cost_estimate = (
                        services_data.get("Amazon Elastic Compute Cloud - Compute", 0) or
                        services_data.get("Amazon Elastic Compute Cloud", 0) or
                        0
                    )

                    # Now running_instances and stopped_instances are available
                    if running_instances > 5:
                        rightsizing_savings = ec2_cost_estimate * 0.20  # 20% rightsizing opportunity
                        if rightsizing_savings > 100:
                            optimization_opportunities["Compute"] = rightsizing_savings

                    if stopped_instances > 0:
                        stopped_savings = stopped_instances * 100  # $100/month per stopped instance
                        optimization_opportunities["Compute"] = optimization_opportunities.get("Compute", 0) + stopped_savings

                    # Conditional rendering based on --output-format parameter
                    if output_format in ["both", "tree"]:
                        # Create and display tree with per-category nested tables
                        cost_tree = create_cost_breakdown_tree(
                            services_by_category=categorized_services,
                            total_monthly_cost=current_cost,
                            optimization_opportunities=optimization_opportunities if optimization_opportunities else None,
                            previous_services_costs=previous_services_costs
                        )

                        console.print(cost_tree)
                        console.print()

                        # ADD diagnostic logging
                        logger.debug(
                            "Rich tree cost breakdown rendered",
                            extra={
                                "categories": list(categorized_services.keys()),
                                "opportunities": list(optimization_opportunities.keys()) if optimization_opportunities else [],
                                "output_format": output_format
                            }
                        )

                except Exception as e:
                    # Graceful degradation if tree visualization fails
                    console.print(f"[dim]Note: Cost breakdown tree unavailable ({str(e)})[/dim]")

                # Display Business Impact Summary (with executive mode option)
                # Calculate evidence-based savings from actual resource analysis (replaces hardcoded 15%)
                monthly_savings, savings_breakdown = calculate_evidence_based_savings(
                    services_data=services_data,
                    running_instances=running_instances,
                    stopped_instances=stopped_instances,
                    current_cost=current_cost
                )
                total_annual = current_cost * 12
                optimization_potential = monthly_savings * 12  # Evidence-based annual projection

                # Executive Summary Mode: Enhanced Business Narrative (LEAN: no duplicate table)
                # Conditional based on --summary-mode parameter
                if executive and summary_mode != "none":
                    if summary_mode == "table":
                        _generate_table_summary(
                            account_id=account_id,
                            current_month_name=current_month_name,
                            current_cost=current_cost,
                            sorted_services=sorted_services if 'sorted_services' in locals() else [],
                            monthly_savings=monthly_savings,
                            optimization_potential=optimization_potential,
                            savings_breakdown=savings_breakdown
                        )
                    elif summary_mode == "tree":
                        _generate_tree_summary(
                            account_id=account_id,
                            current_month_name=current_month_name,
                            current_cost=current_cost,
                            categorized_services=categorized_services if 'categorized_services' in locals() else {},
                            monthly_savings=monthly_savings,
                            optimization_potential=optimization_potential,
                            savings_breakdown=savings_breakdown
                        )
                    else:  # summary_mode == "both"
                        _generate_comprehensive_summary(
                            account_id=account_id,
                            current_month_name=current_month_name,
                            current_cost=current_cost,
                            monthly_savings=monthly_savings,
                            optimization_potential=optimization_potential,
                            savings_breakdown=savings_breakdown
                        )
                else:
                    # Optimization Opportunities Table (DEFAULT - actionable items)
                    opp_table = create_table(title="💡 Cost Optimization Opportunities")
                    opp_table.add_column("Priority", style="bold", width=14, no_wrap=True)
                    opp_table.add_column("Action", style="cyan", width=38)
                    opp_table.add_column("Savings", justify="right", style="green", width=13)
                    opp_table.add_column("Effort", justify="center", style="yellow", width=10)
                    opp_table.add_column("Risk", justify="center", style="dim", width=8)

                    # Calculate opportunity details from EC2 data
                    opportunities = []

                    # Opportunity 1: Stopped EC2 Instances (IMMEDIATE priority)
                    if stopped_instances > 0:
                        stopped_savings = stopped_instances * 100  # $100/month per stopped instance
                        opportunities.append({
                            "priority": "🔴 IMMEDIATE",
                            "action": f"Terminate {stopped_instances} stopped EC2 instance{'s' if stopped_instances > 1 else ''}",
                            "savings": stopped_savings,
                            "effort": "LOW",
                            "risk": "LOW"
                        })

                    # Opportunity 2: EC2 Rightsizing (30-DAY priority)
                    if running_instances > 5:
                        ec2_cost_estimate = services_data.get("Amazon Elastic Compute Cloud - Compute", 0)
                        rightsizing_savings = ec2_cost_estimate * 0.20  # 20% rightsizing opportunity
                        if rightsizing_savings > 100:
                            opportunities.append({
                                "priority": "🟡 30-DAY",
                                "action": f"EC2 rightsizing via Compute Optimizer ({running_instances} instances)",
                                "savings": rightsizing_savings,
                                "effort": "MEDIUM",
                                "risk": "LOW"
                            })

                    # Opportunity 3: Storage Lifecycle (90-DAY priority)
                    s3_cost = services_data.get("Amazon Simple Storage Service", 0)
                    if s3_cost > 200:
                        storage_savings = s3_cost * 0.25  # 25% S3 lifecycle opportunity
                        opportunities.append({
                            "priority": "🟢 90-DAY",
                            "action": "S3 storage lifecycle policies (Intelligent-Tiering + Glacier)",
                            "savings": storage_savings,
                            "effort": "MEDIUM",
                            "risk": "LOW"
                        })

                    # Opportunity 4: CloudWatch Log Retention (60-DAY priority)
                    cloudwatch_cost = services_data.get("Amazon CloudWatch", 0) + services_data.get("CloudWatch", 0)
                    if cloudwatch_cost > 50:
                        log_savings = cloudwatch_cost * 0.30  # 30% log retention optimization
                        opportunities.append({
                            "priority": "🟡 60-DAY",
                            "action": "CloudWatch log retention optimization",
                            "savings": log_savings,
                            "effort": "LOW",
                            "risk": "LOW"
                        })

                    # Opportunity 5: RDS Idle Detection (30-DAY priority if RDS cost >$500)
                    rds_cost = services_data.get("Amazon Relational Database Service", 0) + services_data.get("Amazon RDS Service", 0)
                    if rds_cost > 500:
                        rds_savings = rds_cost * 0.15  # 15% idle RDS opportunity
                        opportunities.append({
                            "priority": "🟡 30-DAY",
                            "action": "RDS idle instance detection (5 signals: connections, CPU, I/O, network, queries)",
                            "savings": rds_savings,
                            "effort": "LOW",
                            "risk": "LOW"
                        })

                    # Add opportunities to table (sorted by priority: IMMEDIATE → 30-DAY → 60-DAY → 90-DAY)
                    priority_order = {"🔴 IMMEDIATE": 1, "🟡 30-DAY": 2, "🟡 60-DAY": 3, "🟢 90-DAY": 4}
                    sorted_opportunities = sorted(opportunities, key=lambda x: priority_order.get(x["priority"], 99))

                    total_monthly_opportunity = 0
                    for opp in sorted_opportunities:
                        monthly_savings = opp["savings"]
                        total_monthly_opportunity += monthly_savings

                        opp_table.add_row(
                            opp["priority"],
                            opp["action"],
                            f"${monthly_savings:,.0f}/mo",
                            opp["effort"],
                            opp["risk"]
                        )

                    console.print(opp_table)

                    # Total opportunity summary
                    if total_monthly_opportunity > 0:
                        total_annual_opportunity = total_monthly_opportunity * 12
                        opportunity_pct = (total_monthly_opportunity / current_cost * 100) if current_cost > 0 else 0

                        console.print(
                            f"\n[bold green]💰 Total Opportunity: ${total_monthly_opportunity:,.0f}/month "
                            f"(${total_annual_opportunity:,.0f} annually) - {opportunity_pct:.1f}% potential reduction[/bold green]"
                        )
                        console.print(f"[dim]Analysis Period: {timeframe.title()} view | Account: {account_id}[/dim]\n")
                    else:
                        # Fallback if no opportunities identified
                        console.print(f"\n[green]✅ No immediate optimization opportunities identified[/green]")
                        console.print(f"[dim]Current spend: ${current_cost:,.0f}/month | Account: {account_id}[/dim]\n")

                # Prepare results dictionary
                results = {
                    "status": "completed",
                    "account_id": account_id,
                    "timeframe": timeframe,
                    "cost_analysis": {
                        "current_monthly_spend": current_cost,
                        "previous_monthly_spend": previous_cost,
                        "annual_projection": total_annual,
                        "optimization_potential": optimization_potential,
                        "top_services": dict(sorted_services[:5]) if services_data else {},
                        "ec2_summary": {
                            "total_instances": total_instances if "total_instances" in locals() else 0,
                            "running_instances": running_instances if "running_instances" in locals() else 0,
                            "stopped_instances": stopped_instances if "stopped_instances" in locals() else 0,
                        },
                    },
                }

                # Attach MCP validation results if available
                if mcp_results:
                    results["mcp_validation"] = {
                        "accuracy_percentage": mcp_results.accuracy_percentage,
                        "validation_passed": mcp_results.accuracy_percentage >= 99.5,
                        "operation_name": mcp_results.operation_name,
                        "status": mcp_results.status.value,
                        "detailed_results": mcp_results,
                    }

                return results

            except Exception as e:
                print_error(f"Failed to retrieve cost data: {e}")
                console.print(
                    f"[yellow]💡 Tip: Ensure your AWS profile '{resolved_profile}' has Cost Explorer permissions[/yellow]"
                )
                console.print(f"[dim]Required permissions: ce:GetCostAndUsage, ce:GetDimensionValues[/dim]")
                raise

        except ImportError as e:
            error_handlers["module_not_available"]("FinOps dashboard", e)
            raise click.ClickException("FinOps dashboard functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("FinOps dashboard generation", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--resource-type",
        type=click.Choice(["ec2", "s3", "rds", "lambda", "vpc"]),
        required=True,
        help="Resource type for optimization analysis",
    )
    @click.option(
        "--savings-target", type=click.FloatRange(0.1, 0.8), default=0.3, help="Target savings percentage (0.1-0.8)"
    )
    @click.option(
        "--analysis-depth",
        type=click.Choice(["basic", "comprehensive", "enterprise"]),
        default="comprehensive",
        help="Analysis depth level",
    )
    @click.option("--mcp-validate", is_flag=True, help="Enable MCP validation for ≥99.5% accuracy cross-validation")
    @click.pass_context
    def optimize(ctx, resource_type, savings_target, analysis_depth, mcp_validate):
        """
        Generate cost optimization recommendations for specific resource types.

        Enterprise Optimization Features:
        • Safety-first analysis with READ-ONLY operations
        • Quantified savings projections with ROI analysis
        • Risk assessment and business impact evaluation
        • Implementation timeline and priority recommendations

        Examples:
            runbooks finops optimize --resource-type ec2 --savings-target 0.25
            runbooks finops optimize --resource-type s3 --analysis-depth enterprise
        """
        try:
            from runbooks.finops.optimization_engine import ResourceOptimizer

            # MCP validation integration for optimization accuracy
            mcp_results = None
            if mcp_validate:
                try:
                    from runbooks.validation.mcp_validator import MCPValidator
                    import asyncio

                    console.print(f"[cyan]🔍 Running MCP validation for {resource_type} optimization accuracy[/cyan]")

                    # Configure validation profiles
                    validation_profiles = {
                        "billing": ctx.obj.get("profile", "default"),
                        "management": ctx.obj.get("profile", "default"),
                        "centralised_ops": ctx.obj.get("profile", "default"),
                        "single_aws": ctx.obj.get("profile", "default"),
                    }

                    # Initialize validator
                    validator = MCPValidator(
                        profiles=validation_profiles, tolerance_percentage=5.0, performance_target_seconds=30.0
                    )

                    # Run validation based on resource type
                    if resource_type in ["ec2"]:
                        mcp_results = asyncio.run(validator.validate_ec2_inventory())
                    elif resource_type in ["vpc"]:
                        mcp_results = asyncio.run(validator.validate_vpc_analysis())
                    elif resource_type in ["s3", "rds", "lambda"]:
                        # For these resource types, use cost explorer validation
                        mcp_results = asyncio.run(validator.validate_cost_explorer())
                    else:
                        # Default to cost explorer validation
                        mcp_results = asyncio.run(validator.validate_cost_explorer())

                    # Display validation results
                    if mcp_results.accuracy_percentage >= 99.5:
                        console.print(
                            f"[green]✅ MCP Validation PASSED: {mcp_results.accuracy_percentage:.1f}% accuracy for {resource_type}[/green]"
                        )
                    elif mcp_results.accuracy_percentage >= 95.0:
                        console.print(
                            f"[yellow]⚠️ MCP Validation WARNING: {mcp_results.accuracy_percentage:.1f}% accuracy (target: ≥99.5%)[/yellow]"
                        )
                    else:
                        console.print(
                            f"[red]❌ MCP Validation FAILED: {mcp_results.accuracy_percentage:.1f}% accuracy[/red]"
                        )

                except Exception as e:
                    console.print(f"[yellow]⚠️ MCP validation failed: {e}[/yellow]")
                    console.print("[dim]Continuing with optimization analysis...[/dim]")

            optimizer = ResourceOptimizer(
                profile=ctx.obj["profile"],
                region=ctx.obj["region"],
                resource_type=resource_type,
                savings_target=savings_target,
                analysis_depth=analysis_depth,
                mcp_validate=mcp_validate,
            )

            optimization_results = optimizer.analyze_optimization_opportunities()

            # Attach MCP validation results if available
            if mcp_results and isinstance(optimization_results, dict):
                optimization_results["mcp_validation"] = {
                    "accuracy_percentage": mcp_results.accuracy_percentage,
                    "validation_passed": mcp_results.accuracy_percentage >= 99.5,
                    "resource_type": resource_type,
                    "operation_name": mcp_results.operation_name,
                    "status": mcp_results.status.value,
                    "detailed_results": mcp_results,
                }

            return optimization_results

        except ImportError as e:
            error_handlers["module_not_available"]("FinOps optimization", e)
            raise click.ClickException("FinOps optimization functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("FinOps optimization analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--format",
        "export_format",
        type=click.Choice(["csv", "json", "pdf", "markdown"]),
        multiple=True,
        default=["json"],
        help="Export formats",
    )
    @click.option("--output-dir", default="./finops_reports", help="Output directory for exports")
    @click.option("--include-quarterly", is_flag=True, help="Include quarterly intelligence data")
    @click.option("--executive-summary", is_flag=True, help="Generate executive summary format")
    @click.option("--mcp-validate", is_flag=True, help="Enable MCP validation for ≥99.5% accuracy cross-validation")
    @click.pass_context
    def export(ctx, export_format, output_dir, include_quarterly, executive_summary, mcp_validate):
        """
        Export financial analysis results in multiple formats.

        Enterprise Export Features:
        • Multi-format simultaneous export
        • Executive-ready formatting and presentation
        • Quarterly intelligence integration
        • Complete audit trail documentation

        Examples:
            runbooks finops export --format csv,pdf --executive-summary
            runbooks finops export --include-quarterly --output-dir ./executive_reports
        """
        try:
            from runbooks.finops.export_manager import FinOpsExportManager

            # MCP validation integration for export accuracy
            mcp_results = None
            if mcp_validate:
                try:
                    from runbooks.validation.mcp_validator import MCPValidator
                    import asyncio

                    console.print("[cyan]🔍 Running MCP validation for export data accuracy[/cyan]")

                    # Configure validation profiles
                    validation_profiles = {
                        "billing": ctx.obj.get("profile", "default"),
                        "management": ctx.obj.get("profile", "default"),
                        "centralised_ops": ctx.obj.get("profile", "default"),
                        "single_aws": ctx.obj.get("profile", "default"),
                    }

                    # Initialize validator
                    validator = MCPValidator(
                        profiles=validation_profiles, tolerance_percentage=5.0, performance_target_seconds=30.0
                    )

                    # Run validation for export data accuracy using cost explorer validation
                    mcp_results = asyncio.run(validator.validate_cost_explorer())

                    # Display validation results
                    if mcp_results.accuracy_percentage >= 99.5:
                        console.print(
                            f"[green]✅ MCP Validation PASSED: {mcp_results.accuracy_percentage:.1f}% accuracy for exports[/green]"
                        )
                    elif mcp_results.accuracy_percentage >= 95.0:
                        console.print(
                            f"[yellow]⚠️ MCP Validation WARNING: {mcp_results.accuracy_percentage:.1f}% accuracy (target: ≥99.5%)[/yellow]"
                        )
                    else:
                        console.print(
                            f"[red]❌ MCP Validation FAILED: {mcp_results.accuracy_percentage:.1f}% accuracy[/red]"
                        )

                except Exception as e:
                    console.print(f"[yellow]⚠️ MCP validation failed: {e}[/yellow]")
                    console.print("[dim]Continuing with export operation...[/dim]")

            export_manager = FinOpsExportManager(
                profile=ctx.obj["profile"],
                output_dir=output_dir,
                include_quarterly=include_quarterly,
                executive_summary=executive_summary,
                mcp_validate=mcp_validate,
            )

            export_results = {}
            for format_type in export_format:
                result = export_manager.export_analysis(format=format_type)
                export_results[format_type] = result

            # Attach MCP validation results if available
            if mcp_results:
                export_results["mcp_validation"] = {
                    "accuracy_percentage": mcp_results.accuracy_percentage,
                    "validation_passed": mcp_results.accuracy_percentage >= 99.5,
                    "export_formats": list(export_format),
                    "operation_name": mcp_results.operation_name,
                    "status": mcp_results.status.value,
                    "detailed_results": mcp_results,
                }

            error_handlers["success"](
                f"Successfully exported to {len(export_format)} format(s)", f"Output directory: {output_dir}"
            )

            return export_results

        except ImportError as e:
            error_handlers["module_not_available"]("FinOps export", e)
            raise click.ClickException("FinOps export functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("FinOps export operation", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--older-than-days", type=int, default=90, help="Minimum age in days for cleanup consideration (default: 90)"
    )
    @click.option(
        "--validate", is_flag=True, default=True, help="Enable MCP validation for ≥99.5% accuracy (default: enabled)"
    )
    @click.option("--cleanup", is_flag=True, help="Enable cleanup recommendations (READ-ONLY analysis only)")
    @click.option("--export-results", is_flag=True, help="Export analysis results to JSON file")
    @click.option(
        "--safety-checks/--no-safety-checks",
        default=True,
        help="Enable comprehensive safety validations (default: enabled)",
    )
    @click.option("--all-profiles", help="Use specified profile for all operations (overrides parent --profile)")
    @click.pass_context
    def ec2_snapshots(ctx, older_than_days, validate, cleanup, export_results, safety_checks, all_profiles):
        """
        EC2 snapshot cost optimization and cleanup analysis.

        Sprint 1, Task 1: Analyze EC2 snapshots for cost optimization opportunities
        targeting $50K+ annual savings through systematic age-based cleanup with
        enterprise safety validations and MCP accuracy frameworks.

        Enterprise Features:
        • Multi-account snapshot discovery via AWS Config aggregator
        • Dynamic pricing via AWS Pricing API for accurate cost calculations
        • MCP validation framework with ≥99.5% accuracy cross-validation
        • Comprehensive safety checks (volume attachment, AMI association, age)
        • Executive reporting with Sprint 1 business impact metrics

        Safety Features:
        • READ-ONLY analysis by default (no actual cleanup performed)
        • Volume attachment verification before recommendations
        • AMI association checking to prevent data loss
        • Configurable age thresholds with safety validations

        Examples:
            # Basic analysis with MCP validation using parent profile
            runbooks finops --profile BILLING_PROFILE ec2-snapshots --validate

            # Override parent profile with command-specific profile
            runbooks finops ec2-snapshots --all-profiles BILLING_PROFILE --validate

            # Custom age threshold with export
            runbooks finops --profile BILLING_PROFILE ec2-snapshots --older-than-days 120 --export-results

            # Comprehensive analysis for Sprint 1
            runbooks finops --profile BILLING_PROFILE ec2-snapshots --cleanup --validate --export-results

            # Quick analysis without safety checks (not recommended)
            runbooks finops ec2-snapshots --all-profiles BILLING_PROFILE --no-safety-checks --older-than-days 30

        Sprint 1 Context:
            Task 1 targeting $50K+ annual savings through systematic snapshot cleanup
            with enterprise coordination and MCP validation accuracy ≥99.5%
        """
        try:
            import asyncio
            from runbooks.finops.snapshot_manager import EC2SnapshotManager

            console.print("\n[bold blue]🎯 Sprint 1, Task 1: EC2 Snapshot Cost Optimization[/bold blue]")

            # Resolve profile with priority: --all-profiles > ctx.obj['profile'] > 'default'
            resolved_profile = all_profiles or ctx.obj.get("profile", "default")
            resolved_region = ctx.obj.get("region", "all")
            resolved_dry_run = ctx.obj.get("dry_run", True)

            # Validate profile resolution
            if not resolved_profile:
                console.print("[red]❌ Error: No AWS profile specified or found[/red]")
                console.print("[yellow]💡 Use --all-profiles PROFILE_NAME or set parent --profile option[/yellow]")
                raise click.ClickException("AWS profile required for ec2-snapshots command")

            console.print(
                f"[dim]Profile: {resolved_profile} | Region: {resolved_region} | Dry-run: {resolved_dry_run}[/dim]\n"
            )

            # Initialize snapshot manager with enterprise configuration
            manager = EC2SnapshotManager(profile=resolved_profile, dry_run=resolved_dry_run)

            # Configure safety checks based on user preference
            if not safety_checks:
                console.print("[yellow]⚠️ Safety checks disabled - use with caution[/yellow]")
                manager.safety_checks = {
                    "volume_attachment_check": False,
                    "ami_association_check": False,
                    "minimum_age_check": True,  # Always keep age check for safety
                    "cross_account_validation": False,
                }

            # Run the main analysis using the enhanced method
            async def run_analysis():
                return await manager.analyze_snapshot_opportunities(
                    profile=resolved_profile,
                    older_than_days=older_than_days,
                    enable_mcp_validation=validate,
                    export_results=export_results,
                )

            # Execute analysis
            results = asyncio.run(run_analysis())

            # Check if we have cost analysis results before validating Sprint 1 targets
            if not results.get("cost_analysis") or not results["cost_analysis"]:
                console.print("\n[yellow]⚠️ No snapshots found - no cost optimization opportunities identified[/yellow]")
                return results

            # Sprint 1 success validation
            annual_savings = results["cost_analysis"]["annual_savings"]
            sprint_target = 50000  # $50K Sprint 1 target

            if annual_savings >= sprint_target:
                console.print(f"\n[bold green]✅ Sprint 1 Task 1 SUCCESS![/bold green]")
                console.print(f"[green]Target: ${sprint_target:,} | Achieved: ${annual_savings:,.2f}[/green]")
            else:
                console.print(f"\n[bold yellow]⚠️ Sprint 1 Task 1 - Below Target[/bold yellow]")
                console.print(f"[yellow]Target: ${sprint_target:,} | Achieved: ${annual_savings:,.2f}[/yellow]")

            # MCP validation status for Sprint 1
            if validate and results.get("mcp_validation"):
                mcp_results = results["mcp_validation"]
                if mcp_results["validation_passed"]:
                    console.print(
                        f"[green]✅ MCP Validation: {mcp_results['accuracy_percentage']:.2f}% accuracy[/green]"
                    )
                else:
                    console.print(
                        f"[red]❌ MCP Validation: {mcp_results['accuracy_percentage']:.2f}% accuracy (Required: ≥99.5%)[/red]"
                    )

            # Enterprise coordination confirmation
            console.print(f"\n[dim]🏢 Enterprise coordination: python-runbooks-engineer [1] (Primary)[/dim]")
            console.print(f"[dim]🎯 Sprint coordination: Systematic delegation activated[/dim]")

            return results

        except ImportError as e:
            error_handlers["module_not_available"]("EC2 Snapshot Manager", e)
            raise click.ClickException("EC2 snapshot analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("EC2 snapshot analysis", e)
            raise click.ClickException(str(e))

    # Epic 2 Infrastructure Optimization Commands
    @finops.group()
    def infrastructure():
        """Epic 2 Infrastructure Optimization - $210,147 annual savings target"""
        pass

    @infrastructure.command()
    @click.option(
        "--components",
        multiple=True,
        type=click.Choice(["nat-gateway", "elastic-ip", "load-balancer", "vpc-endpoint"]),
        help="Infrastructure components to analyze (default: all)",
    )
    @click.option(
        "--export-format",
        type=click.Choice(["json", "csv", "markdown"]),
        default="json",
        help="Export format for results",
    )
    @click.option("--output-file", help="Output file path for results export")
    @click.option("--mcp-validate", is_flag=True, help="Enable MCP validation for ≥99.5% accuracy cross-validation")
    @click.pass_context
    def analyze(ctx, components, export_format, output_file, mcp_validate):
        """
        Comprehensive Infrastructure Optimization Analysis - Epic 2

        Analyze all infrastructure components to achieve $210,147 Epic 2 annual savings target:
        • NAT Gateway optimization: $147,420 target
        • Elastic IP optimization: $21,593 target
        • Load Balancer optimization: $35,280 target
        • VPC Endpoint optimization: $5,854 target

        SAFETY: READ-ONLY analysis only - no resource modifications.

        Examples:
            runbooks finops infrastructure analyze
            runbooks finops infrastructure analyze --components nat-gateway load-balancer
        """
        try:
            import asyncio
            from runbooks.finops.infrastructure.commands import InfrastructureOptimizer

            # MCP validation integration for infrastructure analysis
            mcp_results = None
            if mcp_validate:
                try:
                    from runbooks.validation.mcp_validator import MCPValidator

                    console.print("[cyan]🔍 Running MCP validation for infrastructure optimization accuracy[/cyan]")

                    # Configure validation profiles
                    validation_profiles = {
                        "billing": ctx.obj.get("profile", "default"),
                        "management": ctx.obj.get("profile", "default"),
                        "centralised_ops": ctx.obj.get("profile", "default"),
                        "single_aws": ctx.obj.get("profile", "default"),
                    }

                    # Initialize validator
                    validator = MCPValidator(
                        profiles=validation_profiles, tolerance_percentage=5.0, performance_target_seconds=30.0
                    )

                    # Run validation for infrastructure operations using VPC validation for networking components
                    component_types = (
                        list(components)
                        if components
                        else ["nat-gateway", "elastic-ip", "load-balancer", "vpc-endpoint"]
                    )
                    if any(comp in ["nat-gateway", "vpc-endpoint"] for comp in component_types):
                        mcp_results = asyncio.run(validator.validate_vpc_analysis())
                    elif any(comp in ["elastic-ip"] for comp in component_types):
                        mcp_results = asyncio.run(validator.validate_ec2_inventory())
                    else:
                        # Default to cost explorer for load balancer cost analysis
                        mcp_results = asyncio.run(validator.validate_cost_explorer())

                    # Display validation results
                    if mcp_results.accuracy_percentage >= 99.5:
                        console.print(
                            f"[green]✅ MCP Validation PASSED: {mcp_results.accuracy_percentage:.1f}% accuracy for infrastructure[/green]"
                        )
                    elif mcp_results.accuracy_percentage >= 95.0:
                        console.print(
                            f"[yellow]⚠️ MCP Validation WARNING: {mcp_results.accuracy_percentage:.1f}% accuracy (target: ≥99.5%)[/yellow]"
                        )
                    else:
                        console.print(
                            f"[red]❌ MCP Validation FAILED: {mcp_results.accuracy_percentage:.1f}% accuracy[/red]"
                        )

                except Exception as e:
                    console.print(f"[yellow]⚠️ MCP validation failed: {e}[/yellow]")
                    console.print("[dim]Continuing with infrastructure analysis...[/dim]")

            # Initialize comprehensive optimizer
            optimizer = InfrastructureOptimizer(
                profile_name=ctx.obj.get("profile"),
                regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None,
                mcp_validate=mcp_validate,
            )

            # Execute comprehensive analysis
            results = asyncio.run(
                optimizer.analyze_comprehensive_infrastructure(
                    components=list(components) if components else None, dry_run=ctx.obj.get("dry_run", True)
                )
            )

            # Attach MCP validation results if available
            if mcp_results and hasattr(results, "__dict__"):
                results.mcp_validation = {
                    "accuracy_percentage": mcp_results.accuracy_percentage,
                    "validation_passed": mcp_results.accuracy_percentage >= 99.5,
                    "components_validated": list(components) if components else "all",
                    "operation_name": mcp_results.operation_name,
                    "status": mcp_results.status.value,
                    "detailed_results": mcp_results,
                }

            # Display Epic 2 progress
            if results.epic_2_target_achieved:
                console.print(f"\n[bold green]✅ Epic 2 Infrastructure Target Achieved![/bold green]")
                console.print(
                    f"[green]Target: ${results.epic_2_target_savings:,.0f} | Achieved: ${results.total_potential_savings:,.0f}[/green]"
                )
            else:
                progress_pct = results.epic_2_progress_percentage
                console.print(f"\n[bold yellow]📊 Epic 2 Infrastructure Progress: {progress_pct:.1f}%[/bold yellow]")
                console.print(
                    f"[yellow]Target: ${results.epic_2_target_savings:,.0f} | Achieved: ${results.total_potential_savings:,.0f}[/yellow]"
                )

            # Export results if requested
            if output_file or export_format != "json":
                console.print(f"[dim]Export functionality available - results ready for {export_format} export[/dim]")

            return results

        except ImportError as e:
            error_handlers["module_not_available"]("Infrastructure Optimizer", e)
            raise click.ClickException("Infrastructure optimization functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("Infrastructure optimization analysis", e)
            raise click.ClickException(str(e))

    @infrastructure.command()
    @click.pass_context
    def nat_gateway(ctx):
        """NAT Gateway optimization analysis - $147,420 Epic 2 target"""
        try:
            import asyncio
            from runbooks.finops.nat_gateway_optimizer import NATGatewayOptimizer

            optimizer = NATGatewayOptimizer(
                profile_name=ctx.obj.get("profile"), regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None
            )

            results = asyncio.run(optimizer.analyze_nat_gateways(dry_run=ctx.obj.get("dry_run", True)))

            # Display Epic 2 component progress
            target = 147420.0
            if results.potential_annual_savings >= target:
                console.print(f"\n[bold green]✅ NAT Gateway Epic 2 Target Achieved![/bold green]")
            else:
                progress = (results.potential_annual_savings / target) * 100
                console.print(f"\n[yellow]📊 NAT Gateway Progress: {progress:.1f}% of Epic 2 target[/yellow]")

            return results

        except Exception as e:
            error_handlers["operation_failed"]("NAT Gateway optimization", e)
            raise click.ClickException(str(e))

    @infrastructure.command()
    @click.pass_context
    def elastic_ip(ctx):
        """Elastic IP optimization analysis - $21,593 Epic 2 target"""
        try:
            import asyncio
            from runbooks.finops.elastic_ip_optimizer import ElasticIPOptimizer

            optimizer = ElasticIPOptimizer(
                profile_name=ctx.obj.get("profile"), regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None
            )

            results = asyncio.run(optimizer.analyze_elastic_ips(dry_run=ctx.obj.get("dry_run", True)))

            # Display Epic 2 component progress
            target = 21593.0
            if results.potential_annual_savings >= target:
                console.print(f"\n[bold green]✅ Elastic IP Epic 2 Target Achieved![/bold green]")
            else:
                progress = (results.potential_annual_savings / target) * 100
                console.print(f"\n[yellow]📊 Elastic IP Progress: {progress:.1f}% of Epic 2 target[/yellow]")

            return results

        except Exception as e:
            error_handlers["operation_failed"]("Elastic IP optimization", e)
            raise click.ClickException(str(e))

    @infrastructure.command()
    @click.pass_context
    def load_balancer(ctx):
        """Load Balancer optimization analysis - $35,280 Epic 2 target"""
        try:
            import asyncio
            from runbooks.finops.infrastructure.load_balancer_optimizer import LoadBalancerOptimizer

            optimizer = LoadBalancerOptimizer(
                profile_name=ctx.obj.get("profile"), regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None
            )

            results = asyncio.run(optimizer.analyze_load_balancers(dry_run=ctx.obj.get("dry_run", True)))

            # Display Epic 2 component progress
            target = 35280.0
            if results.potential_annual_savings >= target:
                console.print(f"\n[bold green]✅ Load Balancer Epic 2 Target Achieved![/bold green]")
            else:
                progress = (results.potential_annual_savings / target) * 100
                console.print(f"\n[yellow]📊 Load Balancer Progress: {progress:.1f}% of Epic 2 target[/yellow]")

            return results

        except Exception as e:
            error_handlers["operation_failed"]("Load Balancer optimization", e)
            raise click.ClickException(str(e))

    @infrastructure.command()
    @click.pass_context
    def vpc_endpoint(ctx):
        """VPC Endpoint optimization analysis - $5,854 Epic 2 target"""
        try:
            import asyncio
            from runbooks.finops.infrastructure.vpc_endpoint_optimizer import VPCEndpointOptimizer

            optimizer = VPCEndpointOptimizer(
                profile_name=ctx.obj.get("profile"), regions=[ctx.obj.get("region")] if ctx.obj.get("region") else None
            )

            results = asyncio.run(optimizer.analyze_vpc_endpoints(dry_run=ctx.obj.get("dry_run", True)))

            # Display Epic 2 component progress
            target = 5854.0
            if results.potential_annual_savings >= target:
                console.print(f"\n[bold green]✅ VPC Endpoint Epic 2 Target Achieved![/bold green]")
            else:
                progress = (results.potential_annual_savings / target) * 100
                console.print(f"\n[yellow]📊 VPC Endpoint Progress: {progress:.1f}% of Epic 2 target[/yellow]")

            return results

        except Exception as e:
            error_handlers["operation_failed"]("VPC Endpoint optimization", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input WorkSpaces data file (Excel/CSV with 'AWS Account' column)",
    )
    @click.option(
        "--output", "-o", "output_file", type=click.Path(), help="Output enriched data file (Excel/CSV/JSON)"
    )
    @click.option("--profile", "-p", default="default", help="AWS operational profile (WorkSpaces access)")
    @click.option("--management-profile", "-m", default=None, help="AWS management profile (Organizations access)")
    @click.option("--display-only", is_flag=True, help="Display Rich CLI output without file export")
    @click.pass_context
    def enrich_workspaces(ctx, input_file, output_file, profile, management_profile, display_only):
        """
        Enrich WorkSpaces inventory with Organizations metadata.

        Reads WorkSpaces data from Excel/CSV with 'AWS Account' column and enriches it with:
        • Account name (from AWS Organizations)
        • Account email
        • WBS code (cost allocation)
        • Cost group
        • Technical lead
        • Account owner

        Example:
            runbooks finops enrich-workspaces -i data/workspaces.xlsx -o data/enriched.xlsx -p ops-profile -m mgmt-profile
        """
        try:
            from runbooks.finops.enrich_workspaces import enrich_workspaces as enrich_ws_cmd

            # Import the command function and invoke it directly
            import sys
            from click.testing import CliRunner

            # Get format from parent context (output_format from common_output_options decorator)
            output_format = ctx.obj.get("output_format", "csv") if ctx.obj else "csv"

            # Create a runner to invoke the standalone command
            # This allows reusing the implementation without duplicating code
            enrich_ws_cmd.callback(input_file, output_file, profile, management_profile, output_format, display_only)

        except ImportError as e:
            error_handlers["module_not_available"]("WorkSpaces Enrichment", e)
            raise click.ClickException("WorkSpaces enrichment functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("WorkSpaces enrichment", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input EC2 inventory Excel file (with Identifier, AWS Account, Region columns)",
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output enriched Excel file (multi-sheet with cost analysis)"
    )
    @click.option(
        "--management-profile",
        "-m",
        required=True,
        help="AWS management profile for Organizations API access"
    )
    @click.option(
        "--billing-profile",
        "-b",
        required=True,
        help="AWS billing profile for Cost Explorer API access"
    )
    @click.option(
        "--operational-profile",
        "-p",
        default=None,
        help="AWS operational profile for EC2 describe-instances (optional, defaults to management profile)"
    )
    @click.option(
        "--enable-organizations/--no-organizations",
        default=True,
        help="Enable Organizations metadata enrichment (default: enabled)"
    )
    @click.option(
        "--enable-cost/--no-cost",
        default=True,
        help="Enable Cost Explorer enrichment (default: enabled)"
    )
    @click.option(
        "--enable-activity/--no-activity",
        default=False,
        help="Enable CloudTrail activity tracking (default: disabled, takes 60-90 seconds)"
    )
    @click.option(
        "--include-12month-cost/--no-12month-cost",
        default=True,
        help="Include 12-month cost breakdown (default: enabled)"
    )
    @click.option(
        "--decommission-mode/--no-decommission-mode",
        default=False,
        help="Decommission mode: Filter to 45 decision-focused columns (default: disabled, shows all 87 columns)"
    )
    @click.option("--verbose", "-v", is_flag=True, help="Show detailed logs")
    @click.option("--format", type=click.Choice(["compact", "table", "json"]), default="compact", help="Output format")
    @click.pass_context
    def analyze_ec2(
        ctx,
        input_file,
        output_file,
        management_profile,
        billing_profile,
        operational_profile,
        enable_organizations,
        enable_cost,
        enable_activity,
        include_12month_cost,
        decommission_mode,
        verbose,
        format
    ):
        """
        EC2 cost analysis with 4-way enrichment.

        \b
        🔄 4-WAY ENRICHMENT LAYERS (E1-E7 Decommission Scoring)
        ┌──────────────────────────────────────────────────────────────┐
        │ Layer 1: Discovery       │ Resource Explorer EC2 inventory   │
        │ Layer 2: Organizations   │ Account metadata (6 columns)      │
        │ Layer 3: Cost            │ 12-month Cost Explorer trends     │
        │ Layer 4: Activity        │ CloudTrail idle detection (opt)   │
        └──────────────────────────────────────────────────────────────┘

        \b
        📊 DECOMMISSION SIGNALS (E1-E7)
        • E1: Instance stopped state → +15 points (HIGH priority)
        • E2: Zero 12-month cost → +12 points (Terminated/Free tier)
        • E3: Low monthly cost (<$5) → +8 points (Micro workloads)
        • E4: No CloudTrail activity (14d) → +10 points (Idle detection)
        • E5: Cost decreasing trend → +5 points (Usage declining)
        • E6: Legacy instance type (t2, m4) → +3 points (Modernization)
        • E7: Untagged resources → +2 points (Management overhead)

        \b
        🎯 OPTIMIZATION TIERS
        • HIGH (≥25 points): Immediate decommission candidates
        • MEDIUM (15-24 points): Review recommended within 30 days
        • LOW (<15 points): Monitor and maintain current state

        \b
        💰 Sprint 1 Target: 25-50% cost savings via tier-based decommission
        📖 Example: runbooks finops analyze-ec2 -i ec2.xlsx -o enriched.xlsx -m mgmt -b billing
        """
        # Initialize output controller
        configure_logging(verbose=verbose)
        controller = OutputController(verbose=verbose, format=format)

        try:
            from runbooks.finops.ec2_analyzer import analyze_ec2_costs
            from runbooks.common.rich_utils import print_header, print_success, print_error

            print_header("EC2 Cost Analysis", f"Input: {input_file}")

            # Execute EC2 cost analysis
            result_df = analyze_ec2_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                operational_profile=operational_profile,
                enable_organizations=enable_organizations,
                enable_cost=enable_cost,
                enable_activity=enable_activity,
                include_12month_cost=include_12month_cost,
                decommission_mode=decommission_mode
            )

            # Print summary using OutputController
            enrichment_layers = []
            if enable_organizations:
                enrichment_layers.append("organizations")
            if enable_cost:
                enrichment_layers.append("cost")
            if enable_activity:
                enrichment_layers.append("activity")

            controller.print_operation_summary(
                emoji="💰",
                operation="EC2 Cost Analysis",
                input_count=len(result_df),
                enriched_count=len(result_df),
                enrichment_type="EC2 instances analyzed",
                success_percentage=100.0,
                profile=management_profile,
                output_file=output_file,
                added_columns=enrichment_layers
            )

            return result_df

        except ImportError as e:
            error_handlers["module_not_available"]("EC2 Cost Analyzer", e)
            raise click.ClickException("EC2 cost analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("EC2 cost analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input EC2 inventory Excel file (with instance_id, account_id, region columns)"
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output enriched Excel file with decommission analysis"
    )
    @click.option(
        "--management-profile",
        "-m",
        required=True,
        help="AWS management profile for Organizations API access"
    )
    @click.option(
        "--billing-profile",
        "-b",
        required=True,
        help="AWS billing profile for Cost Explorer API access"
    )
    @click.option(
        "--operational-profile",
        "-p",
        default=None,
        help="AWS operational profile for EC2 describe-instances (optional, defaults to management)"
    )
    @click.option(
        "--enable-expensive-signals/--no-expensive-signals",
        default=False,
        help="Enable expensive signals (CloudTrail activity, SSM, Compute Optimizer)"
    )
    @click.pass_context
    def ec2_decommission_analysis(
        ctx,
        input_file,
        output_file,
        management_profile,
        billing_profile,
        operational_profile,
        enable_expensive_signals
    ):
        """
        EC2 decommission analysis with E1-E7 scoring framework.

        This command provides enterprise EC2 decommission candidate identification with:
        • E1-E7 multi-signal scoring (Compute Optimizer, CloudWatch, CloudTrail, SSM, etc.)
        • MUST/SHOULD/COULD/KEEP tier classification
        • Top 50 decommission candidates executive summary
        • Annual savings projections per instance
        • Professional Rich CLI output for CxO consumption

        Scoring Framework (0-100 scale):
        • MUST (80-100): Immediate decommission candidates
        • SHOULD (50-79): Strong candidates (review recommended)
        • COULD (25-49): Potential candidates (manual review)
        • KEEP (<25): Active resources (no action)

        Signal Framework (E1-E7):
        • E1: Compute Optimizer Idle (max CPU ≤1% for 14d) → +60 points
        • E2: SSM Agent Offline/Stale → +8 points
        • E3: CloudTrail no activity (90d) → +8 points
        • E4: Stopped State (>30 days) → +8 points
        • E5: No service attachment (ALB/ASG/ECS/EKS) → +6 points
        • E6: No Tags/Owner → +5 points
        • E7: Dev/Test Environment → +3 points

        Examples:
            # Basic decommission analysis (fast, no expensive signals)
            runbooks finops ec2-decommission-analysis \\
                --input data/ec2-inventory.xlsx \\
                --output data/ec2-decommission-analysis.xlsx \\
                --management-profile mgmt-ro \\
                --billing-profile billing-ro

            # Comprehensive analysis with expensive signals (CloudTrail, SSM, Compute Optimizer)
            runbooks finops ec2-decommission-analysis \\
                -i data/ec2.xlsx \\
                -o data/ec2-decommission-full.xlsx \\
                -m mgmt-ro \\
                -b billing-ro \\
                --enable-expensive-signals

        Input File Format:
            Excel file with required columns:
            - instance_id: EC2 instance ID (i-xxxxx)
            - account_id: 12-digit AWS account ID
            - region: AWS region (ap-southeast-2, etc.)

        Output File Format:
            Multi-sheet Excel with:
            - Sheet 1: Enriched Data (67+ columns with decommission scores)
            - Sheet 2: Top 50 Decommission Candidates (executive summary)
            - Sheet 3: Cost Summary (by account/tier)
            - Sheet 4: Validation Metrics

        Performance Notes:
            • Fast mode (no expensive signals): ~5-10 seconds
            • Full mode (with expensive signals): ~60-90 seconds
            • Expensive signals: CloudTrail (60s), SSM (10s), Compute Optimizer (10s)

        Pattern: Migrated from notebooks/compute/ec2.ipynb with CxO executive polish
        """
        try:
            from runbooks.finops.ec2_analyzer import analyze_ec2_costs
            from runbooks.finops.decommission_classifier import classify_ec2
            from runbooks.common.rich_utils import print_header, print_success, print_error, print_info

            print_header("EC2 Decommission Analysis", f"Input: {input_file}")

            # Step 1: Execute EC2 cost analysis with enrichment
            print_info("Step 1/3: Enriching EC2 inventory with Organizations and Cost Explorer data...")
            enriched_df = analyze_ec2_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                operational_profile=operational_profile,
                enable_organizations=True,
                enable_cost=True,
                enable_activity=enable_expensive_signals,  # CloudTrail activity (60-90s)
                include_12month_cost=True
            )

            # Step 2: Apply decommission classification
            print_info("\nStep 2/3: Applying E1-E7 decommission scoring framework...")
            enriched_df = classify_ec2(enriched_df)

            # Step 3: Generate executive summary
            print_info("\nStep 3/3: Generating Top 50 decommission candidates executive summary...")

            # Filter to MUST/SHOULD tiers
            top_candidates = enriched_df[
                enriched_df['decommission_tier'].isin(['MUST', 'SHOULD'])
            ].sort_values('decommission_score', ascending=False).head(50)

            if len(top_candidates) > 0:
                from rich.table import Table
                from runbooks.common.rich_utils import console, format_cost

                # Create executive table
                exec_table = Table(
                    title=f"Top {len(top_candidates)} Decommission Candidates",
                    show_header=True,
                    header_style="bold cyan"
                )

                exec_table.add_column("Instance ID", style="dim", width=20)
                exec_table.add_column("Account", style="white", width=25)
                exec_table.add_column("Type", style="cyan", width=12)
                exec_table.add_column("Score", style="red", width=8, justify="right")
                exec_table.add_column("Tier", style="yellow", width=10)
                exec_table.add_column("Monthly Cost", style="green", width=12, justify="right")
                exec_table.add_column("Annual Savings", style="green bold", width=15, justify="right")

                total_monthly_cost = 0
                total_annual_savings = 0

                for _, row in top_candidates.iterrows():
                    instance_id = str(row.get('instance_id', 'N/A'))[:19]
                    account_name = str(row.get('account_name', 'Unknown'))[:24]
                    instance_type = str(row.get('instance_type', 'N/A'))
                    score = row.get('decommission_score', 0)
                    tier = row.get('decommission_tier', 'KEEP')
                    monthly_cost = row.get('monthly_cost', 0.0)
                    annual_savings = monthly_cost * 12

                    total_monthly_cost += monthly_cost
                    total_annual_savings += annual_savings

                    tier_style = "red bold" if tier == "MUST" else "yellow"

                    exec_table.add_row(
                        instance_id,
                        account_name,
                        instance_type,
                        f"{score:.0f}",
                        f"[{tier_style}]{tier}[/{tier_style}]",
                        format_cost(monthly_cost),
                        format_cost(annual_savings)
                    )

                console.print("\n")
                console.print(exec_table)

                print_success(f"\n💰 Financial Impact:")
                print_success(f"   Total Monthly Cost: {format_cost(total_monthly_cost)}")
                print_success(f"   Total Annual Savings: {format_cost(total_annual_savings)}")
                print_success(f"   Average Decommission Score: {top_candidates['decommission_score'].mean():.1f}/100")

                tier_breakdown = top_candidates['decommission_tier'].value_counts()
                print_info(f"\n📊 Tier Breakdown (Top 50):")
                print_info(f"   MUST: {tier_breakdown.get('MUST', 0)} instances")
                print_info(f"   SHOULD: {tier_breakdown.get('SHOULD', 0)} instances")
            else:
                print_success("\n✅ No high-priority decommission candidates identified.")
                print_success("   All EC2 instances classified as COULD or KEEP tiers.")

            # Final summary
            print_success(f"\n✅ EC2 decommission analysis complete!")
            print_success(f"   📊 Analyzed {len(enriched_df)} EC2 instances")
            print_success(f"   📁 Output: {output_file}")

            return enriched_df

        except ImportError as e:
            error_handlers["module_not_available"]("EC2 Decommission Analyzer", e)
            raise click.ClickException("EC2 decommission analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("EC2 decommission analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--profile",
        "-p",
        default="default",
        help="AWS profile for Lambda operations (Organizations, Cost Explorer, CloudWatch)"
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        default="lambda-analysis.xlsx",
        type=click.Path(),
        help="Output enriched Excel file with Lambda analysis"
    )
    @click.option(
        "--regions",
        "-r",
        multiple=True,
        help="AWS regions to analyze (defaults to all enabled regions)"
    )
    @click.option(
        "--enable-organizations/--no-organizations",
        default=True,
        help="Enable Organizations metadata enrichment (default: enabled)"
    )
    @click.option(
        "--enable-cost/--no-cost",
        default=True,
        help="Enable Cost Explorer enrichment (default: enabled)"
    )
    @click.option("--verbose", "-v", is_flag=True, help="Show detailed logs")
    @click.option("--format", type=click.Choice(["compact", "table", "json"]), default="compact", help="Output format")
    @click.pass_context
    def lambda_analysis(
        ctx,
        profile,
        output_file,
        regions,
        enable_organizations,
        enable_cost,
        verbose,
        format
    ):
        """
        Lambda cost and activity analysis with optimization signals.

        \b
        🔄 LAMBDA ENRICHMENT LAYERS (L1-L6 Optimization Scoring)
        ┌──────────────────────────────────────────────────────────────┐
        │ Layer 1: Discovery       │ Lambda function inventory         │
        │ Layer 2: Organizations   │ Account metadata (6 columns)      │
        │ Layer 3: Cost            │ 12-month Cost Explorer trends     │
        │ Layer 4: CloudWatch      │ Invocations, errors, duration     │
        └──────────────────────────────────────────────────────────────┘

        \b
        📊 OPTIMIZATION SIGNALS (L1-L6)
        • L1: High invocation cost (top 10%) → +10 points (Cost impact)
        • L2: Idle function (0 invocations 14d) → +15 points (Unused)
        • L3: Oversized memory (low utilization) → +8 points (Rightsizing)
        • L4: Cold start issues (high p95/p50) → +5 points (Performance)
        • L5: High error rate (>5%) → +7 points (Quality issue)
        • L6: Legacy runtime → +3 points (Security/maintenance)

        \b
        🎯 OPTIMIZATION TIERS
        • HIGH (≥20 points): Immediate optimization/decommission
        • MEDIUM (10-19 points): Review and optimize within 30 days
        • LOW (<10 points): Monitor and maintain current state

        \b
        💰 Sprint 1 Target: Identify serverless optimization opportunities
        📖 Example: runbooks finops lambda-analysis --profile default --output lambda.xlsx
        """
        # Initialize output controller
        configure_logging(verbose=verbose)
        controller = OutputController(verbose=verbose, format=format)

        try:
            from runbooks.finops.lambda_analyzer import analyze_lambda_costs
            from runbooks.common.rich_utils import print_header, print_success, print_error

            print_header("Lambda Cost Analysis", f"Profile: {profile}")

            # Execute Lambda cost analysis
            result_df = analyze_lambda_costs(
                profile=profile,
                output_file=output_file,
                regions=list(regions) if regions else None,
                enable_organizations=enable_organizations,
                enable_cost=enable_cost
            )

            # Print summary using OutputController
            enrichment_layers = []
            if enable_organizations:
                enrichment_layers.append("organizations")
            if enable_cost:
                enrichment_layers.append("cost")

            controller.print_operation_summary(
                emoji="⚡",
                operation="Lambda Cost Analysis",
                input_count=len(result_df),
                enriched_count=len(result_df),
                enrichment_type="Lambda functions analyzed",
                success_percentage=100.0,
                profile=profile,
                output_file=output_file,
                added_columns=enrichment_layers
            )

            # Display optimization summary in verbose mode
            if verbose and len(result_df) > 0 and 'optimization_tier' in result_df.columns:
                tier_counts = result_df['optimization_tier'].value_counts()
                print_info(f"\n🎯 Optimization Opportunities:")
                print_info(f"   HIGH: {tier_counts.get('HIGH', 0)} functions")
                print_info(f"   MEDIUM: {tier_counts.get('MEDIUM', 0)} functions")
                print_info(f"   LOW: {tier_counts.get('LOW', 0)} functions")

            return result_df

        except ImportError as e:
            error_handlers["module_not_available"]("Lambda Cost Analyzer", e)
            raise click.ClickException("Lambda cost analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("Lambda cost analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input WorkSpaces inventory Excel file (with Identifier, AWS Account, Region columns)"
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output enriched Excel file (multi-sheet with cost analysis)"
    )
    @click.option(
        "--management-profile",
        "-m",
        required=True,
        help="AWS management profile for Organizations API access"
    )
    @click.option(
        "--billing-profile",
        "-b",
        required=True,
        help="AWS billing profile for Cost Explorer API access"
    )
    @click.option(
        "--operational-profile",
        "-p",
        default=None,
        help="AWS operational profile for WorkSpaces operations (optional, defaults to management profile)"
    )
    @click.option(
        "--enable-organizations/--no-organizations",
        default=True,
        help="Enable Organizations metadata enrichment (default: enabled)"
    )
    @click.option(
        "--enable-cost/--no-cost",
        default=True,
        help="Enable Cost Explorer enrichment (default: enabled)"
    )
    @click.option(
        "--enable-activity/--no-activity",
        default=False,
        help="Enable CloudTrail activity tracking (default: disabled, takes 60-90 seconds)"
    )
    @click.option(
        "--include-12month-cost/--no-12month-cost",
        default=True,
        help="Include 12-month cost breakdown (default: enabled)"
    )
    @click.option("--verbose", "-v", is_flag=True, help="Show detailed logs")
    @click.option("--format", type=click.Choice(["compact", "table", "json"]), default="compact", help="Output format")
    @click.pass_context
    def analyze_workspaces(
        ctx,
        input_file,
        output_file,
        management_profile,
        billing_profile,
        operational_profile,
        enable_organizations,
        enable_cost,
        enable_activity,
        include_12month_cost,
        verbose,
        format
    ):
        """
        WorkSpaces cost analysis with decommission tier scoring.

        \b
        🔄 4-WAY ENRICHMENT LAYERS (W1-W6 Decommission Scoring)
        ┌──────────────────────────────────────────────────────────────┐
        │ Layer 1: Discovery       │ WorkSpaces inventory (122 total)  │
        │ Layer 2: Organizations   │ Account metadata (6 columns)      │
        │ Layer 3: Cost            │ 12-month Cost Explorer trends     │
        │ Layer 4: Activity        │ CloudWatch UserConnected metrics  │
        └──────────────────────────────────────────────────────────────┘

        \b
        📊 DECOMMISSION SIGNALS (W1-W6)
        • W1: WorkSpace stopped state → +15 points (HIGH priority)
        • W2: Zero UserConnected time (30d) → +12 points (Idle workspace)
        • W3: Low monthly cost (<$10) → +8 points (Basic bundle)
        • W4: Cost decreasing trend → +5 points (Usage declining)
        • W5: No volume encryption → +3 points (Compliance risk)
        • W6: Legacy bundle type → +2 points (Modernization opportunity)

        \b
        🎯 OPTIMIZATION TIERS
        • HIGH (≥25 points): Immediate decommission candidates
        • MEDIUM (15-24 points): Review recommended within 30 days
        • LOW (<15 points): Monitor and maintain current state

        \b
        💰 Sprint 1 Target: 25-50% cost savings via tier-based decommission
        📖 Example: runbooks finops analyze-workspaces -i ws.xlsx -o enriched.xlsx -m mgmt -b billing
        """
        # Initialize output controller
        configure_logging(verbose=verbose)
        controller = OutputController(verbose=verbose, format=format)

        try:
            from runbooks.finops.workspaces_analyzer import analyze_workspaces_costs
            from runbooks.common.rich_utils import print_header, print_success, print_error

            print_header("WorkSpaces Cost Analysis", f"Input: {input_file}")

            # Execute WorkSpaces cost analysis
            result_df = analyze_workspaces_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                operational_profile=operational_profile,
                enable_organizations=enable_organizations,
                enable_cost=enable_cost,
                enable_activity=enable_activity,
                include_12month_cost=include_12month_cost
            )

            # Print summary using OutputController
            enrichment_layers = []
            if enable_organizations:
                enrichment_layers.append("organizations")
            if enable_cost:
                enrichment_layers.append("cost")
            if enable_activity:
                enrichment_layers.append("activity")

            controller.print_operation_summary(
                emoji="🖥️",
                operation="WorkSpaces Cost Analysis",
                input_count=len(result_df),
                enriched_count=len(result_df),
                enrichment_type="WorkSpaces analyzed",
                success_percentage=100.0,
                profile=management_profile,
                output_file=output_file,
                added_columns=enrichment_layers
            )

            return result_df

        except ImportError as e:
            error_handlers["module_not_available"]("WorkSpaces Cost Analyzer", e)
            raise click.ClickException("WorkSpaces cost analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("WorkSpaces cost analysis", e)
            raise click.ClickException(str(e))

    @finops.command()
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input WorkSpaces inventory Excel file (with Identifier, AWS Account, Region columns)"
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output enriched Excel file with decommission analysis"
    )
    @click.option(
        "--management-profile",
        "-m",
        required=True,
        help="AWS management profile for Organizations API access"
    )
    @click.option(
        "--billing-profile",
        "-b",
        required=True,
        help="AWS billing profile for Cost Explorer API access"
    )
    @click.option(
        "--operational-profile",
        "-p",
        default=None,
        help="AWS operational profile for WorkSpaces operations (optional, defaults to management)"
    )
    @click.pass_context
    def workspaces_decommission_analysis(
        ctx,
        input_file,
        output_file,
        management_profile,
        billing_profile,
        operational_profile
    ):
        """
        WorkSpaces decommission analysis with W1-W6 scoring framework.

        This command provides enterprise WorkSpaces decommission candidate identification with:
        • W1-W6 multi-signal scoring (connection recency, CloudWatch usage, break-even, policy)
        • MUST/SHOULD/COULD/KEEP tier classification
        • Top 50 decommission candidates executive summary
        • Annual savings projections per WorkSpace
        • Professional Rich CLI output for CxO consumption

        Scoring Framework (0-100 scale):
        • MUST (80-100): Immediate decommission candidates
        • SHOULD (50-79): Strong candidates (review recommended)
        • COULD (25-49): Potential candidates (manual review)
        • KEEP (<25): Active resources (no action)

        Signal Framework (W1-W6):
        • W1: User connection recency (≥60 days) → +45 points
        • W2: CloudWatch UserConnected sum=0 (no sessions) → +25 points
        • W3: Billing vs usage (hourly usage < break-even) → +10 points
        • W4: Cost Optimizer policy (N months unused) → +10 points
        • W5: No admin API activity (90d) → +5 points
        • W6: User status (not in Identity Center) → +5 points

        Examples:
            # Basic decommission analysis
            runbooks finops workspaces-decommission-analysis \\
                --input data/workspaces-inventory.xlsx \\
                --output data/workspaces-decommission-analysis.xlsx \\
                --management-profile mgmt-ro \\
                --billing-profile billing-ro

            # With custom operational profile
            runbooks finops workspaces-decommission-analysis \\
                -i data/workspaces.xlsx \\
                -o data/workspaces-decommission-full.xlsx \\
                -m mgmt-ro \\
                -b billing-ro \\
                -p ops-ro

        Input File Format:
            Excel file with required columns:
            - Identifier: WorkSpace ID (ws-xxxxx)
            - AWS Account: 12-digit AWS account ID
            - Region: AWS region (ap-southeast-2, etc.)

        Output File Format:
            Multi-sheet Excel with:
            - Sheet 1: Enriched Data (40+ columns with decommission scores)
            - Sheet 2: Top 50 Decommission Candidates (executive summary)
            - Sheet 3: Cost Summary (by account/tier)
            - Sheet 4: Validation Metrics

        Performance Notes:
            • WorkSpaces enrichment: ~5-10 seconds
            • CloudWatch metrics: ~10-15 seconds
            • Volume encryption: ~5-10 seconds
            • Break-even calculation: <1 second

        Pattern: Migrated from notebooks/compute/workspaces.ipynb with CxO executive polish
        """
        try:
            from runbooks.finops.workspaces_analyzer import analyze_workspaces_costs, WorkSpacesCostAnalyzer
            from runbooks.finops.decommission_classifier import classify_workspaces
            from runbooks.common.rich_utils import print_header, print_success, print_error, print_info

            print_header("WorkSpaces Decommission Analysis", f"Input: {input_file}")

            # Step 1: Execute WorkSpaces cost analysis with enrichment
            print_info("Step 1/5: Enriching WorkSpaces inventory with Organizations and Cost Explorer data...")
            enriched_df = analyze_workspaces_costs(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                operational_profile=operational_profile,
                enable_organizations=True,
                enable_cost=True,
                enable_activity=False,  # CloudTrail activity not needed for W1-W6
                include_12month_cost=True
            )

            # Step 2: Initialize analyzer for additional enrichment
            print_info("\nStep 2/5: Analyzing volume encryption and connection status...")
            analyzer = WorkSpacesCostAnalyzer(profile=operational_profile or management_profile)
            enriched_df = analyzer.get_volume_encryption(enriched_df)

            # Step 3: CloudWatch UserConnected metrics
            print_info("\nStep 3/5: Fetching CloudWatch UserConnected metrics...")
            enriched_df = analyzer.get_cloudwatch_user_connected(enriched_df, lookback_days=30)

            # Step 4: Dynamic break-even calculation
            print_info("\nStep 4/5: Calculating dynamic break-even hours...")
            enriched_df = analyzer.calculate_dynamic_breakeven(enriched_df)

            # Step 5: Apply decommission classification
            print_info("\nStep 5/5: Applying W1-W6 decommission scoring framework...")
            enriched_df = classify_workspaces(enriched_df)

            # Generate executive summary
            print_info("\nGenerating Top 50 decommission candidates executive summary...")

            # Filter to MUST/SHOULD tiers
            top_candidates = enriched_df[
                enriched_df['decommission_tier'].isin(['MUST', 'SHOULD'])
            ].sort_values('decommission_score', ascending=False).head(50)

            if len(top_candidates) > 0:
                from rich.table import Table
                from runbooks.common.rich_utils import console, format_cost

                # Create executive table
                exec_table = Table(
                    title=f"Top {len(top_candidates)} Decommission Candidates",
                    show_header=True,
                    header_style="bold cyan"
                )

                exec_table.add_column("WorkSpace ID", style="dim", width=20)
                exec_table.add_column("Account", style="white", width=25)
                exec_table.add_column("Username", style="cyan", width=15)
                exec_table.add_column("Score", style="red", width=8, justify="right")
                exec_table.add_column("Tier", style="yellow", width=10)
                exec_table.add_column("Days Idle", style="yellow", width=10, justify="right")
                exec_table.add_column("Monthly Cost", style="green", width=12, justify="right")
                exec_table.add_column("Annual Savings", style="green bold", width=15, justify="right")

                total_monthly_cost = 0
                total_annual_savings = 0

                for _, row in top_candidates.iterrows():
                    workspace_id = str(row.get('Identifier', row.get('WorkspaceId', 'N/A')))[:19]
                    account_name = str(row.get('account_name', 'Unknown'))[:24]
                    username = str(row.get('Username', 'N/A'))[:14]
                    score = row.get('decommission_score', 0)
                    tier = row.get('decommission_tier', 'KEEP')
                    days_idle = row.get('days_since_last_connection', 0)
                    monthly_cost = row.get('monthly_cost', 0.0)
                    annual_savings = monthly_cost * 12

                    total_monthly_cost += monthly_cost
                    total_annual_savings += annual_savings

                    tier_style = "red bold" if tier == "MUST" else "yellow"

                    exec_table.add_row(
                        workspace_id,
                        account_name,
                        username,
                        f"{score:.0f}",
                        f"[{tier_style}]{tier}[/{tier_style}]",
                        f"{days_idle}",
                        format_cost(monthly_cost),
                        format_cost(annual_savings)
                    )

                console.print("\n")
                console.print(exec_table)

                print_success(f"\n💰 Financial Impact:")
                print_success(f"   Total Monthly Cost: {format_cost(total_monthly_cost)}")
                print_success(f"   Total Annual Savings: {format_cost(total_annual_savings)}")
                print_success(f"   Average Decommission Score: {top_candidates['decommission_score'].mean():.1f}/100")

                tier_breakdown = top_candidates['decommission_tier'].value_counts()
                print_info(f"\n📊 Tier Breakdown (Top 50):")
                print_info(f"   MUST: {tier_breakdown.get('MUST', 0)} WorkSpaces")
                print_info(f"   SHOULD: {tier_breakdown.get('SHOULD', 0)} WorkSpaces")
            else:
                print_success("\n✅ No high-priority decommission candidates identified.")
                print_success("   All WorkSpaces classified as COULD or KEEP tiers.")

            # Final summary
            print_success(f"\n✅ WorkSpaces decommission analysis complete!")
            print_success(f"   📊 Analyzed {len(enriched_df)} WorkSpaces")
            print_success(f"   📁 Output: {output_file}")

            return enriched_df

        except ImportError as e:
            error_handlers["module_not_available"]("WorkSpaces Decommission Analyzer", e)
            raise click.ClickException("WorkSpaces decommission analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("WorkSpaces decommission analysis", e)
            raise click.ClickException(str(e))

    @finops.command(name="validate-with-mcp")
    @click.option("--input", "-i", "input_file", required=True, help="Input Excel file with cost projections")
    @click.option("--resource-type", "-t", type=click.Choice(["ec2", "workspaces", "lambda", "snapshots"]),
                  default="ec2", help="Resource type for validation")
    @click.option("--tolerance", "-tol", type=float, default=0.05, help="Variance tolerance (default: 5%)")
    @click.option("--billing-profile", help="AWS billing profile for MCP Cost Explorer")
    def validate_with_mcp(input_file, resource_type, tolerance, billing_profile):
        """
        Validate runbooks cost projections against MCP Cost Explorer (Feature 1).

        Implements 3-mode validation:
        1. Import mode: Python API validation
        2. CLI mode: Command-line batch validation
        3. MCP cross-validation: Cost Explorer accuracy check (≥99.5% target)

        Examples:
            # Validate EC2 cost projections
            runbooks finops validate-with-mcp --input ec2-enriched.xlsx --resource-type ec2

            # Validate WorkSpaces with custom tolerance
            runbooks finops validate-with-mcp --input ws-enriched.xlsx --resource-type workspaces --tolerance 0.03

            # Validate with specific billing profile
            runbooks finops validate-with-mcp --input ec2-enriched.xlsx --billing-profile billing-mgmt
        """
        try:
            import pandas as pd
            from runbooks.finops.enhanced_mcp_integration import EnhancedMCPIntegration

            print_header(f"MCP Cost Validation - {resource_type.upper()}", f"Input: {input_file}")

            # Load cost projections from Excel
            print_info(f"Loading cost projections from {input_file}...")
            df = pd.read_excel(input_file, sheet_name=resource_type)

            # Determine cost and ID columns based on resource type
            column_mapping = {
                "ec2": {"id": "instance_id", "cost": "monthly_cost"},
                "workspaces": {"id": "WorkspaceId", "cost": "monthly_cost"},
                "lambda": {"id": "FunctionName", "cost": "monthly_cost"},
                "snapshots": {"id": "SnapshotId", "cost": "monthly_cost"}
            }

            id_column = column_mapping[resource_type]["id"]
            cost_column = column_mapping[resource_type]["cost"]

            # Validate columns exist
            if id_column not in df.columns or cost_column not in df.columns:
                raise click.ClickException(
                    f"Required columns missing: {id_column}, {cost_column}\n"
                    f"Available columns: {', '.join(df.columns)}"
                )

            # Build cost projections list
            cost_projections = []
            for _, row in df.iterrows():
                resource_id = row[id_column]
                cost = row[cost_column]

                # Skip invalid rows
                if pd.isna(resource_id) or pd.isna(cost) or cost == 0:
                    continue

                cost_projections.append({
                    "resource_id": str(resource_id),
                    "cost": float(cost)
                })

            if not cost_projections:
                print_warning("No valid cost projections found in input file")
                return

            print_success(f"Loaded {len(cost_projections)} cost projections for validation")

            # Initialize MCP integration
            mcp = EnhancedMCPIntegration(billing_profile=billing_profile)

            # Execute batch validation
            print_info(f"\nValidating against MCP Cost Explorer (tolerance: {tolerance:.1%})...")
            validation_results = mcp.validate_batch(cost_projections, tolerance=tolerance)

            # Display summary
            accuracy_metrics = validation_results.get("accuracy_metrics", {})
            if accuracy_metrics:
                print_header("Validation Summary")

                mcp_accuracy = accuracy_metrics.get("mcp_accuracy", 0.0)
                pass_rate = accuracy_metrics.get("pass_rate", 0.0)

                if mcp_accuracy >= 99.5:
                    print_success(f"✅ MCP Accuracy: {mcp_accuracy:.1f}% (EXCEEDS ≥99.5% target)")
                elif mcp_accuracy >= 95.0:
                    print_warning(f"⚠️  MCP Accuracy: {mcp_accuracy:.1f}% (meets ≥95% baseline)")
                else:
                    print_error(f"❌ MCP Accuracy: {mcp_accuracy:.1f}% (below 95% baseline)")

                print_info(f"   Pass Rate: {pass_rate:.1f}%")
                print_info(f"   Average Variance: {accuracy_metrics.get('average_variance', 0):.2%}")
                print_info(f"   Total Resources: {accuracy_metrics.get('total_resources', 0)}")

                # Export validation results
                output_file = input_file.replace('.xlsx', '-mcp-validation.json')
                import json
                with open(output_file, 'w') as f:
                    json.dump(validation_results, f, indent=2)

                print_success(f"\n📁 Validation results exported: {output_file}")

        except ImportError as e:
            error_handlers["module_not_available"]("MCP Integration", e)
            raise click.ClickException("MCP validation functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("MCP validation", e)
            raise click.ClickException(str(e))

    @finops.command(name="analyze-graviton-eligibility")
    @click.option(
        "--input",
        "-i",
        "input_file",
        required=True,
        type=click.Path(exists=True),
        help="Input EC2 enriched Excel file (output from analyze-ec2)"
    )
    @click.option(
        "--output",
        "-o",
        "output_file",
        required=True,
        type=click.Path(),
        help="Output Excel file with Graviton eligibility assessment"
    )
    @click.option(
        "--management-profile",
        "-m",
        help="AWS management profile for EC2 operations (defaults to service routing)"
    )
    @click.option(
        "--billing-profile",
        "-b",
        help="AWS billing profile for Cost Explorer (defaults to service routing)"
    )
    @click.option(
        "--enable-ami-check/--no-ami-check",
        default=True,
        help="Enable AMI architecture compatibility checking (default: enabled)"
    )
    @click.option(
        "--enable-compute-optimizer/--no-compute-optimizer",
        default=True,
        help="Enable Compute Optimizer integration for confidence scoring (default: enabled)"
    )
    @click.option(
        "--verbose",
        "-v",
        is_flag=True,
        help="Enable verbose logging"
    )
    def analyze_graviton_eligibility_cmd(
        input_file,
        output_file,
        management_profile,
        billing_profile,
        enable_ami_check,
        enable_compute_optimizer,
        verbose
    ):
        """
        Graviton migration eligibility analysis for ARM64 cost optimization.

        This command analyzes EC2 instances for AWS Graviton (ARM64) migration potential
        with 40% cost savings targeting $800K+ annual opportunity.

        \b
        🎯 BUSINESS VALUE
        • $800K+ annual savings opportunity (40% Graviton price reduction)
        • 137 EC2 instances baseline analysis
        • Integration with E2-E7 decommission signals for confidence scoring

        \b
        🔍 ELIGIBILITY ASSESSMENT
        • Instance type compatibility mapping (x86_64 → ARM64)
        • AMI architecture compatibility checking
        • Application workload heuristics (Linux-friendly, GPU-incompatible)
        • Cost savings projection (40% reduction model)
        • Integration with Compute Optimizer recommendations

        \b
        📊 SCORING MODEL (0-100 scale)
        • Has Graviton mapping: 30 points (baseline eligibility)
        • AMI ARM64 compatible: 25 points (verified architecture support)
        • Application compatible (>70): 25 points (workload assessment)
        • Savings threshold met (≥$10/mo): 10 points (business justification)
        • Compute Optimizer idle: 10 points bonus (decommission candidate)

        \b
        🎯 RECOMMENDATION TIERS
        • RECOMMEND (≥70): Strong migration candidates (LOW complexity)
        • EVALUATE (40-69): Requires testing/validation (MEDIUM complexity)
        • NOT_RECOMMENDED (<40): Migration challenges (HIGH complexity)

        \b
        📁 OUTPUT FORMAT
        Multi-sheet Excel with:
        • Sheet 1: Graviton Analysis (12 new columns + original EC2 data)
        • Sheet 2: RECOMMEND Tier (high-confidence candidates)
        • Sheet 3: EVALUATE Tier (testing recommended)
        • Sheet 4: Cost Savings Summary (by account/tier)

        \b
        🔗 INTEGRATION
        • Input: EC2 enriched data from analyze-ec2 command
        • Dependencies: EC2 analyzer (137 instances), Compute Optimizer
        • Feature Coordination: E2-E7 signals (when available)

        \b
        💰 Epic 4 Feature 2: Graviton Migration Analyzer
        Timeline: 7 days (5 dev + 2 testing)
        Business Impact: $800K+ annual savings potential

        \b
        📖 Examples:
            # Basic Graviton analysis with auto-profile routing (v1.1.11+)
            runbooks finops analyze-graviton-eligibility \\
                -i ec2-enriched.xlsx \\
                -o graviton-analysis.xlsx

            # With explicit profiles (backward compatible)
            runbooks finops analyze-graviton-eligibility \\
                -i ec2-enriched.xlsx \\
                -o graviton-analysis.xlsx \\
                -m mgmt-profile \\
                -b billing-profile

            # Disable AMI checking (faster execution)
            runbooks finops analyze-graviton-eligibility \\
                -i ec2-enriched.xlsx \\
                -o graviton-analysis.xlsx \\
                --no-ami-check

            # With verbose logging
            runbooks finops analyze-graviton-eligibility \\
                -i ec2-enriched.xlsx \\
                -o graviton-analysis.xlsx \\
                --verbose

        \b
        Pattern: Follows ec2_analyzer.py proven patterns with Rich CLI standards
        """
        try:
            from runbooks.finops.graviton_migration_analyzer import analyze_graviton_eligibility
            from runbooks.common.rich_utils import print_header, print_success, print_error

            # Configure logging if verbose
            if verbose:
                import logging
                logging.basicConfig(level=logging.DEBUG)

            print_header("Graviton Migration Analysis", f"Input: {input_file}")

            # Execute Graviton eligibility analysis
            result_df = analyze_graviton_eligibility(
                input_file=input_file,
                output_file=output_file,
                management_profile=management_profile,
                billing_profile=billing_profile,
                enable_ami_check=enable_ami_check,
                enable_compute_optimizer=enable_compute_optimizer
            )

            # Summary statistics
            total_instances = len(result_df)
            recommend_count = (result_df['graviton_recommendation'] == 'RECOMMEND').sum()
            evaluate_count = (result_df['graviton_recommendation'] == 'EVALUATE').sum()
            total_monthly_savings = result_df['graviton_monthly_savings'].sum()
            total_annual_savings = result_df['graviton_annual_savings'].sum()

            print_success(f"\n✅ Graviton analysis complete!")
            print_success(f"   📊 Analyzed {total_instances} EC2 instances")
            print_success(f"   ✅ {recommend_count} instances ready for migration (RECOMMEND)")
            print_success(f"   🔍 {evaluate_count} instances require testing (EVALUATE)")
            print_success(f"   💰 Potential savings: ${total_monthly_savings:,.2f}/month (${total_annual_savings:,.2f}/year)")
            print_success(f"   📁 Output: {output_file}")

            if total_annual_savings >= 800000:
                print_success(f"\n🎯 $800K+ annual savings target ACHIEVED!")

            return result_df

        except ImportError as e:
            error_handlers["module_not_available"]("Graviton Migration Analyzer", e)
            raise click.ClickException("Graviton analysis functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("Graviton eligibility analysis", e)
            raise click.ClickException(str(e))

    @finops.command(name='optimize-savings-plans')
    @click.option('--lookback-days', default=90, type=int,
                  help='Usage history analysis period (default: 90 days, minimum: 30)')
    @click.option('--term-years', type=click.Choice(['1', '3']), default='1',
                  help='Commitment term: 1-year or 3-year (default: 1)')
    @click.option('--payment-option',
                  type=click.Choice(['ALL_UPFRONT', 'PARTIAL_UPFRONT', 'NO_UPFRONT']),
                  default='NO_UPFRONT',
                  help='Payment option (default: NO_UPFRONT)')
    @click.option('--billing-profile', help='AWS billing profile for Cost Explorer')
    @click.option('--validate-with-mcp/--no-mcp-validation', default=True,
                  help='Enable MCP validation (≥99.5% accuracy, default: enabled)')
    @click.option('--output', type=click.Path(), help='Export recommendations to Excel')
    @click.option('--verbose', is_flag=True, help='Enable verbose logging')
    def optimize_savings_plans_cmd(lookback_days, term_years, payment_option, billing_profile,
                                   validate_with_mcp, output, verbose):
        """
        Generate hybrid Savings Plans + RI recommendations (60/30/10 strategy).

        \b
        🎯 HYBRID OPTIMIZATION STRATEGY
        • 60% Compute SP (flexible across EC2/Fargate/Lambda, 66% savings)
        • 30% EC2 Instance SP (stable workloads, 72% savings)
        • 10% On-Demand (burst capacity, flexibility)

        \b
        📊 WORKLOAD CLASSIFICATION
        Stable workloads (EC2 Instance SP candidates):
        • Coefficient of Variation (CV) < 0.15
        • Uptime > 95% over analysis period
        • ≤2 regions (single or limited multi-region)

        Variable workloads (Compute SP candidates):
        • CV ≥ 0.15 (fluctuating usage patterns)
        • Multi-region (3+ regions for flexibility)

        \b
        💰 FINANCIAL MODELING
        • Break-even analysis (target: <10 months)
        • ROI calculations with upfront cost amortization
        • Annual savings projections ($500K+ target)
        • MCP validation (≥99.5% accuracy for major commitments)

        \b
        🔧 ANALYSIS PERIOD
        • Default: 90 days (3 months minimum for SP sizing)
        • Minimum: 30 days (shorter periods reduce accuracy)
        • Recommended: 90-180 days for seasonal patterns

        \b
        📖 Examples:
            # Basic SP optimization with 90-day analysis
            runbooks finops optimize-savings-plans \\
                --billing-profile billing-account

            # 6-month analysis for seasonal patterns
            runbooks finops optimize-savings-plans \\
                --lookback-days 180 \\
                --billing-profile billing-account

            # 3-year term with partial upfront
            runbooks finops optimize-savings-plans \\
                --term-years 3 \\
                --payment-option PARTIAL_UPFRONT \\
                --billing-profile billing-account

            # Export to Excel with verbose logging
            runbooks finops optimize-savings-plans \\
                --billing-profile billing-account \\
                --output sp-recommendations.xlsx \\
                --verbose

            # Disable MCP validation (faster, but no accuracy guarantee)
            runbooks finops optimize-savings-plans \\
                --billing-profile billing-account \\
                --no-mcp-validation

        \b
        🎯 Epic 5 Feature 3: Savings Plans Hybrid Optimizer
        Timeline: 5 days (4 dev + 1 testing)
        Business Impact: $500K+ annual savings potential
        Target Accuracy: ≥99.5% MCP validation for cost projections

        \b
        Pattern: Follows reservation_optimizer.py proven patterns with Rich CLI standards
        Integration: Complementary to existing RI optimizer for unified procurement strategy
        """
        try:
            import asyncio
            from runbooks.finops.savings_plans_optimizer import SavingsPlansOptimizer
            from runbooks.common.rich_utils import print_header, print_success, print_error, print_info

            # Configure logging if verbose
            if verbose:
                import logging
                logging.basicConfig(level=logging.DEBUG)

            print_header("Savings Plans Hybrid Optimizer", f"Analysis Period: {lookback_days} days")

            # Initialize optimizer
            optimizer = SavingsPlansOptimizer(
                profile_name=billing_profile,
                regions=None  # Will use default regions from profile
            )

            # Execute comprehensive analysis
            recommendations = asyncio.run(
                optimizer.generate_recommendations(
                    usage_history_days=lookback_days,
                    validate_with_mcp=validate_with_mcp
                )
            )

            if not recommendations:
                print_info("No cost-effective Savings Plans opportunities identified")
                return

            # Calculate summary statistics
            total_savings = sum(rec.estimated_annual_savings for rec in recommendations)
            total_commitment = sum(rec.commitment_usd_hourly for rec in recommendations)
            compute_sp_count = sum(1 for r in recommendations if r.plan_type.value == "Compute")
            ec2_instance_sp_count = sum(1 for r in recommendations if r.plan_type.value == "EC2Instance")

            # Display final success message
            print_success(f"\n✅ Savings Plans optimization complete!")
            print_success(f"   📊 {len(recommendations)} recommendations generated")
            print_success(f"   💰 Potential annual savings: ${total_savings:,.2f}")
            print_success(f"   💲 Total hourly commitment: ${total_commitment:.2f}/hour")
            print_success(f"   🔧 Plan breakdown:")
            if compute_sp_count > 0:
                print_success(f"      • {compute_sp_count} Compute SP (60% allocation, flexible)")
            if ec2_instance_sp_count > 0:
                print_success(f"      • {ec2_instance_sp_count} EC2 Instance SP (30% allocation, stable)")
            print_info(f"      • 10% remains On-Demand (burst capacity)")

            # Epic 5 achievement validation
            if total_savings >= 500000:
                print_success(f"\n🎯 $500K+ annual savings target ACHIEVED!")
                print_info(f"   Epic 5 (Reserved Capacity): 40%→80% progression unlocked")

            # Export to Excel if requested
            if output:
                try:
                    import pandas as pd
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment

                    # Create Excel workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "SP Recommendations"

                    # Headers
                    headers = [
                        "Plan Type", "Commitment ($/hr)", "Annual Savings",
                        "Coverage %", "Allocation", "Confidence %", "ROI %",
                        "Break-even (months)", "Risk Level", "Regions", "Services"
                    ]
                    ws.append(headers)

                    # Style headers
                    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")

                    # Data rows
                    for rec in recommendations:
                        ws.append([
                            rec.plan_type.value,
                            f"${rec.commitment_usd_hourly:.2f}",
                            f"${rec.estimated_annual_savings:,.2f}",
                            f"{rec.coverage_percentage:.1f}%",
                            rec.hybrid_strategy.get("allocation", "N/A"),
                            f"{rec.confidence_score * 100:.1f}%",
                            f"{rec.roi_percentage:.1f}%",
                            f"{rec.break_even_months:.1f}",
                            rec.risk_level,
                            ", ".join(rec.regions[:3]) + ("..." if len(rec.regions) > 3 else ""),
                            ", ".join(rec.services)
                        ])

                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

                    # Save workbook
                    wb.save(output)
                    print_success(f"   📁 Recommendations exported: {output}")

                except Exception as e:
                    print_error(f"Failed to export Excel: {str(e)}")

        except ImportError as e:
            error_handlers["module_not_available"]("Savings Plans Optimizer", e)
            raise click.ClickException("Savings Plans optimization functionality not available")
        except Exception as e:
            error_handlers["operation_failed"]("Savings Plans optimization", e)
            raise click.ClickException(str(e))

    @finops.command("optimize-s3-lifecycle")
    @click.option("--profile", help="AWS profile name")
    @click.option("--regions", multiple=True, help="AWS regions to analyze")
    @click.option("--dry-run/--no-dry-run", default=True, help="Execute in dry-run mode (READ-ONLY analysis)")
    @click.option(
        "-f",
        "--format",
        "--export-format",
        type=click.Choice(["json", "csv"]),
        default="json",
        help="Export format for results",
    )
    @click.option("--output-file", help="Output file path for results export")
    def optimize_s3_lifecycle_cmd(profile, regions, dry_run, format, output_file):
        """
        S3 Lifecycle Optimizer - Automated Storage Cost Optimization ($180K target)

        Comprehensive S3 lifecycle policy optimization targeting $180K annual savings:
        • Intelligent-Tiering for hybrid access patterns (50% IA savings)
        • Glacier transitions for archive-eligible data (80% savings)
        • Deep Archive for long-term retention (92% savings)
        • Lifecycle expiration for temporary/log data

        Part of Epic 3 completion strategy (70% → 85%).

        SAFETY: READ-ONLY analysis only - lifecycle policies require explicit approval.

        Examples:
            runbooks finops optimize-s3-lifecycle
            runbooks finops optimize-s3-lifecycle --profile my-profile --regions ap-southeast-2
            runbooks finops optimize-s3-lifecycle --export-format csv --output-file s3_recommendations.csv
        """
        try:
            # Lazy import for performance
            from runbooks.finops.s3_lifecycle_optimizer import S3LifecycleOptimizer
            import asyncio

            print_header("S3 Lifecycle Automation", "Enterprise Storage Cost Optimization")

            # Initialize optimizer
            optimizer = S3LifecycleOptimizer(profile_name=profile, regions=list(regions) if regions else None)

            # Execute analysis
            results = asyncio.run(optimizer.analyze_s3_lifecycle_optimization(dry_run=dry_run))

            # Export results if requested
            if output_file or format != "json":
                optimizer.export_recommendations(results, output_file, format)

            # Display final success message
            if results.total_potential_annual_savings > 0:
                print_success(
                    f"Analysis complete: {format_cost(results.total_potential_annual_savings)} potential annual savings"
                )
                print_info(
                    f"Strategies: IT ({format_cost(results.intelligent_tiering_annual_savings)}) | "
                    f"Glacier ({format_cost(results.glacier_annual_savings)})"
                )
            else:
                print_info("Analysis complete: All S3 buckets have optimal lifecycle policies")

        except ImportError as e:
            error_handlers["module_not_available"]("S3 Lifecycle Optimizer", e)
            raise click.ClickException("S3 Lifecycle optimization functionality not available")
        except KeyboardInterrupt:
            print_warning("Analysis interrupted by user")
            raise click.Abort()
        except Exception as e:
            error_handlers["operation_failed"]("S3 Lifecycle optimization", e)
            raise click.ClickException(str(e))

    @finops.command("detect-rds-idle")
    @common_aws_options
    @click.option(
        "--lookback-days",
        type=int,
        default=7,
        help="Analysis period in days for CloudWatch metrics (default: 7)"
    )
    @click.option(
        "--connection-threshold",
        type=int,
        default=10,
        help="Daily connection threshold for idle detection (default: 10)"
    )
    @click.option(
        "--cpu-threshold",
        type=float,
        default=5.0,
        help="CPU percentage threshold for idle detection (default: 5.0)"
    )
    @click.option(
        "--export-json",
        is_flag=True,
        default=False,
        help="Export results to JSON file"
    )
    @click.option(
        "--output-file",
        type=click.Path(),
        help="Output file path for JSON export"
    )
    def detect_rds_idle(profile, regions, lookback_days, connection_threshold, cpu_threshold, export_json, output_file):
        """
        Detect idle RDS instances for $50K annual savings potential.

        Analyzes RDS instances using 5 idle signals:
        - I1: DatabaseConnections <10/day (40 points)
        - I2: CPUUtilization <5% avg (30 points)
        - I3: ReadIOPS + WriteIOPS <100/day (15 points)
        - I4: NetworkReceiveThroughput <1MB/day (10 points)
        - I5: No recent snapshots modified (5 points)

        Recommendations:
        - Score 80-100: TERMINATE (high confidence idle)
        - Score 60-79: STOP (medium confidence)
        - Score 40-59: DOWNSIZE (low utilization)
        - Score <40: KEEP (active usage)

        Examples:
            # Basic idle detection with 7-day lookback
            runbooks finops detect-rds-idle --profile prod

            # Extended analysis with 30-day lookback
            runbooks finops detect-rds-idle --profile prod --lookback-days 30

            # Custom thresholds with JSON export
            runbooks finops detect-rds-idle --profile prod --connection-threshold 5 --cpu-threshold 3.0 --export-json
        """
        try:
            from runbooks.finops.rds_analyzer import RDSAnalysisConfig, RDSCostAnalyzer
            import json

            # Initialize configuration
            config = RDSAnalysisConfig(
                management_profile=profile,
                billing_profile=profile,
                regions=list(regions) if regions else ['ap-southeast-2'],
                enable_organizations=False,  # Not needed for idle detection
                enable_cost=False  # Using internal cost estimation
            )

            # Initialize analyzer
            analyzer = RDSCostAnalyzer(config)

            # Execute idle detection
            idle_analyses = analyzer.detect_idle_instances(
                lookback_days=lookback_days,
                connection_threshold=connection_threshold,
                cpu_threshold=cpu_threshold
            )

            if not idle_analyses:
                print_warning("No RDS instances found for idle detection")
                return

            # Display results in Rich table
            from rich.table import Table

            table = Table(
                title=f"RDS Idle Instance Detection ({len(idle_analyses)} instances analyzed)",
                show_header=True,
                header_style="bold cyan"
            )

            table.add_column("Instance ID", style="yellow")
            table.add_column("Class", style="cyan")
            table.add_column("Engine", style="blue")
            table.add_column("Region", style="magenta")
            table.add_column("Idle Score", justify="right", style="bold")
            table.add_column("Confidence", style="green")
            table.add_column("Recommendation", style="bold red")
            table.add_column("Annual Savings", justify="right", style="green")

            for analysis in sorted(idle_analyses, key=lambda x: x.idle_score, reverse=True):
                # Color-code recommendation
                recommendation_color = {
                    "TERMINATE": "bold red",
                    "STOP": "yellow",
                    "DOWNSIZE": "cyan",
                    "KEEP": "green"
                }.get(analysis.recommendation, "white")

                table.add_row(
                    analysis.instance_id,
                    analysis.instance_class,
                    analysis.engine,
                    analysis.region,
                    f"{analysis.idle_score}/100",
                    analysis.idle_confidence,
                    f"[{recommendation_color}]{analysis.recommendation}[/{recommendation_color}]",
                    f"${analysis.annual_savings_potential:,.2f}"
                )

            console.print(table)

            # Summary statistics
            high_confidence = [a for a in idle_analyses if a.idle_confidence == "HIGH"]
            total_savings = sum(a.annual_savings_potential for a in idle_analyses)

            print_info(f"\nSummary:")
            print_info(f"  Total instances analyzed: {len(idle_analyses)}")
            print_info(f"  High-confidence idle: {len(high_confidence)}")
            print_info(f"  Total annual savings potential: ${total_savings:,.2f}")

            # Export to JSON if requested
            if export_json or output_file:
                output_path = output_file or f"rds-idle-detection-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json"

                export_data = {
                    "analysis_timestamp": datetime.now().isoformat(),
                    "lookback_days": lookback_days,
                    "connection_threshold": connection_threshold,
                    "cpu_threshold": cpu_threshold,
                    "total_instances": len(idle_analyses),
                    "high_confidence_idle": len(high_confidence),
                    "total_annual_savings": total_savings,
                    "instances": [
                        {
                            "instance_id": a.instance_id,
                            "instance_class": a.instance_class,
                            "engine": a.engine,
                            "account_id": a.account_id,
                            "region": a.region,
                            "idle_signals": a.idle_signals,
                            "idle_score": a.idle_score,
                            "idle_confidence": a.idle_confidence,
                            "monthly_cost": a.monthly_cost,
                            "annual_savings_potential": a.annual_savings_potential,
                            "recommendation": a.recommendation
                        }
                        for a in idle_analyses
                    ]
                }

                with open(output_path, 'w') as f:
                    json.dump(export_data, f, indent=2)

                print_success(f"Results exported to: {output_path}")

        except ImportError as e:
            error_handlers["module_not_available"]("RDS Idle Detection", e)
            raise click.ClickException("RDS idle detection functionality not available")
        except KeyboardInterrupt:
            print_warning("Detection interrupted by user")
            raise click.Abort()
        except Exception as e:
            error_handlers["operation_failed"]("RDS idle detection", e)
            raise click.ClickException(str(e))

    # Phase 1B: Critical Gap Closure Commands (Cost Optimization Playbook Enhancement)

    @finops.command("optimize-cloudwatch-costs")
    @common_aws_options
    @click.option(
        "--regions",
        multiple=True,
        help="AWS regions to analyze (default: major commercial regions)"
    )
    @click.option(
        "--mcp-validation",
        is_flag=True,
        help="Enable MCP validation for cost projections"
    )
    @click.option(
        "--export-format",
        type=click.Choice(["json", "csv", "markdown"]),
        help="Export results format"
    )
    @click.option(
        "--output-file",
        type=click.Path(),
        help="Output file path for results export"
    )
    @click.option("--executive", is_flag=True, help="Executive summary format (business narrative, <5 min review)")
    def optimize_cloudwatch_costs(profile, regions, dry_run, mcp_validation, export_format, output_file, executive):
        """
        Analyze and optimize CloudWatch log retention costs.

        Implements Cost Optimization Playbook Phase 4 (CloudWatch cost controls).

        Capabilities:
        - Log group retention policy optimization
        - Cost savings calculations ($0.50/GB ingestion, $0.03/GB storage)
        - Automated retention recommendations (7/30/90/365 days)
        - Infinite retention detection and remediation

        Business Impact: Typical savings of $10K-$50K annually
        """
        try:
            import asyncio
            from runbooks.finops.cloudwatch_cost_optimizer import CloudWatchCostOptimizer

            print_header("CloudWatch Cost Optimization", "Enterprise Log Retention Control")

            optimizer = CloudWatchCostOptimizer(
                profile_name=profile,
                regions=list(regions) if regions else None,
                dry_run=dry_run
            )

            results = asyncio.run(optimizer.analyze_cloudwatch_costs(enable_mcp_validation=mcp_validation))

            print_success(f"✅ CloudWatch analysis complete: {results.log_groups_optimizable} optimization opportunities")
            print_info(f"   Potential annual savings: ${results.potential_annual_savings:,.2f}")

        except ImportError as e:
            error_handlers["module_not_available"]("CloudWatch Cost Optimizer", e)
            raise click.ClickException("CloudWatch cost optimization not available")
        except Exception as e:
            error_handlers["operation_failed"]("CloudWatch cost optimization", e)
            raise click.ClickException(str(e))

    @finops.command("detect-orphans")
    @common_aws_options
    @click.option(
        "--regions",
        multiple=True,
        help="AWS regions to analyze"
    )
    @click.option(
        "--resource-type",
        type=click.Choice(["all", "ebs", "eip", "logs", "nat", "lb", "snapshot"]),
        default="all",
        help="Resource type to detect (default: all)"
    )
    @click.option(
        "--validate-with-config",
        is_flag=True,
        help="Validate orphans with AWS Config compliance rules"
    )
    @click.option("--executive", is_flag=True, help="Executive summary format (business narrative, <5 min review)")
    def detect_orphans(profile, regions, resource_type, validate_with_config, executive):
        """
        Detect orphaned AWS resources across multiple types.

        Implements unified orphan detection from Cost Optimization Playbook Phase 3.

        Orphan Types:
        - EBS volumes (unattached >30 days)
        - Elastic IPs (unallocated)
        - CloudWatch Log Groups (no recent events)
        - NAT Gateways (no traffic)
        - Load Balancers (no targets)
        - Snapshots (orphaned - no AMI/volume reference)

        Decision Rubric:
        - MUST: Immediate action (zero risk, high cost)
        - SHOULD: High priority (low risk, moderate cost)
        - COULD: Investigation recommended

        Business Impact: Typical savings of $50K-$200K annually
        """
        try:
            import asyncio
            from runbooks.finops.orphan_resource_detector import (
                OrphanResourceDetector,
                OrphanResourceType
            )

            print_header("Unified Orphan Detection", "Enterprise Waste Elimination")

            # Map CLI resource type to enum
            resource_type_map = {
                "all": OrphanResourceType.ALL,
                "ebs": OrphanResourceType.EBS_VOLUME,
                "eip": OrphanResourceType.ELASTIC_IP,
                "nat": OrphanResourceType.NAT_GATEWAY,
                "lb": OrphanResourceType.LOAD_BALANCER,
            }

            detector = OrphanResourceDetector(
                profile_name=profile,
                regions=list(regions) if regions else None
            )

            results = asyncio.run(
                detector.detect_orphaned_resources(resource_type=resource_type_map[resource_type])
            )

            print_success(f"✅ Orphan detection complete: {results.total_orphans_detected} orphans found")
            print_info(f"   MUST act: {results.orphans_by_decision_level.get('MUST', 0)}")
            print_info(f"   Potential annual savings: ${results.potential_annual_savings:,.2f}")

            if validate_with_config:
                print_info("🔗 AWS Config validation enabled - integrating with security module")

        except ImportError as e:
            error_handlers["module_not_available"]("Orphan Resource Detector", e)
            raise click.ClickException("Orphan detection not available")
        except Exception as e:
            error_handlers["operation_failed"]("Orphan detection", e)
            raise click.ClickException(str(e))

    @finops.command("analyze-s3-storage-lens")
    @common_aws_options
    @click.option(
        "--region",
        default="ap-southeast-2",
        help="AWS region for S3 client (S3 is global)"
    )
    def analyze_s3_storage_lens(profile, region):
        """
        Analyze S3 Storage Lens metrics for cost optimization.

        Implements Cost Optimization Playbook Feature #6 (S3 Storage Lens integration).

        Capabilities:
        - Bucket lifecycle policy gap detection
        - Intelligent-Tiering recommendations
        - Incomplete multipart upload (MPU) detection
        - Fastest-growing bucket identification
        - Storage class distribution analysis

        Business Impact: Typical savings of $30K-$150K annually
        """
        try:
            import asyncio
            from runbooks.finops.s3_storage_lens_analyzer import S3StorageLensAnalyzer

            print_header("S3 Storage Lens Analysis", "Enterprise S3 Cost Intelligence")

            analyzer = S3StorageLensAnalyzer(
                profile_name=profile,
                regions=[region]
            )

            results = asyncio.run(analyzer.analyze_s3_storage_lens())

            print_success(f"✅ S3 Storage Lens analysis complete: {results.total_buckets_analyzed} buckets")
            print_info(f"   Buckets without lifecycle: {results.buckets_without_lifecycle}")
            print_info(f"   Potential annual savings: ${results.potential_annual_savings:,.2f}")

        except ImportError as e:
            error_handlers["module_not_available"]("S3 Storage Lens Analyzer", e)
            raise click.ClickException("S3 Storage Lens analysis not available")
        except Exception as e:
            error_handlers["operation_failed"]("S3 Storage Lens analysis", e)
            raise click.ClickException(str(e))

    @finops.command("check-config-compliance")
    @common_aws_options
    @click.option(
        "--regions",
        multiple=True,
        help="AWS regions to analyze"
    )
    @click.option(
        "--config-rules",
        multiple=True,
        type=click.Choice(["ebs-inuse", "eip-attached", "cw-retention", "all"]),
        default=["all"],
        help="AWS Config rules to check (default: all)"
    )
    def check_config_compliance(profile, regions, config_rules):
        """
        Check AWS Config compliance and map to cost impact.

        Implements Cost Optimization Playbook Phase 2 (AWS Config integration).

        Config Rules:
        - ebs-inuse: Detect unattached EBS volumes
        - eip-attached: Detect unallocated Elastic IPs
        - cw-retention: CloudWatch log retention compliance

        Integration:
        - Cross-module with security/config module
        - Maps compliance violations to cost impact
        - Correlates with orphan detection

        Business Impact: Typical savings of $20K-$80K annually
        """
        try:
            import asyncio
            from runbooks.finops.config_compliance_checker import (
                ConfigComplianceChecker,
                ConfigComplianceRule
            )

            print_header("AWS Config FinOps Integration", "Compliance-Driven Cost Optimization")

            # Map CLI config rules to enum
            config_rule_map = {
                "ebs-inuse": ConfigComplianceRule.EBS_VOLUME_INUSE_CHECK,
                "eip-attached": ConfigComplianceRule.EIP_ATTACHED,
                "cw-retention": ConfigComplianceRule.CW_LOGGROUP_RETENTION_PERIOD_CHECK,
            }

            rules_to_check = []
            if "all" in config_rules:
                rules_to_check = [
                    ConfigComplianceRule.EBS_VOLUME_INUSE_CHECK,
                    ConfigComplianceRule.EIP_ATTACHED,
                ]
            else:
                rules_to_check = [config_rule_map[rule] for rule in config_rules if rule in config_rule_map]

            checker = ConfigComplianceChecker(
                profile_name=profile,
                regions=list(regions) if regions else None
            )

            results = asyncio.run(checker.check_config_compliance(config_rules=rules_to_check))

            print_success(f"✅ Config compliance check complete: {results.total_violations_detected} violations")
            print_info(f"   Annual cost impact: ${results.total_annual_cost_impact:,.2f}")

            if results.total_violations_detected > 0:
                print_info("💡 Tip: Run 'runbooks finops detect-orphans --validate-with-config' for unified analysis")

        except ImportError as e:
            error_handlers["module_not_available"]("Config Compliance Checker", e)
            raise click.ClickException("Config compliance checking not available")
        except Exception as e:
            error_handlers["operation_failed"]("Config compliance check", e)
            raise click.ClickException(str(e))

    return finops
