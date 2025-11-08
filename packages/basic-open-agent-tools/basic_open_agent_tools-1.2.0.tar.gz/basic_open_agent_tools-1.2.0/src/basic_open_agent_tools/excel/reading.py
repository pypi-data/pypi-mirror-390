"""Excel spreadsheet reading and extraction functions for AI agents.

This module provides functions for reading and extracting data from
Excel (.xlsx) spreadsheets.
"""

import os

from ..decorators import strands_tool

try:
    from openpyxl import load_workbook  # type: ignore[import-untyped, import-not-found]

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Maximum file size: 100MB
MAX_FILE_SIZE = 100 * 1024 * 1024


@strands_tool
def read_excel_sheet(file_path: str, sheet_name: str) -> list[list[str]]:
    """Read Excel sheet as 2D list of strings.

    This function reads all data from a specific sheet and returns it
    as a 2D list where each inner list represents a row.

    Args:
        file_path: Path to Excel file to read
        sheet_name: Name of sheet to read

    Returns:
        2D list of strings representing sheet data

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, or sheet doesn't exist

    Example:
        >>> data = read_excel_sheet("/path/to/file.xlsx", "Sheet1")
        >>> data[0]  # First row
        ['Name', 'Age', 'City']
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        # Load workbook
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        # Get sheet
        ws = wb[sheet_name]

        # Extract all data
        data = []
        for row in ws.iter_rows(values_only=True):
            # Convert row to strings, handling None values
            str_row = [str(cell) if cell is not None else "" for cell in row]
            data.append(str_row)

        wb.close()
        return data

    except Exception as e:
        raise ValueError(f"Failed to read Excel sheet: {e}")


@strands_tool
def get_excel_sheet_names(file_path: str) -> list[str]:
    """Get list of all sheet names in Excel workbook.

    Args:
        file_path: Path to Excel file

    Returns:
        List of sheet names

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large or unreadable

    Example:
        >>> sheets = get_excel_sheet_names("/path/to/file.xlsx")
        >>> sheets
        ['Sheet1', 'Sheet2', 'Data']
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True)
        sheet_names = list(wb.sheetnames)  # Cast to list[str] for type safety
        wb.close()
        return sheet_names

    except Exception as e:
        raise ValueError(f"Failed to get sheet names: {e}")


@strands_tool
def read_excel_as_dicts(
    file_path: str, sheet_name: str, header_row: int
) -> list[dict[str, str]]:
    """Read Excel sheet as list of dictionaries using header row as keys.

    This function reads a sheet and returns each row as a dictionary where
    keys come from the specified header row.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to read
        header_row: Row number containing headers (1-indexed)

    Returns:
        List of dictionaries, one per data row

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        ValueError: If header_row invalid, file unreadable, or sheet doesn't exist
        FileNotFoundError: If file doesn't exist

    Example:
        >>> data = read_excel_as_dicts("/path/to/file.xlsx", "Sheet1", 1)
        >>> data[0]
        {'Name': 'Alice', 'Age': '30', 'City': 'NYC'}
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(header_row, int):
        raise TypeError("header_row must be an integer")

    if header_row < 1:
        raise ValueError("header_row must be >= 1 (1-indexed)")

    # Read sheet data
    data = read_excel_sheet(file_path, sheet_name)

    if len(data) < header_row:
        raise ValueError(
            f"header_row {header_row} out of range (sheet has {len(data)} rows)"
        )

    # Extract headers (convert to 0-indexed)
    headers = data[header_row - 1]

    # Convert rows to dicts, skipping header row
    result = []
    for i, row in enumerate(data):
        if i < header_row:
            continue  # Skip rows before and including header

        # Create dict with headers as keys
        row_dict = {}
        for j, cell in enumerate(row):
            if j < len(headers):
                key = headers[j] if headers[j] else f"Column{j + 1}"
                row_dict[key] = cell
            else:
                # Handle extra cells beyond headers
                row_dict[f"Column{j + 1}"] = cell

        result.append(row_dict)

    return result


@strands_tool
def get_excel_cell_value(file_path: str, sheet_name: str, cell_reference: str) -> str:
    """Get value from single cell using A1 notation.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        cell_reference: Cell reference in A1 notation (e.g., "B5", "AA10")

    Returns:
        Cell value as string

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        ValueError: If cell_reference invalid or sheet doesn't exist
        FileNotFoundError: If file doesn't exist

    Example:
        >>> value = get_excel_cell_value("/path/to/file.xlsx", "Sheet1", "B5")
        >>> value
        '42'
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(cell_reference, str):
        raise TypeError("cell_reference must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]
        cell_value = ws[cell_reference].value
        wb.close()

        return str(cell_value) if cell_value is not None else ""

    except KeyError:
        raise ValueError(f"Invalid cell reference: {cell_reference}")
    except Exception as e:
        raise ValueError(f"Failed to get cell value: {e}")


@strands_tool
def get_excel_cell_range(
    file_path: str, sheet_name: str, start_cell: str, end_cell: str
) -> list[list[str]]:
    """Get range of cells as 2D list using A1 notation.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        start_cell: Start cell reference (e.g., "A1")
        end_cell: End cell reference (e.g., "C10")

    Returns:
        2D list of strings for the specified range

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        ValueError: If cell references invalid or sheet doesn't exist
        FileNotFoundError: If file doesn't exist

    Example:
        >>> data = get_excel_cell_range("/path/to/file.xlsx", "Sheet1", "A1", "C3")
        >>> data
        [['Name', 'Age', 'City'], ['Alice', '30', 'NYC'], ['Bob', '25', 'LA']]
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(start_cell, str):
        raise TypeError("start_cell must be a string")

    if not isinstance(end_cell, str):
        raise TypeError("end_cell must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get range
        cell_range = f"{start_cell}:{end_cell}"
        data = []

        for row in ws[cell_range]:
            str_row = [
                str(cell.value) if cell.value is not None else "" for cell in row
            ]
            data.append(str_row)

        wb.close()
        return data

    except Exception as e:
        raise ValueError(f"Failed to get cell range: {e}")


@strands_tool
def search_excel_text(
    file_path: str, search_term: str, case_sensitive: bool
) -> list[dict[str, object]]:
    """Search for text across all sheets in Excel workbook.

    Args:
        file_path: Path to Excel file
        search_term: Text to search for
        case_sensitive: Whether search should be case-sensitive

    Returns:
        List of dicts with keys: sheet_name, row, column, cell_reference, value

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large or unreadable

    Example:
        >>> matches = search_excel_text("/path/to/file.xlsx", "Python", False)
        >>> matches[0]
        {'sheet_name': 'Sheet1', 'row': 5, 'column': 2, 'cell_reference': 'B5', 'value': 'Python Developer'}
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(search_term, str):
        raise TypeError("search_term must be a string")

    if not isinstance(case_sensitive, bool):
        raise TypeError("case_sensitive must be a boolean")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        matches = []

        # Prepare search term for comparison
        search_compare = search_term if case_sensitive else search_term.lower()

        # Search all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value is None:
                        continue

                    cell_str = str(cell.value)
                    cell_compare = cell_str if case_sensitive else cell_str.lower()

                    if search_compare in cell_compare:
                        matches.append(
                            {
                                "sheet_name": sheet_name,
                                "row": row_idx,
                                "column": col_idx,
                                "cell_reference": cell.coordinate,
                                "value": cell_str,
                            }
                        )

        wb.close()
        return matches

    except Exception as e:
        raise ValueError(f"Failed to search Excel file: {e}")


@strands_tool
def get_excel_metadata(file_path: str) -> dict[str, str]:
    """Get Excel workbook metadata and properties.

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary containing metadata (creator, title, subject, etc.)

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large or unreadable

    Example:
        >>> metadata = get_excel_metadata("/path/to/file.xlsx")
        >>> metadata['creator']
        'John Doe'
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True)
        props = wb.properties

        metadata = {
            "creator": str(props.creator) if props.creator else "",
            "title": str(props.title) if props.title else "",
            "subject": str(props.subject) if props.subject else "",
            "description": str(props.description) if props.description else "",
            "keywords": str(props.keywords) if props.keywords else "",
            "category": str(props.category) if props.category else "",
            "last_modified_by": str(props.lastModifiedBy)
            if props.lastModifiedBy
            else "",
            "created": str(props.created) if props.created else "",
            "modified": str(props.modified) if props.modified else "",
        }

        wb.close()
        return metadata

    except Exception as e:
        raise ValueError(f"Failed to get Excel metadata: {e}")


@strands_tool
def get_excel_info(file_path: str) -> dict[str, object]:
    """Get comprehensive information about Excel workbook.

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary with file info, sheets, dimensions, formula count, etc.

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large or unreadable

    Example:
        >>> info = get_excel_info("/path/to/file.xlsx")
        >>> info['sheet_count']
        3
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True)

        # Get sheet information
        sheet_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info.append(
                {
                    "name": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                }
            )

        info = {
            "file_path": file_path,
            "file_size_bytes": file_size,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheet_info,
        }

        wb.close()
        return info

    except Exception as e:
        raise ValueError(f"Failed to get Excel info: {e}")


@strands_tool
def get_sheet_info(file_path: str, sheet_name: str) -> dict[str, str]:
    """Get metadata for specific Excel sheet without loading data.

    This function efficiently retrieves sheet dimensions and metadata
    without loading cell data, saving tokens when you only need
    information about one sheet.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to inspect

    Returns:
        Dictionary with sheet metadata: row_count, column_count, file_size

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, or sheet doesn't exist

    Example:
        >>> info = get_sheet_info("/path/to/file.xlsx", "Sheet1")
        >>> info['row_count']
        '1000'
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        info = {
            "sheet_name": sheet_name,
            "row_count": str(ws.max_row),
            "column_count": str(ws.max_column),
            "file_size_bytes": str(file_size),
        }

        wb.close()
        return info

    except Exception as e:
        raise ValueError(f"Failed to get sheet info: {e}")


@strands_tool
def get_sheet_schema(
    file_path: str, sheet_name: str, sample_rows: int
) -> dict[str, str]:
    """Get Excel sheet schema with inferred data types by sampling rows.

    This function efficiently determines column types by examining
    a sample of rows without loading the entire sheet.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to analyze
        sample_rows: Number of rows to sample for type inference (after header)

    Returns:
        Dictionary mapping column names to inferred types
        (integer, float, boolean, string)

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, sheet doesn't exist, or no data

    Example:
        >>> schema = get_sheet_schema("/path/to/file.xlsx", "Sheet1", 100)
        >>> schema
        {'Name': 'string', 'Age': 'integer', 'Salary': 'float'}
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(sample_rows, int) or sample_rows < 1:
        raise TypeError("sample_rows must be a positive integer")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get headers from first row
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        if not headers or all(h == "" for h in headers):
            raise ValueError("Sheet has no header row")

        # Sample data rows for type inference
        sample_data = []
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row_count >= sample_rows:
                break
            sample_data.append(row)
            row_count += 1

        if not sample_data:
            raise ValueError("Sheet has no data rows")

        # Infer types for each column
        schema = {}
        for col_idx, header in enumerate(headers):
            # Collect non-None values for this column
            values = [
                row[col_idx]
                for row in sample_data
                if col_idx < len(row) and row[col_idx] is not None
            ]

            if not values:
                schema[header] = "string"
                continue

            # Type inference logic
            all_bool = all(isinstance(v, bool) for v in values)
            all_int = all(
                isinstance(v, int) and not isinstance(v, bool) for v in values
            )
            all_float = all(
                isinstance(v, (int, float)) and not isinstance(v, bool) for v in values
            )

            if all_bool:
                schema[header] = "boolean"
            elif all_int:
                schema[header] = "integer"
            elif all_float:
                schema[header] = "float"
            else:
                schema[header] = "string"

        wb.close()
        return schema

    except Exception as e:
        raise ValueError(f"Failed to get sheet schema: {e}")


@strands_tool
def preview_sheet_rows(
    file_path: str, sheet_name: str, num_rows: int
) -> list[dict[str, str]]:
    """Get first N rows of Excel sheet for preview without loading entire sheet.

    This function efficiently reads only the requested number of rows,
    making it ideal for large sheets where you want a quick preview.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to preview
        num_rows: Number of data rows to preview (not including header)

    Returns:
        List of dictionaries, each representing a row with column headers as keys

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, sheet doesn't exist, or no data

    Example:
        >>> rows = preview_sheet_rows("/path/to/file.xlsx", "Sheet1", 5)
        >>> rows[0]
        {'Name': 'Alice', 'Age': '30', 'City': 'NYC'}
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(num_rows, int) or num_rows < 1:
        raise TypeError("num_rows must be a positive integer")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        if not headers:
            raise ValueError("Sheet has no header row")

        # Read preview rows
        result = []
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row_count >= num_rows:
                break

            row_dict = {}
            for idx, header in enumerate(headers):
                value = row[idx] if idx < len(row) and row[idx] is not None else ""
                row_dict[header] = str(value)

            result.append(row_dict)
            row_count += 1

        wb.close()
        return result

    except Exception as e:
        raise ValueError(f"Failed to preview sheet rows: {e}")


@strands_tool
def select_sheet_columns(
    file_path: str, sheet_name: str, columns: list[str]
) -> list[dict[str, str]]:
    """Read only specific columns from Excel sheet, discarding others.

    This function efficiently reads only the requested columns,
    making it ideal for wide sheets where you only need a subset of data.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to read
        columns: List of column names to include

    Returns:
        List of dictionaries with only the specified columns

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, sheet doesn't exist, or columns not found

    Example:
        >>> rows = select_sheet_columns("/path/to/file.xlsx", "Sheet1", ["Name", "Age"])
        >>> rows[0]
        {'Name': 'Alice', 'Age': '30'}
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(columns, list) or not all(isinstance(c, str) for c in columns):
        raise TypeError("columns must be a list of strings")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        # Find column indices
        col_indices = {}
        for col in columns:
            try:
                col_indices[col] = headers.index(col)
            except ValueError:
                raise ValueError(
                    f"Column '{col}' not found in sheet. Available: {headers}"
                )

        # Read data with only selected columns
        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for col in columns:
                idx = col_indices[col]
                value = row[idx] if idx < len(row) and row[idx] is not None else ""
                row_dict[col] = str(value)
            result.append(row_dict)

        wb.close()
        return result

    except Exception as e:
        raise ValueError(f"Failed to select sheet columns: {e}")


@strands_tool
def filter_sheet_rows(
    file_path: str, sheet_name: str, column: str, value: str, operator: str
) -> list[dict[str, str]]:
    """Read only Excel rows matching filter criteria.

    This function efficiently filters rows based on criteria,
    avoiding loading rows that don't match.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to filter
        column: Column name to filter on
        value: Value to compare against
        operator: Comparison operator (equals, contains, startswith, endswith,
                 greater_than, less_than)

    Returns:
        List of dictionaries containing only matching rows

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, sheet doesn't exist, column not found,
                   or invalid operator

    Example:
        >>> rows = filter_sheet_rows("/path/to/file.xlsx", "Sheet1", "Age", "30", "greater_than")
        >>> rows[0]
        {'Name': 'Bob', 'Age': '35', 'City': 'LA'}
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(column, str):
        raise TypeError("column must be a string")

    if not isinstance(value, str):
        raise TypeError("value must be a string")

    if not isinstance(operator, str):
        raise TypeError("operator must be a string")

    valid_operators = [
        "equals",
        "contains",
        "startswith",
        "endswith",
        "greater_than",
        "less_than",
    ]
    if operator not in valid_operators:
        raise ValueError(f"operator must be one of: {valid_operators}")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        # Find column index
        try:
            col_idx = headers.index(column)
        except ValueError:
            raise ValueError(
                f"Column '{column}' not found in sheet. Available: {headers}"
            )

        # Filter rows
        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            cell_value = (
                str(row[col_idx])
                if col_idx < len(row) and row[col_idx] is not None
                else ""
            )

            # Apply filter
            matches = False
            if operator == "equals":
                matches = cell_value == value
            elif operator == "contains":
                matches = value in cell_value
            elif operator == "startswith":
                matches = cell_value.startswith(value)
            elif operator == "endswith":
                matches = cell_value.endswith(value)
            elif operator == "greater_than":
                try:
                    matches = float(cell_value) > float(value)
                except ValueError:
                    matches = False
            elif operator == "less_than":
                try:
                    matches = float(cell_value) < float(value)
                except ValueError:
                    matches = False

            if matches:
                row_dict = {}
                for idx, header in enumerate(headers):
                    row_value = (
                        row[idx] if idx < len(row) and row[idx] is not None else ""
                    )
                    row_dict[header] = str(row_value)
                result.append(row_dict)

        wb.close()
        return result

    except Exception as e:
        raise ValueError(f"Failed to filter sheet rows: {e}")


@strands_tool
def get_sheet_row_range(
    file_path: str, sheet_name: str, start_row: int, end_row: int
) -> list[dict[str, str]]:
    """Get specific range of rows from Excel sheet (pagination support).

    This function efficiently reads only the requested range of rows,
    ideal for implementing pagination or processing sheets in chunks.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to read
        start_row: Starting row number (1-based, not including header)
        end_row: Ending row number (1-based, inclusive)

    Returns:
        List of dictionaries for rows in the specified range

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, sheet doesn't exist, or invalid range

    Example:
        >>> rows = get_sheet_row_range("/path/to/file.xlsx", "Sheet1", 10, 20)
        >>> len(rows)
        11
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(start_row, int) or start_row < 1:
        raise TypeError("start_row must be a positive integer")

    if not isinstance(end_row, int) or end_row < start_row:
        raise TypeError("end_row must be an integer >= start_row")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        # Read range of rows
        result = []
        row_num = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if row_num < start_row:
                continue
            if row_num > end_row:
                break

            row_dict = {}
            for idx, header in enumerate(headers):
                value = row[idx] if idx < len(row) and row[idx] is not None else ""
                row_dict[header] = str(value)
            result.append(row_dict)

        wb.close()
        return result

    except Exception as e:
        raise ValueError(f"Failed to get sheet row range: {e}")


@strands_tool
def sample_sheet_rows(
    file_path: str, sheet_name: str, sample_size: int, method: str
) -> list[dict[str, str]]:
    """Get a representative sample of Excel sheet rows without loading entire sheet.

    This function provides different sampling strategies to get a
    representative subset of data efficiently.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to sample
        sample_size: Number of rows to include in sample
        method: Sampling method (first, random, systematic)
                first: First N rows
                random: Random N rows using reservoir sampling
                systematic: Every Kth row to reach N rows

    Returns:
        List of dictionaries containing sampled rows

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, sheet doesn't exist, or invalid method

    Example:
        >>> rows = sample_sheet_rows("/path/to/file.xlsx", "Sheet1", 100, "random")
        >>> len(rows)
        100
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(sample_size, int) or sample_size < 1:
        raise TypeError("sample_size must be a positive integer")

    if not isinstance(method, str):
        raise TypeError("method must be a string")

    valid_methods = ["first", "random", "systematic"]
    if method not in valid_methods:
        raise ValueError(f"method must be one of: {valid_methods}")

    if method == "first":
        # Just use preview_sheet_rows for first N rows
        result_preview: list[dict[str, str]] = preview_sheet_rows(
            file_path, sheet_name, sample_size
        )
        return result_preview

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        if method == "random":
            # Reservoir sampling
            import random

            result: list[dict[str, str]] = []
            row_num = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for idx, header in enumerate(headers):
                    value = row[idx] if idx < len(row) and row[idx] is not None else ""
                    row_dict[header] = str(value)

                if row_num < sample_size:
                    result.append(row_dict)
                else:
                    # Reservoir sampling algorithm
                    j = random.randint(0, row_num)
                    if j < sample_size:
                        result[j] = row_dict

                row_num += 1

            wb.close()
            return result

        else:  # systematic
            # Get total rows first
            total_rows = ws.max_row - 1  # Exclude header
            if total_rows <= sample_size:
                # Return all rows if total is less than sample size
                result_all = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        value = (
                            row[idx] if idx < len(row) and row[idx] is not None else ""
                        )
                        row_dict[header] = str(value)
                    result_all.append(row_dict)
                wb.close()
                return result_all

            # Calculate step size
            step = total_rows // sample_size

            result_systematic = []
            row_num = 0
            next_sample = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row_num == next_sample:
                    row_dict = {}
                    for idx, header in enumerate(headers):
                        value = (
                            row[idx] if idx < len(row) and row[idx] is not None else ""
                        )
                        row_dict[header] = str(value)
                    result_systematic.append(row_dict)
                    next_sample += step

                    if len(result_systematic) >= sample_size:
                        break

                row_num += 1

            wb.close()
            return result_systematic

    except Exception as e:
        raise ValueError(f"Failed to sample sheet rows: {e}")


@strands_tool
def get_sheet_column_stats(
    file_path: str, sheet_name: str, column: str, sample_size: int
) -> dict[str, str]:
    """Get statistics for an Excel sheet column by sampling rows.

    This function efficiently computes statistics by sampling,
    avoiding the need to load all data for large sheets.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        column: Column name to analyze
        sample_size: Number of rows to sample for statistics

    Returns:
        Dictionary with statistics: total_rows, unique_count, null_count,
        sample_values, min_value (if numeric), max_value (if numeric)

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, sheet doesn't exist, or column not found

    Example:
        >>> stats = get_sheet_column_stats("/path/to/file.xlsx", "Sheet1", "Age", 1000)
        >>> stats['unique_count']
        '45'
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(column, str):
        raise TypeError("column must be a string")

    if not isinstance(sample_size, int) or sample_size < 1:
        raise TypeError("sample_size must be a positive integer")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        # Find column index
        try:
            col_idx = headers.index(column)
        except ValueError:
            raise ValueError(
                f"Column '{column}' not found in sheet. Available: {headers}"
            )

        # Collect sample values
        values = []
        null_count = 0
        row_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row_count >= sample_size:
                break

            if col_idx < len(row) and row[col_idx] is not None:
                values.append(row[col_idx])
            else:
                null_count += 1

            row_count += 1

        # Get total row count
        total_rows = ws.max_row - 1  # Exclude header

        # Calculate statistics
        unique_values = {str(v) for v in values}
        stats = {
            "total_rows": str(total_rows),
            "unique_count": str(len(unique_values)),
            "null_count": str(null_count),
            "sample_size": str(len(values)),
            "sample_values": ", ".join(str(v) for v in list(unique_values)[:5]),
        }

        # Try to calculate min/max for numeric columns
        try:
            numeric_values = [float(v) for v in values if v != ""]
            if numeric_values:
                stats["min_value"] = str(min(numeric_values))
                stats["max_value"] = str(max(numeric_values))
        except (ValueError, TypeError):
            pass  # Not a numeric column

        wb.close()
        return stats

    except Exception as e:
        raise ValueError(f"Failed to get sheet column stats: {e}")


@strands_tool
def count_sheet_rows(
    file_path: str, sheet_name: str, filter_column: str, filter_value: str
) -> int:
    """Count Excel sheet rows, optionally matching filter criteria.

    This function efficiently counts rows without loading all data,
    ideal for large sheets where you need row counts.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to count
        filter_column: Column name to filter on (empty string for no filter)
        filter_value: Value to match (ignored if filter_column is empty)

    Returns:
        Count of matching rows

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, sheet doesn't exist, or column not found

    Example:
        >>> count = count_sheet_rows("/path/to/file.xlsx", "Sheet1", "", "")
        >>> count
        1000
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(filter_column, str):
        raise TypeError("filter_column must be a string")

    if not isinstance(filter_value, str):
        raise TypeError("filter_value must be a string")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # No filter - return total count
        if not filter_column:
            count_total: int = ws.max_row - 1  # Exclude header
            wb.close()
            return count_total

        # With filter - count matching rows
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        # Find column index
        try:
            col_idx = headers.index(filter_column)
        except ValueError:
            raise ValueError(
                f"Column '{filter_column}' not found in sheet. Available: {headers}"
            )

        # Count matching rows
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            cell_value = (
                str(row[col_idx])
                if col_idx < len(row) and row[col_idx] is not None
                else ""
            )
            if cell_value == filter_value:
                count += 1

        wb.close()
        return count

    except Exception as e:
        raise ValueError(f"Failed to count sheet rows: {e}")


@strands_tool
def get_sheet_value_counts(
    file_path: str, sheet_name: str, column: str, top_n: int
) -> dict[str, str]:
    """Get frequency counts for Excel column values (top N most common).

    This function efficiently computes value distributions,
    useful for understanding categorical data without loading all rows.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        column: Column name to analyze
        top_n: Number of top values to return

    Returns:
        Dictionary mapping values to their frequency counts (top N)

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If file too large, unreadable, sheet doesn't exist, or column not found

    Example:
        >>> counts = get_sheet_value_counts("/path/to/file.xlsx", "Sheet1", "City", 5)
        >>> counts
        {'NYC': '150', 'LA': '120', 'Chicago': '90'}
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(column, str):
        raise TypeError("column must be a string")

    if not isinstance(top_n, int) or top_n < 1:
        raise TypeError("top_n must be a positive integer")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check file size
    file_size = os.path.getsize(file_path)
    if file_size > MAX_FILE_SIZE:
        raise ValueError(
            f"File too large: {file_size} bytes (max {MAX_FILE_SIZE} bytes)"
        )

    try:
        from collections import Counter

        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Get headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
            headers.append(str(cell) if cell is not None else "")

        # Find column index
        try:
            col_idx = headers.index(column)
        except ValueError:
            raise ValueError(
                f"Column '{column}' not found in sheet. Available: {headers}"
            )

        # Count values
        values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if col_idx < len(row) and row[col_idx] is not None:
                values.append(str(row[col_idx]))

        counter = Counter(values)
        top_values = dict(counter.most_common(top_n))

        # Convert to string values
        result = {k: str(v) for k, v in top_values.items()}

        wb.close()
        return result

    except Exception as e:
        raise ValueError(f"Failed to get sheet value counts: {e}")
