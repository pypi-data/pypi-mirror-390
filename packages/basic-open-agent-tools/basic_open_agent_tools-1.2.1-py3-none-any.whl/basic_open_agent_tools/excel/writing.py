"""Excel spreadsheet creation and modification functions for AI agents.

This module provides functions for creating and modifying Excel (.xlsx) spreadsheets.
"""

import csv
import os

from ..decorators import strands_tool

try:
    from openpyxl import (  # type: ignore[import-untyped, import-not-found]
        Workbook,
        load_workbook,
    )

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Maximum file size: 100MB
MAX_FILE_SIZE = 100 * 1024 * 1024


@strands_tool
def create_simple_excel(
    file_path: str, data: list[list[str]], skip_confirm: bool
) -> str:
    """Create simple Excel workbook with single sheet from 2D list.

    This function creates a new Excel file with a single sheet containing
    the provided data.

    Args:
        file_path: Path for new Excel file
        data: 2D list of strings (rows and columns)
        skip_confirm: If False, raises error if file exists; if True, overwrites

    Returns:
        Success message with file path

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        ValueError: If file exists and skip_confirm is False, or data invalid
        PermissionError: If lacking write permission

    Example:
        >>> data = [['Name', 'Age'], ['Alice', '30'], ['Bob', '25']]
        >>> msg = create_simple_excel("/tmp/data.xlsx", data, True)
        >>> "Created" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(data, list):
        raise TypeError("data must be a list")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Validate data structure
    if not data:
        raise ValueError("data must not be empty")

    for row in data:
        if not isinstance(row, list):
            raise TypeError("each row in data must be a list")

    # Check if file exists
    if os.path.exists(file_path) and not skip_confirm:
        raise ValueError(
            f"File already exists: {file_path}. Set skip_confirm=True to overwrite."
        )

    # Check parent directory exists
    parent_dir = os.path.dirname(file_path)
    if parent_dir and not os.path.exists(parent_dir):
        raise ValueError(f"Parent directory does not exist: {parent_dir}")

    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active

        # Write data
        for row_data in data:
            ws.append(row_data)

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Created Excel file at {file_path} with {len(data)} rows"

    except PermissionError:
        raise PermissionError(f"Permission denied writing to: {file_path}")
    except Exception as e:
        raise ValueError(f"Failed to create Excel file: {e}")


@strands_tool
def create_excel_with_headers(
    file_path: str, headers: list[str], data: list[list[str]], skip_confirm: bool
) -> str:
    """Create Excel workbook with header row and data.

    Args:
        file_path: Path for new Excel file
        headers: List of header strings for first row
        data: 2D list of data rows
        skip_confirm: If False, raises error if file exists; if True, overwrites

    Returns:
        Success message with file path

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        ValueError: If file exists and skip_confirm is False, or headers/data invalid
        PermissionError: If lacking write permission

    Example:
        >>> headers = ['Name', 'Age', 'City']
        >>> data = [['Alice', '30', 'NYC'], ['Bob', '25', 'LA']]
        >>> msg = create_excel_with_headers("/tmp/data.xlsx", headers, data, True)
        >>> "Created" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(headers, list):
        raise TypeError("headers must be a list")

    if not isinstance(data, list):
        raise TypeError("data must be a list")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Validate headers
    if not headers:
        raise ValueError("headers must not be empty")

    for header in headers:
        if not isinstance(header, str):
            raise TypeError("each header must be a string")

    # Combine headers and data
    combined_data = [headers] + data

    return create_simple_excel(file_path, combined_data, skip_confirm)


@strands_tool
def create_excel_from_dicts(
    file_path: str, data: list[dict[str, str]], skip_confirm: bool
) -> str:
    """Create Excel workbook from list of dictionaries.

    This function creates an Excel file where the dictionary keys become
    headers and values populate the rows.

    Args:
        file_path: Path for new Excel file
        data: List of dictionaries with consistent keys
        skip_confirm: If False, raises error if file exists; if True, overwrites

    Returns:
        Success message with file path

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        ValueError: If file exists and skip_confirm is False, or data invalid
        PermissionError: If lacking write permission

    Example:
        >>> data = [{'Name': 'Alice', 'Age': '30'}, {'Name': 'Bob', 'Age': '25'}]
        >>> msg = create_excel_from_dicts("/tmp/data.xlsx", data, True)
        >>> "Created" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(data, list):
        raise TypeError("data must be a list")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Validate data
    if not data:
        raise ValueError("data must not be empty")

    for item in data:
        if not isinstance(item, dict):
            raise TypeError("each item in data must be a dict")

    # Extract headers from first dict
    headers = list(data[0].keys())

    # Convert dicts to rows
    rows = []
    for item in data:
        row = [item.get(key, "") for key in headers]
        rows.append(row)

    return create_excel_with_headers(file_path, headers, rows, skip_confirm)


@strands_tool
def add_sheet_to_excel(
    file_path: str, sheet_name: str, data: list[list[str]], skip_confirm: bool
) -> str:
    """Add new sheet to existing Excel workbook.

    Args:
        file_path: Path to existing Excel file
        sheet_name: Name for new sheet
        data: 2D list of data for new sheet
        skip_confirm: If False, raises error if sheet exists; if True, overwrites

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet exists and skip_confirm is False, or data invalid

    Example:
        >>> data = [['Product', 'Price'], ['Apple', '1.00']]
        >>> msg = add_sheet_to_excel("/tmp/data.xlsx", "Products", data, True)
        >>> "Added sheet" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(data, list):
        raise TypeError("data must be a list")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Validate data
    if not data:
        raise ValueError("data must not be empty")

    for row in data:
        if not isinstance(row, list):
            raise TypeError("each row in data must be a list")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check if sheet exists
        if sheet_name in wb.sheetnames and not skip_confirm:
            raise ValueError(
                f"Sheet '{sheet_name}' already exists. Set skip_confirm=True to overwrite."
            )

        # Remove sheet if it exists and skip_confirm is True
        if sheet_name in wb.sheetnames and skip_confirm:
            del wb[sheet_name]

        # Create new sheet
        ws = wb.create_sheet(title=sheet_name)

        # Write data
        for row_data in data:
            ws.append(row_data)

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Added sheet '{sheet_name}' to {file_path} with {len(data)} rows"

    except Exception as e:
        raise ValueError(f"Failed to add sheet: {e}")


@strands_tool
def append_rows_to_excel(
    file_path: str, sheet_name: str, rows: list[list[str]], skip_confirm: bool
) -> str:
    """Append rows to existing sheet in Excel workbook.

    Args:
        file_path: Path to existing Excel file
        sheet_name: Name of sheet to append to
        rows: 2D list of rows to append
        skip_confirm: Required for consistency (not used for append operations)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet doesn't exist or rows invalid

    Example:
        >>> rows = [['Charlie', '35', 'SF'], ['Dana', '28', 'LA']]
        >>> msg = append_rows_to_excel("/tmp/data.xlsx", "Sheet1", rows, True)
        >>> "Appended" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(rows, list):
        raise TypeError("rows must be a list")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Validate rows
    if not rows:
        raise ValueError("rows must not be empty")

    for row in rows:
        if not isinstance(row, list):
            raise TypeError("each row must be a list")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Append rows
        for row_data in rows:
            ws.append(row_data)

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Appended {len(rows)} rows to sheet '{sheet_name}' in {file_path}"

    except Exception as e:
        raise ValueError(f"Failed to append rows: {e}")


@strands_tool
def update_excel_cell(
    file_path: str,
    sheet_name: str,
    cell_reference: str,
    value: str,
    skip_confirm: bool,
) -> str:
    """Update single cell value in Excel sheet.

    Args:
        file_path: Path to existing Excel file
        sheet_name: Name of sheet
        cell_reference: Cell reference in A1 notation (e.g., "B5")
        value: New value for cell
        skip_confirm: Required for consistency (not used for cell updates)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet doesn't exist or cell_reference invalid

    Example:
        >>> msg = update_excel_cell("/tmp/data.xlsx", "Sheet1", "B5", "Updated", True)
        >>> "Updated cell" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(cell_reference, str):
        raise TypeError("cell_reference must be a string")

    if not isinstance(value, str):
        raise TypeError("value must be a string")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Update cell
        ws[cell_reference] = value

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Updated cell {cell_reference} in sheet '{sheet_name}' to '{value}'"

    except KeyError:
        raise ValueError(f"Invalid cell reference: {cell_reference}")
    except Exception as e:
        raise ValueError(f"Failed to update cell: {e}")


@strands_tool
def delete_excel_sheet(file_path: str, sheet_name: str, skip_confirm: bool) -> str:
    """Delete sheet from Excel workbook.

    Args:
        file_path: Path to existing Excel file
        sheet_name: Name of sheet to delete
        skip_confirm: If False, raises error for confirmation; if True, deletes

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet doesn't exist, is last sheet, or skip_confirm is False

    Example:
        >>> msg = delete_excel_sheet("/tmp/data.xlsx", "OldSheet", True)
        >>> "Deleted sheet" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    if not skip_confirm:
        raise ValueError(
            f"Deletion requires confirmation. Set skip_confirm=True to delete sheet '{sheet_name}'."
        )

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        # Check not last sheet
        if len(wb.sheetnames) == 1:
            raise ValueError("Cannot delete the last sheet in workbook")

        # Delete sheet
        del wb[sheet_name]

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Deleted sheet '{sheet_name}' from {file_path}"

    except Exception as e:
        raise ValueError(f"Failed to delete sheet: {e}")


@strands_tool
def excel_to_csv(
    file_path: str, sheet_name: str, output_path: str, skip_confirm: bool
) -> str:
    """Export Excel sheet to CSV file.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to export
        output_path: Path for output CSV file
        skip_confirm: If False, raises error if CSV exists; if True, overwrites

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If sheet doesn't exist or CSV exists and skip_confirm is False
        PermissionError: If lacking write permission

    Example:
        >>> msg = excel_to_csv("/tmp/data.xlsx", "Sheet1", "/tmp/data.csv", True)
        >>> "Exported" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(output_path, str):
        raise TypeError("output_path must be a string")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Check Excel file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    # Check if CSV exists
    if os.path.exists(output_path) and not skip_confirm:
        raise ValueError(
            f"CSV file already exists: {output_path}. Set skip_confirm=True to overwrite."
        )

    # Check parent directory exists
    parent_dir = os.path.dirname(output_path)
    if parent_dir and not os.path.exists(parent_dir):
        raise ValueError(f"Parent directory does not exist: {parent_dir}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Write to CSV
        with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)

            for row in ws.iter_rows(values_only=True):
                # Convert None to empty string
                csv_row = [str(cell) if cell is not None else "" for cell in row]
                writer.writerow(csv_row)

        wb.close()

        return f"Exported sheet '{sheet_name}' from {file_path} to {output_path}"

    except PermissionError:
        raise PermissionError(f"Permission denied writing to: {output_path}")
    except Exception as e:
        raise ValueError(f"Failed to export to CSV: {e}")
