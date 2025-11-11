"""Excel spreadsheet formatting and styling functions for AI agents.

This module provides functions for applying formatting and styles to
Excel (.xlsx) spreadsheets.
"""

import os
import re

from ..decorators import strands_tool

try:
    from openpyxl import load_workbook  # type: ignore[import-untyped, import-not-found]
    from openpyxl.styles import (  # type: ignore[import-untyped, import-not-found]
        Alignment,
        Font,
        PatternFill,
    )

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@strands_tool
def apply_excel_bold(
    file_path: str, sheet_name: str, cell_range: str, skip_confirm: bool
) -> str:
    """Apply bold formatting to cells in range.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        cell_range: Cell range in A1 notation (e.g., "A1:C10" or single "B5")
        skip_confirm: Required for consistency (not used for formatting)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet or cell_range invalid

    Example:
        >>> msg = apply_excel_bold("/tmp/data.xlsx", "Sheet1", "A1:A10", True)
        >>> "Applied bold" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(cell_range, str):
        raise TypeError("cell_range must be a string")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Apply bold to range
        for row in ws[cell_range]:
            # Handle single cell or range
            cells = row if isinstance(row, tuple) else [row]
            for cell in cells:
                cell.font = Font(bold=True)

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Applied bold formatting to {cell_range} in sheet '{sheet_name}'"

    except Exception as e:
        raise ValueError(f"Failed to apply bold formatting: {e}")


@strands_tool
def apply_excel_font_size(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    font_size: int,
    skip_confirm: bool,
) -> str:
    """Set font size for cells in range.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        cell_range: Cell range in A1 notation
        font_size: Font size in points (e.g., 12)
        skip_confirm: Required for consistency (not used for formatting)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet, cell_range, or font_size invalid

    Example:
        >>> msg = apply_excel_font_size("/tmp/data.xlsx", "Sheet1", "A1", 14, True)
        >>> "Set font size" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(cell_range, str):
        raise TypeError("cell_range must be a string")

    if not isinstance(font_size, int):
        raise TypeError("font_size must be an integer")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    if font_size < 1 or font_size > 409:
        raise ValueError("font_size must be between 1 and 409")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Apply font size to range
        for row in ws[cell_range]:
            cells = row if isinstance(row, tuple) else [row]
            for cell in cells:
                cell.font = Font(size=font_size)

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Set font size to {font_size} for {cell_range} in sheet '{sheet_name}'"

    except Exception as e:
        raise ValueError(f"Failed to set font size: {e}")


@strands_tool
def apply_excel_alignment(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    horizontal: str,
    vertical: str,
    skip_confirm: bool,
) -> str:
    """Set cell alignment for range.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        cell_range: Cell range in A1 notation
        horizontal: Horizontal alignment (left, center, right, justify)
        vertical: Vertical alignment (top, center, bottom, justify)
        skip_confirm: Required for consistency (not used for formatting)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet, cell_range, or alignment invalid

    Example:
        >>> msg = apply_excel_alignment("/tmp/data.xlsx", "Sheet1", "A1", "center", "center", True)
        >>> "Set alignment" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(cell_range, str):
        raise TypeError("cell_range must be a string")

    if not isinstance(horizontal, str):
        raise TypeError("horizontal must be a string")

    if not isinstance(vertical, str):
        raise TypeError("vertical must be a string")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Validate alignment values
    valid_horizontal = ["left", "center", "right", "justify"]
    valid_vertical = ["top", "center", "bottom", "justify"]

    if horizontal.lower() not in valid_horizontal:
        raise ValueError(
            f"horizontal must be one of {valid_horizontal}, got '{horizontal}'"
        )

    if vertical.lower() not in valid_vertical:
        raise ValueError(f"vertical must be one of {valid_vertical}, got '{vertical}'")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Apply alignment to range
        for row in ws[cell_range]:
            cells = row if isinstance(row, tuple) else [row]
            for cell in cells:
                cell.alignment = Alignment(
                    horizontal=horizontal.lower(), vertical=vertical.lower()
                )

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Set alignment to {horizontal}/{vertical} for {cell_range} in sheet '{sheet_name}'"

    except Exception as e:
        raise ValueError(f"Failed to set alignment: {e}")


@strands_tool
def set_excel_column_width(
    file_path: str, sheet_name: str, column_letter: str, width: int, skip_confirm: bool
) -> str:
    """Set column width in Excel sheet.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        column_letter: Column letter (e.g., "A", "B", "AA")
        width: Width in character units (typically 8-100)
        skip_confirm: Required for consistency (not used for formatting)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet or column invalid, or width out of range

    Example:
        >>> msg = set_excel_column_width("/tmp/data.xlsx", "Sheet1", "A", 20, True)
        >>> "Set column width" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(column_letter, str):
        raise TypeError("column_letter must be a string")

    if not isinstance(width, int):
        raise TypeError("width must be an integer")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    if width < 0 or width > 255:
        raise ValueError("width must be between 0 and 255")

    # Validate column letter format
    if not re.match(r"^[A-Z]+$", column_letter.upper()):
        raise ValueError(f"Invalid column letter: {column_letter}")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Set column width
        ws.column_dimensions[column_letter.upper()].width = width

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Set column {column_letter} width to {width} in sheet '{sheet_name}'"

    except Exception as e:
        raise ValueError(f"Failed to set column width: {e}")


@strands_tool
def set_excel_row_height(
    file_path: str, sheet_name: str, row_number: int, height: int, skip_confirm: bool
) -> str:
    """Set row height in Excel sheet.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        row_number: Row number (1-indexed)
        height: Height in points (typically 12-400)
        skip_confirm: Required for consistency (not used for formatting)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet or row invalid, or height out of range

    Example:
        >>> msg = set_excel_row_height("/tmp/data.xlsx", "Sheet1", 1, 30, True)
        >>> "Set row height" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(row_number, int):
        raise TypeError("row_number must be an integer")

    if not isinstance(height, int):
        raise TypeError("height must be an integer")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    if row_number < 1:
        raise ValueError("row_number must be >= 1")

    if height < 0 or height > 409:
        raise ValueError("height must be between 0 and 409")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Set row height
        ws.row_dimensions[row_number].height = height

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Set row {row_number} height to {height} in sheet '{sheet_name}'"

    except Exception as e:
        raise ValueError(f"Failed to set row height: {e}")


@strands_tool
def apply_excel_cell_color(
    file_path: str,
    sheet_name: str,
    cell_range: str,
    color_hex: str,
    skip_confirm: bool,
) -> str:
    """Apply background color to cells in range.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        cell_range: Cell range in A1 notation
        color_hex: Hex color code without # (e.g., "FFFF00" for yellow)
        skip_confirm: Required for consistency (not used for formatting)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet, cell_range, or color_hex invalid

    Example:
        >>> msg = apply_excel_cell_color("/tmp/data.xlsx", "Sheet1", "A1", "FFFF00", True)
        >>> "Applied color" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(cell_range, str):
        raise TypeError("cell_range must be a string")

    if not isinstance(color_hex, str):
        raise TypeError("color_hex must be a string")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Validate hex color (6 characters, hex digits)
    if not re.match(r"^[0-9A-Fa-f]{6}$", color_hex):
        raise ValueError(f"color_hex must be 6 hex digits without #, got '{color_hex}'")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Apply color to range
        fill = PatternFill(
            start_color=color_hex.upper(),
            end_color=color_hex.upper(),
            fill_type="solid",
        )

        for row in ws[cell_range]:
            cells = row if isinstance(row, tuple) else [row]
            for cell in cells:
                cell.fill = fill

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Applied color {color_hex} to {cell_range} in sheet '{sheet_name}'"

    except Exception as e:
        raise ValueError(f"Failed to apply cell color: {e}")


@strands_tool
def freeze_excel_panes(
    file_path: str, sheet_name: str, cell_reference: str, skip_confirm: bool
) -> str:
    """Freeze panes at specified cell in Excel sheet.

    Rows above and columns to left of the cell will be frozen.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        cell_reference: Cell at which to freeze (e.g., "B2" freezes row 1 and column A)
        skip_confirm: Required for consistency (not used for formatting)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet or cell_reference invalid

    Example:
        >>> msg = freeze_excel_panes("/tmp/data.xlsx", "Sheet1", "B2", True)
        >>> "Froze panes" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(cell_reference, str):
        raise TypeError("cell_reference must be a string")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Freeze panes
        ws.freeze_panes = cell_reference

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Froze panes at {cell_reference} in sheet '{sheet_name}'"

    except Exception as e:
        raise ValueError(f"Failed to freeze panes: {e}")


@strands_tool
def add_excel_formula(
    file_path: str,
    sheet_name: str,
    cell_reference: str,
    formula: str,
    skip_confirm: bool,
) -> str:
    """Add formula to cell in Excel sheet.

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
        cell_reference: Cell reference for formula (e.g., "C1")
        formula: Excel formula (e.g., "=SUM(A1:A10)" or "=A1+B1")
        skip_confirm: Required for consistency (not used for formulas)

    Returns:
        Success message

    Raises:
        ImportError: If openpyxl is not installed
        TypeError: If parameters are wrong type
        FileNotFoundError: If file doesn't exist
        ValueError: If sheet, cell_reference, or formula invalid

    Example:
        >>> msg = add_excel_formula("/tmp/data.xlsx", "Sheet1", "C1", "=SUM(A1:A10)", True)
        >>> "Added formula" in msg
        True
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel operations. "
            "Install with: pip install basic-open-agent-tools[excel]"
        )

    if not isinstance(file_path, str):
        raise TypeError("file_path must be a string")

    if not isinstance(sheet_name, str):
        raise TypeError("sheet_name must be a string")

    if not isinstance(cell_reference, str):
        raise TypeError("cell_reference must be a string")

    if not isinstance(formula, str):
        raise TypeError("formula must be a string")

    if not isinstance(skip_confirm, bool):
        raise TypeError("skip_confirm must be a boolean")

    # Validate formula starts with =
    if not formula.startswith("="):
        raise ValueError(f"Formula must start with '=', got '{formula}'")

    # Check file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    try:
        # Load workbook
        wb = load_workbook(filename=file_path)

        # Check sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )

        ws = wb[sheet_name]

        # Add formula
        ws[cell_reference] = formula

        # Save workbook
        wb.save(file_path)
        wb.close()

        return f"Added formula '{formula}' to {cell_reference} in sheet '{sheet_name}'"

    except KeyError:
        raise ValueError(f"Invalid cell reference: {cell_reference}")
    except Exception as e:
        raise ValueError(f"Failed to add formula: {e}")
