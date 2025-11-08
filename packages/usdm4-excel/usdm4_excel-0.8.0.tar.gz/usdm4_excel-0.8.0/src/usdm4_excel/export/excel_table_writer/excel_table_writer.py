from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import List, Any, Union, Tuple, Optional
import os.path


class ExcelTableWriter:
    """
    A class for adding tables (list of lists) to specific positions within an Excel workbook.

    This class allows you to:
    - Add tables to specific sheets in an existing workbook
    - Create a new workbook if the specified path doesn't exist
    - Position each table at a specified row and column
    - Save the modified workbook
    - Get the last row used after adding a table
    """

    def __init__(self, workbook_path: str, default_sheet_name: str = "Sheet1"):
        """
        Initialize the ExcelTableWriter with the path to an Excel workbook.
        If the workbook doesn't exist, a new one will be created.

        Args:
            workbook_path (str): Path to the Excel workbook
            default_sheet_name (str, optional): Name of the default sheet when creating a new workbook.
                                               Defaults to "Sheet1".
        """
        self.workbook_path = workbook_path
        self.default_sheet_name = default_sheet_name
        self._load_workbook()

    def _load_workbook(self):
        """
        Load the Excel workbook or create a new one if it doesn't exist.
        """
        try:
            if os.path.exists(self.workbook_path):
                self.workbook = load_workbook(self.workbook_path)
            else:
                # Create a new workbook if the file doesn't exist
                self.workbook = Workbook()
                # Rename the default sheet
                ws = self.workbook.active
                ws.title = self.default_sheet_name
        except Exception as e:
            raise Exception(f"Error loading or creating workbook: {str(e)}")

    def add_table(
        self,
        data: List[List[Any]],
        sheet_name: str,
        start_row: int = 1,
        start_col: int = 1,
    ) -> int:
        """
        Add a table (list of lists) to a specific position in a worksheet.
        If the specified sheet does not exist, it will be created.

        Args:
            data (List[List[Any]]): The table data as a list of lists
            sheet_name (str): Name of the sheet to add the table to
            start_row (int, optional): Row number to start adding the table (1-based). Defaults to 1.
            start_col (int, optional): Column number to start adding the table (1-based). Defaults to 1.

        Returns:
            int: The last row used in the sheet after adding the table
        """
        # Check if the sheet exists, create it if it doesn't
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)

        # Get the worksheet
        worksheet = self.workbook[sheet_name]

        # Write the data to the worksheet
        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                worksheet.cell(
                    row=start_row + i, column=start_col + j, value=cell_value
                )

        # Calculate and return the last row used
        last_row_used = start_row + len(data) - 1
        return last_row_used

    def save(self, output_path: str = None) -> None:
        """
        Save the workbook to the specified path or overwrite the original file.

        Args:
            output_path (str, optional): Path to save the workbook to.
                                         If None, overwrites the original file. Defaults to None.
        """
        save_path = output_path if output_path else self.workbook_path
        self.workbook.save(save_path)

    def close(self) -> None:
        """Close the workbook."""
        self.workbook.close()

    def format_cells(
        self,
        sheet_name: str,
        cell_range: Union[Tuple[int, int, int, int], str],
        font_size: int = None,
        font_style: str = None,
        vertical_alignment: str = None,
        horizontal_alignment: str = None,
        wrap_text: bool = None,
        background_color: str = None,
    ) -> None:
        """
        Format a range of cells on a specific worksheet.
        If the specified sheet does not exist, it will be created.

        Args:
            sheet_name (str): Name of the sheet containing the cells to format
            cell_range (Union[Tuple[int, int, int, int], str]): Range of cells to format.
                                                               Can be either:
                                                               - A tuple of (start_row, start_col, end_row, end_col) (1-based)
                                                               - A string in Excel notation (e.g., "A1:C5")
            font_size (int, optional): Font size to apply. Defaults to None (no change).
            font_style (str, optional): Font style to apply. Can be 'bold', 'italic', or 'bold italic'.
                                       Defaults to None (no change).
            vertical_alignment (str, optional): Vertical alignment to apply.
                                              Can be 'top', 'center', or 'bottom'.
                                              Defaults to None (no change).
            horizontal_alignment (str, optional): Horizontal alignment to apply.
                                                Can be 'left', 'center', 'right', or 'justify'.
                                                Defaults to None (no change).
            wrap_text (bool, optional): Whether to enable text wrapping.
                                      Defaults to None (no change).
            background_color (str, optional): Background color as a hex string (e.g., "FF0000" for red).
                                            Defaults to None (no change).

        Raises:
            ValueError: If the cell range format is invalid
        """
        # Check if the sheet exists, create it if it doesn't
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)

        # Get the worksheet
        worksheet = self.workbook[sheet_name]

        # Parse the cell range
        if isinstance(cell_range, str):
            # If cell_range is a string (e.g., "A1:C5"), use worksheet's iter_rows method
            # Convert the range to coordinates
            cells = []
            for row in worksheet[cell_range]:
                cells.append(row)
        elif isinstance(cell_range, tuple) and len(cell_range) == 4:
            # If cell_range is a tuple (start_row, start_col, end_row, end_col)
            start_row, start_col, end_row, end_col = cell_range
            cells = []
            for row in worksheet.iter_rows(
                min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col
            ):
                cells.append(row)
        else:
            raise ValueError(
                "Invalid cell range format. Must be either a string (e.g., 'A1:C5') "
                "or a tuple of (start_row, start_col, end_row, end_col)"
            )

        # Prepare font style parameters
        bold = None
        italic = None
        if font_style:
            font_style = font_style.lower()
            bold = "bold" in font_style
            italic = "italic" in font_style

        # Prepare alignment parameters
        v_align = None
        if vertical_alignment:
            vertical_alignment = vertical_alignment.lower()
            if vertical_alignment == "top":
                v_align = "top"
            elif vertical_alignment == "center":
                v_align = "center"
            elif vertical_alignment == "bottom":
                v_align = "bottom"

        h_align = None
        if horizontal_alignment:
            horizontal_alignment = horizontal_alignment.lower()
            if horizontal_alignment == "left":
                h_align = "left"
            elif horizontal_alignment == "center":
                h_align = "center"
            elif horizontal_alignment == "right":
                h_align = "right"
            elif horizontal_alignment == "justify":
                h_align = "justify"

        # Apply formatting to each cell in the range
        for row in cells:
            for cell in row:
                # Apply font formatting if specified
                if font_size is not None or bold is not None or italic is not None:
                    # Get current font properties
                    current_font = cell.font

                    # Create a new font based on the current font, updating only specified properties
                    new_font = Font(
                        name=current_font.name,
                        size=font_size if font_size is not None else current_font.size,
                        bold=bold if bold is not None else current_font.bold,
                        italic=italic if italic is not None else current_font.italic,
                        color=current_font.color,
                        underline=current_font.underline,
                        strike=current_font.strike,
                        vertAlign=current_font.vertAlign,
                        charset=current_font.charset,
                        scheme=current_font.scheme,
                        family=current_font.family,
                        outline=current_font.outline,
                        shadow=current_font.shadow,
                        condense=current_font.condense,
                        extend=current_font.extend,
                    )
                    cell.font = new_font

                # Apply alignment formatting if specified
                if h_align is not None or v_align is not None or wrap_text is not None:
                    # Get current alignment properties
                    current_alignment = cell.alignment

                    # Create a new alignment based on the current alignment, updating only specified properties
                    new_alignment = Alignment(
                        horizontal=h_align
                        if h_align is not None
                        else current_alignment.horizontal,
                        vertical=v_align
                        if v_align is not None
                        else current_alignment.vertical,
                        textRotation=current_alignment.textRotation,
                        wrapText=wrap_text
                        if wrap_text is not None
                        else current_alignment.wrapText,
                        shrinkToFit=current_alignment.shrinkToFit,
                        indent=current_alignment.indent,
                        relativeIndent=current_alignment.relativeIndent,
                        justifyLastLine=current_alignment.justifyLastLine,
                        readingOrder=current_alignment.readingOrder,
                    )
                    cell.alignment = new_alignment

                # Apply background color if specified
                if background_color is not None:
                    cell.fill = PatternFill(
                        start_color=background_color,
                        end_color=background_color,
                        fill_type="solid",
                    )

    def set_column_width(
        self,
        sheet_name: str,
        columns: Union[int, str, List[Union[int, str]]],
        width: float,
    ) -> None:
        """
        Set the width for one or more columns in a worksheet.
        If the specified sheet does not exist, it will be created.

        Args:
            sheet_name (str): Name of the sheet containing the columns to modify
            columns (Union[int, str, List[Union[int, str]]]): Column(s) to modify.
                                                             Can be either:
                                                             - A single column index (1-based, like Excel)
                                                             - A single column letter (e.g., "A", "B", etc.)
                                                             - A list of column indices or letters
            width (float): The width to set for the column(s)

        Raises:
            ValueError: If the column format is invalid
        """
        # Check if the sheet exists, create it if it doesn't
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)

        # Get the worksheet
        worksheet = self.workbook[sheet_name]

        # Convert columns to a list if it's a single value
        if not isinstance(columns, list):
            columns = [columns]

        # Process each column
        for col in columns:
            # Convert column to index if it's a letter
            if isinstance(col, str):
                try:
                    col_idx = column_index_from_string(col)
                except ValueError:
                    raise ValueError(f"Invalid column letter: '{col}'")
            elif isinstance(col, int):
                if col < 1:
                    raise ValueError(f"Column index must be positive: {col}")
                col_idx = col
            else:
                raise ValueError(
                    f"Invalid column format: {col}. Must be an integer or a string."
                )

            # Get the column letter (required by openpyxl for setting width)
            col_letter = get_column_letter(col_idx)

            # Set the column width
            worksheet.column_dimensions[col_letter].width = width

    def set_sheet_font(
        self,
        sheet_name: str,
        font_name: Optional[str] = None,
        font_size: Optional[int] = None,
        bold: Optional[bool] = None,
        italic: Optional[bool] = None,
        color: Optional[str] = None,
    ) -> None:
        """
        Set the font for all cells in a worksheet.
        If the specified sheet does not exist, it will be created.

        Args:
            sheet_name (str): Name of the sheet to modify
            font_name (Optional[str], optional): Font name (e.g., "Arial", "Calibri"). Defaults to None (no change).
            font_size (Optional[int], optional): Font size in points. Defaults to None (no change).
            bold (Optional[bool], optional): Whether the text should be bold. Defaults to None (no change).
            italic (Optional[bool], optional): Whether the text should be italic. Defaults to None (no change).
            color (Optional[str], optional): Font color as a hex string (e.g., "FF0000" for red). Defaults to None (no change).
        """
        # Check if the sheet exists, create it if it doesn't
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)

        # Get the worksheet
        worksheet = self.workbook[sheet_name]

        # Check if any font properties are specified
        if all(prop is None for prop in [font_name, font_size, bold, italic, color]):
            return  # No changes to make

        # Iterate through all cells in the worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                # Get current font properties
                current_font = cell.font

                # Create a new font based on the current font, updating only specified properties
                new_font = Font(
                    name=font_name if font_name is not None else current_font.name,
                    size=font_size if font_size is not None else current_font.size,
                    bold=bold if bold is not None else current_font.bold,
                    italic=italic if italic is not None else current_font.italic,
                    color=color if color is not None else current_font.color,
                    underline=current_font.underline,
                    strike=current_font.strike,
                    vertAlign=current_font.vertAlign,
                    charset=current_font.charset,
                    scheme=current_font.scheme,
                    family=current_font.family,
                    outline=current_font.outline,
                    shadow=current_font.shadow,
                    condense=current_font.condense,
                    extend=current_font.extend,
                )
                cell.font = new_font
