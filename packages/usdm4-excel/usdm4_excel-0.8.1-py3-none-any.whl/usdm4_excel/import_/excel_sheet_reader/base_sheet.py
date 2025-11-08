import os
import datetime
import pandas as pd
from typing import Type
from openpyxl import load_workbook
from simple_error_log.errors import Errors
from simple_error_log.error_location import KlassMethodLocation
from usdm4_excel.import_.excel_sheet_reader.sheet_location import SheetLocation
from usdm4_excel.import_.excel_sheet_reader.split_complex import (
    SplitComplex,
    SplitStateError,
    SplitFormatError,
)
from usdm4.builder.builder import Builder
from usdm4.api.api_base_model import ApiBaseModelWithId
from usdm4.api.quantity_range import Quantity
from usdm4.api.comment_annotation import CommentAnnotation
from usdm4.api.code import Code
from usdm4_excel.import_.types.quantity_type import QuantityType
from usdm4_excel.import_.types.range_type import RangeType
from usdm4.api.quantity_range import Range
from usdm4.api.address import Address
from usdm4.api.syntax_template_dictionary import SyntaxTemplateDictionary
from usdm4.api.geographic_scope import GeographicScope


class BaseSheet:
    MODULE = "usdm4_excel.import_.excel_sheet_reader.base_sheet.BaseSheet"

    class FormatError(Exception):
        pass

    class StateError(Exception):
        pass

    def __init__(
        self,
        file_path: str,
        sheet_name: str,
        builder: Builder,
        errors: Errors,
        header: int = 0,
        optional: bool = False,
        converters: dict = {},
    ):
        self._errors = errors
        self._builder = builder
        self._file_path = file_path
        self._dir_path, self._filename = os.path.split(file_path)
        self._sheet_name = sheet_name
        self._sheet = None
        self._success = False
        self._sheet_names = None
        if optional and not self._sheet_present(file_path, sheet_name):
            self._errors.warning(
                f"{sheet_name} not found but optional",
                KlassMethodLocation(self.MODULE, "__init__"),
            )
        else:
            self._sheet = pd.read_excel(
                open(file_path, "rb"),
                sheet_name=sheet_name,
                header=header,
                converters=converters,
            )
            self._success = True

    def _empty(self, row_index: int, col_index: int) -> bool:
        return pd.isnull(self._sheet.iloc[row_index, col_index])

    def _column_present(self, names):
        fields = [names] if isinstance(names, str) else names
        for field in fields:
            try:
                col_index = self._sheet.columns.get_loc(field)
                return col_index
            except Exception:
                pass
        columns = ", ".join(fields)
        raise BaseSheet.FormatError(f"Failed to detect column(s) '{columns}' in sheet")

    def _create(
        self,
        klass: Type[ApiBaseModelWithId],
        params: dict,
        cross_reference: bool = True,
    ) -> object:
        try:
            item = self._builder.create(klass, params, cross_reference)
            if not item:
                self._errors.error(f"Failed to create '{klass.__name__}' object")
            return item
        except Exception as e:
            self._errors.exception(
                f"Exception raised creating '{klass.__name__}' object",
                e,
                KlassMethodLocation(self.MODULE, "_create"),
            )
            return None

    def _read_cell_by_name(
        self, row_index, field_name, default=None, must_be_present=True
    ):
        try:
            col_index = self._column_present(field_name)
            return self._read_cell(row_index, col_index)
        except Exception as e:
            if not must_be_present:
                return ""
            elif default:
                return default
            else:
                self._errors.exception(
                    f"Exception raised attempting to read cell '{field_name}'",
                    e,
                    self._location(row_index, None),
                )
                return ""

    def _read_boolean_cell_by_name(self, row_index, field_name, default=False):
        """Read a boolean cell value by field name"""
        try:
            value = self._read_cell_by_name(row_index, field_name)
            if not value or value.strip() == "":
                return default
            value_lower = value.strip().lower()
            if value_lower in ["true", "yes", "1", "y", "t"]:
                return True
            elif value_lower in ["false", "no", "0", "n", "f"]:
                return False
            else:
                self._errors.warning(
                    f"Unrecognized boolean value '{value}', defaulting to {default}",
                    self._location(row_index, self._column_present(field_name)),
                )
                return default
        except Exception as e:
            self._errors.exception(
                f"Exception raised reading boolean cell '{field_name}'",
                e,
                self._location(row_index, None),
            )
            return default

    def _read_cell(self, row_index, col_index, default=None):
        try:
            if pd.isnull(self._sheet.iloc[row_index, col_index]):
                return default if default else ""
            else:
                return str(self._sheet.iloc[row_index, col_index]).strip()
        except Exception as e:
            print(f"READ CELL: {self._sheet_name}, [{row_index}, {col_index}], {e}")
            self._errors.exception(
                "Exception raised reading cell", e, self._location(row_index, col_index)
            )
            if default:
                return default
            else:
                return ""

    def _read_cell_multiple_by_name(self, row_index, field_name, must_be_present=True):
        try:
            col_index = self._column_present(field_name)
            return self._read_cell_multiple(row_index, col_index)
        except Exception as e:
            if not must_be_present:
                return []
            else:
                self._errors.exception(
                    f"Exception raised reading cell multiple '{field_name}'",
                    e,
                    self._location(row_index, None),
                )
                return []

    def _read_cell_multiple(self, rindex, cindex):
        try:
            results = []
            value = self._read_cell(rindex, cindex)
            if value.strip() == "":
                return results
            splitter = SplitComplex(value)
            complex_parts = splitter.split()
            for part in complex_parts:
                results.append(part.strip())
            return results
        except SplitStateError as e:
            self._errors.exception(
                "Exception raised due to state error reading cell multiple",
                e,
                self._location(rindex, cindex),
            )
            return []
        except SplitFormatError as e:
            self._errors.exception(
                "Exception raised due to format error reading cell multiple",
                e,
                self._location(rindex, cindex),
            )
            return []

    def _read_quantity_cell_by_name(
        self, row_index, field_name, allow_missing_units=True, allow_empty=True
    ):
        col_index = self._column_present(field_name)
        return self._read_quantity_cell(
            row_index, col_index, allow_missing_units, allow_empty
        )

    def _read_quantity_cell(
        self, row_index, col_index, allow_missing_units=True, allow_empty=True
    ):
        text = ""
        try:
            text = self._read_cell(row_index, col_index)
            quantity = QuantityType(
                text, self._builder, self._errors, allow_missing_units, allow_empty
            )
            if quantity.valid:
                unit = self._builder.alias_code(quantity.units_code)
                return (
                    None
                    if quantity._empty
                    else self._builder.create(
                        Quantity, {"value": float(quantity.value), "unit": unit}
                    )
                )
            else:
                return None
        except Exception as e:
            self._errors.exception(
                f"Exception raised decoding quantity data '{text}'",
                e,
                self._location(row_index, col_index),
            )
            return None

    def _read_range_cell_by_name(
        self, row_index, field_name, require_units=True, allow_empty=True
    ):
        col_index = self._column_present(field_name)
        return self._read_range_cell(row_index, col_index, require_units, allow_empty)

    def _read_range_cell(
        self, row_index, col_index, require_units=True, allow_empty=True
    ):
        text = ""
        try:
            text = self._read_cell(row_index, col_index)
            range_type = RangeType(
                text, self._builder, self._errors, require_units, allow_empty
            )
            if range_type._empty:
                return None
            elif range_type.valid:
                params = {
                    "minValue": self._builder.create(
                        Quantity,
                        {
                            "value": float(range_type.lower),
                            "unit": range_type.units_code,
                        },
                    ),
                    "maxValue": self._builder.create(
                        Quantity,
                        {
                            "value": float(range_type.upper),
                            "unit": range_type.units_code,
                        },
                    ),
                    "isApproximate": False,
                }
                return self._create(Range, params)
            else:
                return None
        except Exception as e:
            self._errors.exception(
                f"Exception raised decoding range data '{text}'",
                e,
                self._location(row_index, col_index),
            )
            return None

    def _read_cdisc_klass_attribute_cell_by_name(
        self,
        klass: str,
        attribute: str,
        row_index: int,
        field_name: str,
        allow_empty: bool = False,
    ) -> Code:
        col_index = self._column_present(field_name)
        return self._read_cdisc_klass_attribute_cell(
            klass, attribute, row_index, col_index, allow_empty
        )

    def _read_cdisc_klass_attribute_cell(
        self,
        klass: str,
        attribute: str,
        row_index: int,
        col_index: int,
        allow_empty: bool = False,
    ) -> Code:
        code = None
        value = self._read_cell(row_index, col_index)
        if value:
            code = self._builder.klass_and_attribute_value(klass, attribute, value)
            if not code:
                self._errors.error(
                    f"CDISC CT not found for value '{value}'",
                    self._location(row_index, col_index),
                )
        elif not allow_empty:
            self._errors.error(
                "Empty cell detected where CDISC CT value expected",
                self._location(row_index, col_index),
            )
        return code

    def _read_cdisc_klass_attribute_cell_multiple_by_name(
        self, klass, attribute, row_index, field_name
    ):
        col_index = self._column_present(field_name)
        return self._read_cdisc_klass_attribute_cell_multiple(
            klass, attribute, row_index, col_index
        )

    def _read_cdisc_klass_attribute_cell_multiple(
        self, klass, attribute, row_index, col_index
    ):
        result = []
        value = self._read_cell(row_index, col_index)
        if value.strip() == "":
            self._errors.error(
                "Empty cell detected where multiple CDISC CT values expected",
                self._location(row_index, col_index),
            )
            return result
        for item in self._state_split(value):
            code = self._builder.klass_and_attribute_value(
                klass, attribute, item.strip()
            )
            if code is not None:
                result.append(code)
            else:
                self._errors.error(
                    f"CDISC CT not found for value '{item.strip()}'",
                    self._location(row_index, col_index),
                )
        return result

    def _read_iso639_code_cell_by_name(self, row_index, field_name):
        col_index = self._column_present(field_name)
        return self._read_iso639_code_cell(row_index, col_index)

    def _read_iso639_code_cell(self, row_index, col_index):
        value = self._read_cell(row_index, col_index)
        if value.strip() == "":
            return None
        return self._builder.iso639_code_or_decode(value)

    def _read_other_code_cell_by_name(self, row_index, field_name):
        col_index = self._column_present(field_name)
        return self._read_other_code_cell(row_index, col_index)

    def _read_other_code_cell(self, row_index, col_index):
        value = self._read_cell(row_index, col_index)
        if value.strip() == "":
            return None
        return self._decode_other_code(value, row_index, col_index)

    def _read_other_code_cell_multiple_by_name(self, row_index, field_name):
        col_index = self._column_present(field_name)
        return self._read_other_code_cell_multiple(row_index, col_index)

    def _read_other_code_cell_multiple(self, row_index, col_index):
        value = self._read_cell(row_index, col_index)
        result = []
        if value.strip() == "":
            return result
        for item in self._state_split(value):
            code = self._decode_other_code(item.strip(), row_index, col_index)
            if code:
                result.append(code)
        return result

    def _read_date_cell_by_name(self, row_index, field_name, must_be_present=True):
        col_index = self._column_present(field_name)
        return self._read_date_cell(row_index, col_index, must_be_present)

    def _read_date_cell(self, row_index, col_index, must_be_present=True):
        cell = self._read_cell(row_index, col_index)
        try:
            return datetime.datetime.strptime(cell, "%Y-%m-%d %H:%M:%S")
        except Exception as e:
            self._errors.exception(
                "Exception raised reading date cell",
                e,
                self._location(row_index, col_index),
            )
            return None

    def _read_geographic_scopes_cell_by_name(self, row_index, field_name):
        col_index = self._column_present(field_name)
        return self._read_geographic_scopes_cell(row_index, col_index)

    def _read_geographic_scopes_cell(self, row_index, col_index):
        try:
            result = []
            value: str = self._read_cell(row_index, col_index, default="")
            if value.strip() == "":
                self._errors.warning(
                    "Empty cell detected where geographic scope values expected, assuming global scope.",
                    self._location(row_index, col_index),
                )
                result.append(self._scope("Global", None))
            else:
                for item in self._state_split(value):
                    key_value = self._key_value(
                        item, row_index, col_index, allow_single=True
                    )
                    if key_value[0] == "GLOBAL":
                        result.append(self._scope("Global", None))
                    elif key_value[0] == "REGION":
                        code = self._country_region(key_value[1], "Region")
                        if code:
                            scope = self._scope("Region", code)
                            result.append(scope)
                    elif key_value[0] == "COUNTRY":
                        code = self._country_region(key_value[1], "Country")
                        if code:
                            scope = self._scope("Country", code)
                            result.append(scope)
                    else:
                        self._errors.warning(
                            f"Failed to decode geographic scope '{value}'. Formats are 'Global', 'Region: <value>' or 'Country: <value>'. Assuming global scope.",
                            self._location(row_index, col_index),
                        )
                        result.append(self._scope("Global", None))
            return result
        except Exception as e:
            self._errors.exception(
                "No geographic scope column found, assuming global scope.",
                e,
                self._location(row_index, col_index),
            )
            return [self._scope("Global", None)]

    def _add_notes(self, instance: object, note_refs: list) -> None:
        for note_ref in note_refs:
            try:
                note = self._builder.cross_reference.get_by_name(
                    CommentAnnotation, note_ref
                )
                if note:
                    instance.notes.append(note)
                else:
                    self._errors.error(
                        f"Failed to find note with name '{note_ref}'", self._location()
                    )
            except Exception as e:
                self._errors.exception(
                    f"Exception raised adding note to '{object.__class__.__name__}' object",
                    e,
                )

    def _double_link(self, items, prev, next):
        try:
            for idx, item in enumerate(items):
                if idx == 0:
                    setattr(item, prev, None)
                else:
                    the_id = getattr(items[idx - 1], "id")
                    setattr(item, prev, the_id)
                if idx == len(items) - 1:
                    setattr(item, next, None)
                else:
                    the_id = getattr(items[idx + 1], "id")
                    setattr(item, next, the_id)
        except Exception as e:
            self._errors.exception(
                "Exception raised while doubly linking lists",
                e,
                KlassMethodLocation(self.MODULE, "_double_link"),
            )

    def _single_link(self, items, next):
        try:
            for idx, item in enumerate(items):
                if idx == len(items) - 1:
                    setattr(item, next, None)
                else:
                    the_id = getattr(items[idx + 1], "id")
                    setattr(item, next, the_id)
        except Exception as e:
            self._errors.exception(
                "Exception raised while singly linking lists",
                e,
                KlassMethodLocation(self.MODULE, "_singly_link"),
            )

    def _sheet_present(self, file_path, sheet_name):
        sheet_names = self._get_sheet_names(file_path)
        return sheet_name in sheet_names

    def _get_sheet_names(self, file_path):
        if not self._sheet_names:
            wb = load_workbook(file_path, read_only=True, keep_links=False)
            self._sheet_names = wb.sheetnames
        return self._sheet_names

    def _sheet_exception(self, e):
        self._errors.exception(
            f"Exception raised while reading sheet '{self._sheet_name}'",
            e,
            KlassMethodLocation(self.MODULE, "_sheet_exception"),
        )

    def _decode_other_code(self, value, row_index, col_index):
        if value.strip() == "":
            return None
        outer_parts = value.split(":")
        if len(outer_parts) == 2:
            system = outer_parts[0].strip()
            inner_parts = outer_parts[1].strip().split("=")
            if len(inner_parts) == 2:
                version = self._builder.other_ct_version_manager.get(system)
                return self._builder.other_code(
                    code=inner_parts[0].strip(),
                    system=system,
                    version=version,
                    decode=inner_parts[1].strip(),
                )
            else:
                self._errors.error(
                    "Failed to decode code data '%s', no '=' detected" % (value),
                    self._location(row_index, col_index),
                )
        else:
            self._errors.error(
                "Failed to decode code data '%s', no ':' detected" % (value),
                self._location(row_index, col_index),
            )
        return None

    def _country_region_quantity(
        self, text: str, type: str, row_index: int, col_index: int
    ):
        name_value = text.split("=")
        if len(name_value) == 2:
            quantity = self._get_quantity(name_value[1].strip())
            code = self._country_region(name_value[0].strip(), type)
            return code, quantity
        else:
            self._error(
                f"Failed to decode geographic enrollment data '{text}', incorrect format, missing '='?",
                self._location(row_index, col_index),
            )
            return None, None

    def _country_region(self, text: str, type: str):
        return (
            self._builder.iso3166_region_code(text)
            if type == "Region"
            else self._builder.iso3166_code_or_decode(text)
        )

    def _read_address_cell_by_name(self, row_index, field_name, allow_empty=False):
        raw_address = self._read_cell_by_name(row_index, field_name)
        # TODO The '|' separator is preserved for legacy reasons but should be removed in the future
        if not raw_address:
            sep = ","
            parts = []
        elif "|" in raw_address:
            sep = "|"
            parts = raw_address.split(sep)
        else:
            sep = ","
            parts = self._state_split(raw_address)
        if len(parts) >= 6:
            result = self._to_address(
                lines=[x.strip() for x in parts[:-5]],
                district=parts[-5].strip(),
                city=parts[-4].strip(),
                state=parts[-3].strip(),
                postal_code=parts[-2].strip(),
                country_code=parts[-1].strip(),
            )
            return result
        elif allow_empty:
            return None
        else:
            col_index = self._column_present(field_name)
            self._errors.error(
                f"Address '{raw_address}' does not contain the required fields (lines, district, city, state, postal code and country code) using '{sep}' separator characters, only {len(parts)} found",
                self._location(row_index, col_index),
            )
            return None

    def _to_address(self, lines, city, district, state, postal_code, country_code):
        text = f"{(', ').join(lines)}, {city}, {district}, {state}, {postal_code}, {country_code}"
        country = self._builder.iso3166_code(country_code) if country_code else None
        result = self._builder.create(
            Address,
            {
                "text": text,
                "lines": lines,
                "city": city,
                "district": district,
                "state": state,
                "postalCode": postal_code,
                "country": country,
            },
        )
        return result

    def _state_split(self, s):
        OUT = "out"
        IN_QUOTED = "in_quoted"
        OUT_QUOTED = "out_quoted"
        IN_NORMAL = "in_normal"
        ESC = "escape"

        state = OUT
        result = []
        current = []
        exit = ""
        for c in s:
            # print(f"STATE: s: {state}, c: {c}")
            if state == OUT:
                current = []
                if c == ",":
                    result.append("")
                elif c in ['"', "'"]:
                    state = IN_QUOTED
                    exit = c
                elif c.isspace():
                    pass
                else:
                    state = IN_NORMAL
                    current.append(c)
            elif state == IN_QUOTED:
                if c == "\\":
                    state = ESC
                elif c == exit:
                    result.append("".join(current).strip())
                    state = OUT_QUOTED
                else:
                    current.append(c)
            elif state == OUT_QUOTED:
                if c == ",":
                    state = OUT
                else:
                    pass
            elif state == IN_NORMAL:
                if c == ",":
                    result.append("".join(current).strip())
                    state = OUT
                else:
                    current.append(c)
            elif state == ESC:
                if c == exit:
                    current.append(c)
                    state = IN_QUOTED
                else:
                    current.append("\\")
                    current.append(c)
            else:
                raise BaseSheet.StateError

        if state == OUT or state == OUT_QUOTED:
            pass
        elif state == IN_NORMAL:
            result.append("".join(current).strip())
        else:
            raise BaseSheet.FormatError
        return result

    def _location(self, row: int = None, column: int = None) -> SheetLocation:
        return SheetLocation(self._sheet_name, row, column)

    def _get_dictionary_id(self, dictionary_name: str) -> str | None:
        if dictionary_name:
            instance: ApiBaseModelWithId = self._get_cross_reference(
                SyntaxTemplateDictionary, dictionary_name
            )
            return instance.id if instance else None
        else:
            return None

    def _get_cross_reference(
        self, klasses: list[Type | str] | Type | str, name: str
    ) -> Type | None:
        klasses = klasses if isinstance(klasses, list) else [klasses]
        try:
            for klass in klasses:
                klass_name: str = klass if isinstance(klass, str) else klass.__name__
                instance = self._builder.cross_reference.get_by_name(klass, name)
                if instance:
                    return instance
            self._errors.error(
                f"Failed to find an instance of '{klass_name}' with name '{name}'",
                self._location(),
            )
            return None
        except Exception as e:
            self._errors.exception(
                f"Exception raised finding cross reference of '{klass_name}' with name '{name}'",
                e,
                self._location(),
            )
            return None

    def _key_value(self, text: str, row_index: int, col_index: int, allow_single=False):
        if text.strip():
            parts = text.split(":")
            if len(parts) == 2:
                return [parts[0].strip().upper(), parts[1].strip()]
            elif len(parts) == 1 and allow_single:
                return [parts[0].strip().upper(), ""]
        self._errors.error(
            f"Failed to decode key value pair '{text}', incorrect format, missing ':'?",
            self._location(row_index, col_index),
        )
        return ["", ""]

    def _get_quantity(self, text):
        quantity = QuantityType(text, self._builder, self._errors, True, False)
        unit = (
            self._builder.alias_code(quantity.units_code)
            if quantity.units_code
            else None
        )
        return self._create(Quantity, {"value": float(quantity.value), "unit": unit})

    def _scope(self, type, code):
        scope_type = self._builder.klass_and_attribute_value(
            "GeographicScope", "type", type
        )
        alias = self._builder.alias_code(code) if code else None
        return self._create(GeographicScope, {"type": scope_type, "code": alias})
