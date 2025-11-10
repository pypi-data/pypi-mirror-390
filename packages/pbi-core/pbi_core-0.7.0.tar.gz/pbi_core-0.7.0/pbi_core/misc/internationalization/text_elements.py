import json
from typing import TYPE_CHECKING, Literal

import openpyxl
from attrs import define

from pbi_core.logging import get_logger
from pbi_core.ssas.model_tables import Column, Hierarchy, Measure, Table
from pbi_core.static_files.layout.sources.literal import LiteralSource, serialize_literal

if TYPE_CHECKING:
    from _typeshed import StrPath

    from pbi_core.report.base.main import BaseReport
logger = get_logger()


@define()
class TextElement:
    category: str
    """Used to group elements into different Excel sheets"""
    source: Literal["layout", "ssas"]
    """Used to identify whether the xpath should apply to the frontend or backend"""
    xpath: list[str | int]
    """Used to identify the entity to update"""
    field: str
    """Used to identify the field of the updating entity"""
    text: str


class TextElements:
    language: str = "source_text"
    text_elements: list[TextElement]

    def __init__(self, text_elements: list[TextElement] | None = None, language: str = "source_text") -> None:
        self.text_elements = text_elements or []
        self.language = language

    def to_csv(self) -> None:
        pass

    def _grouped(self) -> dict[str, list[TextElement]]:
        grouped: dict[str, list[TextElement]] = {}
        for element in self.text_elements:
            grouped.setdefault(element.category, []).append(element)
        return grouped

    def set_elements(self, report: "BaseReport") -> None:
        """Updates the text elements."""
        for text_element in self.text_elements:
            if text_element.source == "layout":
                node = report.static_files.layout.find_xpath(text_element.xpath)
                if isinstance(node, LiteralSource):
                    node.Literal.Value = serialize_literal(text_element.text)
                else:
                    setattr(node, text_element.field, text_element.text)
            elif text_element.source == "ssas":
                group, idx = text_element.xpath
                assert isinstance(group, str)
                assert isinstance(idx, int)
                group = getattr(report.ssas, group)
                entity = group.find(idx)
                # These entities have DAX implications that need to be handled specially
                if (isinstance(entity, (Measure, Hierarchy, Table)) and text_element.field == "name") or (
                    isinstance(entity, Column) and text_element.field == "explicit_name"
                ):
                    entity.set_name(text_element.text, report.static_files.layout)
                else:
                    setattr(entity, text_element.field, text_element.text)
        if any(te.source == "ssas" for te in self.text_elements):
            report.ssas.sync_to()

    def to_excel(self, path: "StrPath") -> None:
        wb = openpyxl.Workbook()
        for category, objects in self._grouped().items():
            ws = wb.create_sheet(category)
            for j, name in enumerate(["source", "xpath", "field", "source_text"]):
                ws.cell(1, j + 1).value = name
            for i, obj in enumerate(objects):
                ws.cell(2 + i, 1).value = obj.source
                ws.cell(2 + i, 2).value = json.dumps(obj.xpath)
                ws.cell(2 + i, 3).value = obj.field
                ws.cell(2 + i, 4).value = obj.text
        wb.remove(wb["Sheet"])
        wb.save(path)

    @classmethod
    def from_excel(cls, path: "StrPath") -> "list[TextElements]":
        logger.info("Parsing Excel", path=path)
        wb = openpyxl.load_workbook(path)

        headers = next(wb.worksheets[0].rows)
        languages: list[str] = [str(c.value) for c in headers][4:]

        static_elements: dict[str, list[TextElement]] = {}
        for ws in wb.worksheets:
            for row in list(ws.values)[1:]:
                assert isinstance(row[0], str)
                assert row[0] in ("layout", "ssas")  # noqa: PLR6201
                # For now, mypy is not able to infer the type of row[0] when using set notation

                assert isinstance(row[1], str)
                assert isinstance(row[2], str)
                assert isinstance(row[3], str)
                xpath = json.loads(row[1])
                field = row[2]
                for i, lang in enumerate(languages):
                    text = row[4 + i]
                    assert isinstance(text, str)
                    source = row[0]
                    assert source in {"layout", "ssas"}
                    static_elements.setdefault(lang, []).append(
                        TextElement(
                            category=ws.title,
                            source=source,
                            xpath=xpath,
                            field=field,
                            text=text,
                        ),
                    )
        return [
            TextElements(
                text_elements=elements,
                language=lang,
            )
            for lang, elements in static_elements.items()
        ]
