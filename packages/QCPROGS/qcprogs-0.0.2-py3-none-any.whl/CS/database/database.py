
import oracledb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import ALL_TABEL_MAPPING
from model import DataMapItem,Dict
from icecream import ic

class OracleDB:
    DATAMAP:list[DataMapItem] = ALL_TABEL_MAPPING
    VALUE: Dict[str, str] = {}
    DUMMY =["CS_CODE","VENDER_ID","SERVICE_ID"]
    CONFIG={}
    VENDOR_CODE =''
    VENDOR_ID =''
    VENDOR_NAME =''
    SERVICE_ID =''
    USER = 'CS_SUPPORT'
    PASSWORD ='1234'
    HOST='csonldbqa01.counterservice.co.th'
    PORT='1521'
    DATABASE ='ONLPRD'
    def __init__(self,URL=None, **kwargs):
        self._set_url(URL)
    def _set_url(self,URL):
        if not URL:
            self.URL = f'{self.HOST}:{self.PORT}/{self.DATABASE}'
        else:
            self.URL=f'{URL}:{self.PORT}/{self.DATABASE}'
        self.__connect()

    def __connect(self):
        try:
            self.conn = oracledb.connect(
                user=self.USER,
                password=self.PASSWORD,
                dsn=self.URL
            )
            ic("Connected to Oracle Database")
        except Exception as e:
            ic("Connection failed:", e)
    def _get_vendor(self,CSCODE):
        if CSCODE == '':
            return
        cur = self.conn.cursor()
        QUY =f"SELECT VENDOR_CODE ,VENDOR_NAME ,VENDOR_ID,SERVICE_ID FROM ONLSTD.WS_CLIENT_CONFIG WHERE 1=1 {CSCODE} FETCH FIRST 1 ROWS ONLY"
        ic(QUY)
        cur.execute(QUY)
        cols = [c[0] for c in cur.description] # type: ignore
        data = cur.fetchone()
        if data:
            return  dict(zip(cols, data))
        
    def __set_config(self, cs_code: str, vender_code: str, service_id: str,flag=True):
        TEXT : str  = ''
        if cs_code == '':
            TEXT += f"AND VENDOR_ID = '{vender_code}' "
        if vender_code == '':
            TEXT += f"AND VENDOR_CODE = '{cs_code}' "
        if not service_id == '':
            TEXT += f"AND SERVICE_ID = '{service_id}' "
        ic(TEXT)
        item :dict= self._get_vendor(TEXT) # type: ignore
        ic(item)
        for value in self.DATAMAP:
            TEXT=''
            for k, v in value['RULE_MAP'].items():            
                if v.get('STATUS', 'S') == 'A':
                    TEXT += f"AND {v['RESULT']} = '{item.get(v['RESULT'])}' "
            COL = ','.join(value.get('COLUMN', '*')if not flag else '*')
            value['QUERY'] = (
                f"""SELECT {COL} FROM {value['SCHEMA']}.{value['TABEL']} WHERE 1=1 {TEXT} FETCH FIRST 50 ROWS ONLY""")
            self.CONFIG[value['TABEL']] = value['QUERY']
        ic(value['QUERY'])
        return self.CONFIG
    def _run_query(self, query, fetch=True,flag=True):
        """รัน query (SELECT / UPDATE)"""
        try:
            ic(query)
            with self.conn.cursor() as cur:
                ic("""รัน query (SELECT / UPDATE)""")
                cur.execute(query)
                if fetch:
                    cols = [c[0] for c in cur.description] # pyright: ignore[reportOptionalIterable]
                    rows = cur.fetchall()
                    # แปลงแต่ละค่าเป็น str เพื่อรองรับ Unicode
                    rows_clean = [[str(v).replace("\n", " ").replace("\r", " ").strip().encode('latin1').decode('tis-620',errors='ignore') if v is not None else "" for v in r] for r in rows]
                    ic(rows_clean)
                    if flag:
                        return pd.DataFrame(rows_clean, columns=cols)
                    else:
                        return {"Column":cols,"Row":rows_clean}
                else:
                    ic(f"Executed: {query[:80]}...")
        except Exception as e:
            ic(f"Error: {e}")
        return None
    def export_file(self, VENDOR_CODE: str,CS_CODE :str|None=None , SERVICE: str|None = None):
        wb = Workbook()
        ws = wb.active
        ws.title = "config" # pyright: ignore[reportOptionalMemberAccess]
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        column_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        row = 1
        all_results:dict = self.fetch_tables(VENDOR_CODE,str(CS_CODE),str(SERVICE)) # pyright: ignore[reportAssignmentType]
        for table_name, df in all_results.items():
            ws.cell(row=row, column=1, value=table_name) # pyright: ignore[reportOptionalMemberAccess]
            ws.cell(row=row, column=1).fill = header_fill # pyright: ignore[reportOptionalMemberAccess]
            ws.cell(row=row, column=1).font = Font(bold=True) # pyright: ignore[reportOptionalMemberAccess]
            row += 2
            if df is not None and not df.empty:
                for col_num, col_name in enumerate(df.columns, 1):
                    cell = ws.cell(row=row, column=col_num, value=col_name) # pyright: ignore[reportOptionalMemberAccess]
                    cell.fill = column_fill
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
                for r in df.itertuples(index=False):
                    row += 1
                    for col_num, val in enumerate(r, 1):
                        cell = ws.cell(row=row, column=col_num, value=val) # pyright: ignore[reportOptionalMemberAccess]
                        cell.border = border
                row += 3
            else:
                ws.cell(row=row, column=1, value="ไม่พบการ Set ที่ Tabel นี้") # pyright: ignore[reportOptionalMemberAccess]
                row += 3

        # Auto adjust column width
        for col in ws.columns: # pyright: ignore[reportOptionalMemberAccess]
            max_length = 0
            col_letter = col[0].column_letter # pyright: ignore[reportAttributeAccessIssue]
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2 # pyright: ignore[reportOptionalMemberAccess]
        wb.save(f"QCM-COU-Test Case&Result_CS Online_{self.VENDOR_CODE}_SV{str(self.SERVICE_ID).zfill(2)}_{self.VENDOR_NAME}-R1.xlsx")
        return True
    def fetch_tables(self,values: str,cs_code :str|None=None , service: str|None = None,include=None):
        """
        ดึงข้อมูลจากหลาย table + แสดงผลใน Jupyter
        export_excel: ถ้าใส่ชื่อไฟล์ เช่น "output.xlsx" → export เป็น Excel ด้วย
        """
        FLAG = False
        ic(values,cs_code,service,include)
        if  not include:
            FLAG = True
            include = self.DATAMAP
        ic(values,cs_code,service,include)  
        self.__set_config(str(cs_code),values,str(service))
        ic(self.CONFIG)
        all_results = {}
        for ky,vl in self.CONFIG.items():
            if ky not in include :
                continue
            try:
                all_results[ky] = self._run_query(vl,flag=FLAG)
                ic(all_results[ky])
                if not FLAG:
                    return all_results[ky]
            except Exception :
                continue
        return all_results

        