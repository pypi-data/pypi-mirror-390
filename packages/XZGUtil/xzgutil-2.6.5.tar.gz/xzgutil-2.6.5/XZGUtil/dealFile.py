#! /usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time    : 2022-06-16 10:07
# @Site    :
# @File    : dealFile.py
# @Software: PyCharm
"""
读取，写入表格文件
"""
import csv
import os
import traceback

import xlrd
import openpyxl as op
from openpyxl.utils.exceptions import InvalidFileException
from xlrd import XLRDError


def clean_directory(path):
    """清空文件夹"""
    if not os.path.exists(path):
        return
    for i in os.listdir(path):
        path_file = os.path.join(path, i)
        if os.path.isfile(path_file):
            os.remove(path_file)
        else:
            for f in os.listdir(path_file):
                path_file2 = os.path.join(path_file, f)
            if os.path.isfile(path_file2):
                os.remove(path_file2)

def create_excel_xlsx(path, sheet_name):
    """创建excel"""
    workbook = op.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    workbook.save(path)

def write_excel_xlsx_append(path,name, value,sheet_name = 'Sheet1', truncate_sheet=False):
    """
    :param path: ./demo/sycm.xlsx
    :param value: [[],[]]
    :param truncate_sheet:truncate_sheet为True，覆盖原表中的数据
    :return:
    """
    # 如果不存在就创建该excel
    if not os.path.exists(f"{path}"):
        os.makedirs(path)
    if not os.path.exists(f"{path}/{name}", ):
        create_excel_xlsx(f"{path}/{name}", sheet_name)
    data = op.load_workbook(f"{path}/{name}", )
    # 取第一张表
    sheetnames = data.sheetnames
    sheet = data[sheetnames[0]]
    sheet = data.active
    if(truncate_sheet): #truncate_sheet为True，覆盖原表中的数据
        startrows = 0
    else:
        # print(sheet.title)  # 输出表名
        startrows = sheet.max_row  # 获得行数
    index = len(value)
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=startrows + i + 1, column=j + 1, value=str(value[i][j]))
    data.save(f"{path}/{name}",)
    print("xlsx格式表格追加写入数据成功！")

def read_excel_xlsx(path, sheet_name):
    workbook = op.load_workbook(path)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


def read_excel(path: str, sheet: str = None, encoding="utf-8", del_file=False):
    """
    读取文件
    :param path:  文件路径
    :param sheet:  有sheet页的需要指定sheet名字,不指定默认第一个sheet
    :param encoding:  编码格式
    :param del_file:  读取完成后是否删除文件，默认不删除
    :return:
    """
    if path[-4:] == ".csv":
        with open(path, encoding=encoding) as f:
            reader = csv.reader(f)
            lr = tuple([tuple(item) for item in reader])
    elif path[-5:] in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            workbook = op.load_workbook(path)
            if not sheet:
                sheet = workbook.sheetnames[0]
            sheet = workbook[sheet]  # 获取表单
            max_row = sheet.max_row  # 获取最大行数
            max_column = sheet.max_column  # 获取最大列数
            lr = tuple([tuple([sheet.cell(row=row, column=column).value for column in range(1, max_column+1)]) for row in range(1, max_row+1)])
    elif path[-4:] == ".xls":
            workbook = xlrd.open_workbook(path)  # 打开工作簿
            if not sheet:
                sheet = workbook.sheet_names()[0]
            sheet = workbook.sheet_by_name(sheet_name=sheet)  # 获取表单
            max_row = sheet.nrows  # 获取最大行
            max_column = sheet.ncols  # 获取最大列数
            lr = tuple([tuple([sheet.cell_value(rowx=row, colx=column) for column in range(0, max_column)]) for row in range(0, max_row)])
    else:
        traceback.print_exc()
        assert False, "当前文件格式不支持，支持格式：.xlsx,.xlsm,.xltx,.xltm .xls .csv"
    if del_file:
        os.remove(path)
    return lr



if __name__ == '__main__':
    print(read_excel(r"E:\chenfan_work\gitlib\工具类\util\xzg_util\XZGUtil\270977983_创意_上新快_2022-11-10.xls"))
